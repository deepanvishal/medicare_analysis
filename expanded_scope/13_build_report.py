"""
13 - Build the combined multi-state compliance workbook.   [PYTHON / pandas + openpyxl]

Modeled on FL Step4_tab1_tab2.py, multi-state (FL+OH+AZ+IL) in ONE workbook with a
State column so every tab filters. Tab order matches Step4:
  1 Project Overview      2 County Mapping        3 County Type Validation
  4 Compliance Report     5 Summary by Specialty  6 Summary by County
  7 Data Dictionary       8 CMS Rules             9 Methodology
  10 W3 Data Inventory    11 W3 Par Flags
DEFERRED (submarket is an Aetna FL grouping -- likely undefined for OH/AZ/IL):
  the 5 Submarket tabs. Added later once submarket is defined for the new states.

INPUT : ms_fact_gap_analysis, ms_ref_county, ms_ref_time_distance,
        ms_week3_data_inventory, ms_provider_par_flag, ms_stg_providers_multi_specialty,
        + bigquery-public-data (counties, county_2020_5yr)
OUTPUT: medicare_supply_demand_ms.xlsx  (repo root)
Run   : python expanded_scope/13_build_report.py     (needs: pip install db-dtypes openpyxl)
"""

import datetime
import pandas as pd
import config as cfg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

OUT_XLSX = cfg.repo_path("medicare_supply_demand_ms.xlsx")
STATE_NAMES = {"FL": "Florida", "OH": "Ohio", "AZ": "Arizona", "IL": "Illinois"}

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"
LIGHT_GREEN, LIGHT_RED = "E2EFDA", "FFE0E0"


# ---------- styling helpers (from Step4) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def section_header(ws, row, c0, c1, text, bg=MID_BLUE):
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell(ws, f"{get_column_letter(c0)}{row}", text, bold=True, color=WHITE, bg=bg, size=11)
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def title(ws, text, sub=None, ncols=8):
    span = f"A1:{get_column_letter(ncols)}1"
    ws.merge_cells(span)
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        cell(ws, "A2", sub, italic=True, color=DARK_BLUE, size=10, bg=LIGHT_BLUE)
        ws.row_dimensions[2].height = 18


def _int(v, d=0):
    try:
        return d if v is None or pd.isna(v) else int(float(v))
    except (TypeError, ValueError):
        return d


def _float(v, d=0.0):
    try:
        return d if v is None or pd.isna(v) else float(v)
    except (TypeError, ValueError):
        return d


def scope_line():
    return " + ".join(f"{STATE_NAMES[s]} ({s})" for s in cfg.STATE_ABBRS)


# ---------- generic formatted data table ----------
def simple_table(wb, sheet, df, cols, subtitle, status_key=None,
                 good=("MATCH", "COMPLIANT"), bad=("MISMATCH", "NON-COMPLIANT")):
    """cols = list of (df_key, header, width, number_format, align)."""
    ws = wb.create_sheet(sheet)
    n = len(cols)
    title(ws, sheet.split(". ", 1)[-1], subtitle, ncols=n)
    for i, (_, hdr, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}3", hdr, bold=True, color=WHITE, bg=DARK_BLUE, size=9, h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[3].height = 26
    for ridx, (_, row) in enumerate(df.iterrows(), start=4):
        bg = GREY if ridx % 2 == 0 else WHITE
        if status_key is not None:
            sv = row.get(status_key)
            if sv in good:
                bg = LIGHT_GREEN
            elif sv in bad:
                bg = LIGHT_RED
        for i, (key, _, _, num, align) in enumerate(cols):
            v = row.get(key)
            if num and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i+1)}{ridx}", v, bg=bg, size=9, bdr=True, num=num, h_align=align)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(df)+3}"
    return ws


# ---------- data ----------
def load():
    c = cfg.client()
    FACT, CTY, TD = cfg.table("fact_gap_analysis"), cfg.table("ref_county"), cfg.table("ref_time_distance")
    PROV, INV, PAR = cfg.table("stg_providers_multi_specialty"), cfg.table("week3_data_inventory"), cfg.table("provider_par_flag")
    COUNTIES = "bigquery-public-data.geo_us_boundaries.counties"
    ACS = "bigquery-public-data.census_bureau_acs.county_2020_5yr"
    fips = cfg.state_fips_sql()

    d = {}
    d["fact"] = c.query(f"SELECT * FROM `{FACT}` ORDER BY state_cd, county_name, cms_specialty, plan_type").to_dataframe()
    d["td"] = c.query(
        f"SELECT cms_specialty, county_type, MIN(max_time_min) AS base_time, MIN(max_distance_miles) AS base_dist "
        f"FROM `{TD}` GROUP BY cms_specialty, county_type").to_dataframe()
    d["county_map"] = c.query(f"""
        SELECT rc.state_cd, rc.county_name AS hsd_name, cen.county_name AS census_name, a.aetna_names
        FROM `{CTY}` rc
        LEFT JOIN `{COUNTIES}` cen ON rc.county_fips = cen.geo_id
        LEFT JOIN (SELECT county_fips, STRING_AGG(DISTINCT aetna_county_nm, ', ' ORDER BY aetna_county_nm) AS aetna_names
                   FROM `{PROV}` GROUP BY county_fips) a ON rc.county_fips = a.county_fips
        ORDER BY rc.state_cd, rc.county_name""").to_dataframe()
    d["county_val"] = c.query(f"""
        WITH raw AS (SELECT geo_id AS county_fips, area_land_meters/2589988.11 AS area_sq_miles
                     FROM `{COUNTIES}` WHERE state_fips_code IN {fips}),
        pop AS (SELECT geo_id AS county_fips, total_pop FROM `{ACS}` WHERE LEFT(geo_id,2) IN {fips}),
        j AS (SELECT r.county_fips, r.area_sq_miles, p.total_pop AS population,
                     ROUND(p.total_pop/NULLIF(r.area_sq_miles,0),2) AS pop_density
              FROM raw r LEFT JOIN pop p USING (county_fips)),
        cl AS (SELECT *, CASE
                 WHEN (population>=1000000 AND pop_density>=1000) OR (population>=500000 AND pop_density>=1500) OR (pop_density>=5000) THEN 'Large Metro'
                 WHEN (population>=1000000 AND pop_density>=10) OR (population>=500000 AND pop_density>=10) OR (population>=200000 AND pop_density>=10)
                   OR (population>=50000 AND pop_density>=100) OR (population>=10000 AND pop_density>=1000) THEN 'Metro'
                 WHEN (population>=50000 AND pop_density>=10) OR (population>=10000 AND pop_density>=50) THEN 'Micro'
                 WHEN pop_density<10 THEN 'CEAC'
                 ELSE 'Rural' END AS census_derived_type
               FROM j)
        SELECT rc.state_cd, rc.county_name, cl.population, ROUND(cl.area_sq_miles,2) AS area_sq_miles, cl.pop_density,
               cl.census_derived_type, rc.county_type AS hsd_official_type,
               IF(cl.census_derived_type = rc.county_type, 'MATCH', 'MISMATCH') AS status
        FROM cl JOIN `{CTY}` rc ON cl.county_fips = rc.county_fips
        ORDER BY status DESC, rc.state_cd, rc.county_name""").to_dataframe()
    d["w3_inv"] = c.query(
        f"SELECT state_cd, cms_specialty, plan_type, county_name, ma_contracted_providers, "
        f"aetna_participating_providers, cms_medicare_providers FROM `{INV}` "
        f"ORDER BY state_cd, county_name, cms_specialty, plan_type").to_dataframe()
    d["par"] = c.query(f"""
        SELECT state_cd, county_name, cms_specialty, plan_type,
          COUNT(DISTINCT provider_id) AS contracted_total,
          COUNT(DISTINCT CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN provider_id END) AS aetna_par,
          COUNT(DISTINCT CASE WHEN participation_status IN ('ACTIVE BOTH','CONTRACTED NOT ACTIVE - IN ORIGINAL MEDICARE') THEN provider_id END) AS cms_medicare,
          COUNT(DISTINCT CASE WHEN participation_status = 'ACTIVE BOTH' THEN provider_id END) AS both_par,
          SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN COALESCE(claim_count,0) ELSE 0 END) AS aetna_total_claims,
          ROUND(SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN COALESCE(total_allowed_amt,0) ELSE 0 END),2) AS aetna_total_allowed,
          SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','CONTRACTED NOT ACTIVE - IN ORIGINAL MEDICARE') THEN COALESCE(tot_benes,0) ELSE 0 END) AS cms_total_benes
        FROM `{PAR}` GROUP BY state_cd, county_name, cms_specialty, plan_type
        ORDER BY state_cd, county_name, cms_specialty, plan_type""").to_dataframe()

    # submarket crosswalk: county (fips->HSD name) -> submarket, joined by state_cd+county_name
    SM = (f"SELECT DISTINCT rc.state_cd, rc.county_name, p.submarket "
          f"FROM `{PROV}` p JOIN `{CTY}` rc ON p.county_fips = rc.county_fips WHERE p.submarket IS NOT NULL")
    d["sub_compliance"] = c.query(f"""
        SELECT sm.state_cd, sm.submarket, f.county_type, f.cms_specialty, f.plan_type,
          SUM(COALESCE(f.county_total_beneficiaries,0)) AS county_total_beneficiaries,
          SUM(COALESCE(f.required_provider_count,0)) AS required_provider_count,
          SUM(COALESCE(f.actual_count,0)) AS actual_count,
          SAFE_DIVIDE(SUM(COALESCE(f.population_with_access,0)), NULLIF(SUM(COALESCE(f.total_county_population,0)),0)) AS pct_covered,
          COUNTIF(f.compliance_status='COMPLIANT') AS compliant_counties,
          COUNTIF(f.compliance_status='NON-COMPLIANT') AS non_compliant_counties,
          COUNT(DISTINCT f.county_name) AS total_counties,
          CASE WHEN COUNTIF(f.compliance_status='NON-COMPLIANT')=0 THEN 'ALL COMPLIANT'
               WHEN COUNTIF(f.compliance_status='COMPLIANT')=0 THEN 'ALL NON-COMPLIANT' ELSE 'MIXED' END AS submarket_status
        FROM `{FACT}` f JOIN ({SM}) sm ON f.state_cd=sm.state_cd AND f.county_name=sm.county_name
        GROUP BY sm.state_cd, sm.submarket, f.county_type, f.cms_specialty, f.plan_type
        ORDER BY submarket_status DESC, sm.state_cd, sm.submarket, f.cms_specialty, f.plan_type""").to_dataframe()
    d["sub_summary"] = c.query(f"""
        SELECT sm.state_cd, sm.submarket, f.plan_type,
          COUNT(DISTINCT f.county_name) AS total_counties,
          COUNTIF(f.compliance_status='COMPLIANT') AS compliant_specialty_counties,
          COUNTIF(f.compliance_status='NON-COMPLIANT') AS non_compliant_specialty_counties,
          COUNT(*) AS total_specialty_counties,
          SAFE_DIVIDE(COUNTIF(f.compliance_status='COMPLIANT'), COUNT(*)) AS pct_compliant,
          COUNTIF(f.access_compliant=FALSE) AS access_failures,
          COUNTIF(f.count_compliant=FALSE) AS count_failures
        FROM `{FACT}` f JOIN ({SM}) sm ON f.state_cd=sm.state_cd AND f.county_name=sm.county_name
        GROUP BY sm.state_cd, sm.submarket, f.plan_type
        ORDER BY pct_compliant ASC, sm.state_cd, sm.submarket, f.plan_type""").to_dataframe()
    d["sub_inventory"] = c.query(f"""
        SELECT sm.state_cd, sm.submarket, i.cms_specialty, i.plan_type,
          SUM(COALESCE(i.cms_medicare_providers,0)) AS cms_available,
          SUM(COALESCE(i.ma_contracted_providers,0)) AS aetna_contracted,
          SUM(COALESCE(i.aetna_participating_providers,0)) AS aetna_active,
          COUNT(DISTINCT i.county_name) AS county_count
        FROM `{INV}` i JOIN ({SM}) sm ON i.state_cd=sm.state_cd AND i.county_name=sm.county_name
        GROUP BY sm.state_cd, sm.submarket, i.cms_specialty, i.plan_type
        ORDER BY sm.state_cd, sm.submarket, i.cms_specialty, i.plan_type""").to_dataframe()
    d["sub_par"] = c.query(f"""
        SELECT sm.state_cd, sm.submarket, p.cms_specialty, p.plan_type,
          COUNT(DISTINCT p.county_name) AS county_count,
          SUM(p.contracted_total) AS contracted_total, SUM(p.aetna_par) AS aetna_par,
          SUM(p.cms_medicare) AS cms_medicare, SUM(p.both_par) AS both_par,
          SUM(p.aetna_total_claims) AS aetna_total_claims, ROUND(SUM(p.aetna_total_allowed),2) AS aetna_total_allowed,
          SUM(p.cms_total_benes) AS cms_total_benes
        FROM (
          SELECT state_cd, county_name, cms_specialty, plan_type,
            COUNT(DISTINCT provider_id) AS contracted_total,
            COUNT(DISTINCT CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN provider_id END) AS aetna_par,
            COUNT(DISTINCT CASE WHEN participation_status IN ('ACTIVE BOTH','CONTRACTED NOT ACTIVE - IN ORIGINAL MEDICARE') THEN provider_id END) AS cms_medicare,
            COUNT(DISTINCT CASE WHEN participation_status='ACTIVE BOTH' THEN provider_id END) AS both_par,
            SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN COALESCE(claim_count,0) ELSE 0 END) AS aetna_total_claims,
            SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','AETNA ACTIVE - NO NPI MATCH','AETNA ACTIVE - NOT IN ORIGINAL MEDICARE') THEN COALESCE(total_allowed_amt,0) ELSE 0 END) AS aetna_total_allowed,
            SUM(CASE WHEN participation_status IN ('ACTIVE BOTH','CONTRACTED NOT ACTIVE - IN ORIGINAL MEDICARE') THEN COALESCE(tot_benes,0) ELSE 0 END) AS cms_total_benes
          FROM `{PAR}` GROUP BY state_cd, county_name, cms_specialty, plan_type
        ) p JOIN ({SM}) sm ON p.state_cd=sm.state_cd AND p.county_name=sm.county_name
        GROUP BY sm.state_cd, sm.submarket, p.cms_specialty, p.plan_type
        ORDER BY sm.state_cd, sm.submarket, p.cms_specialty, p.plan_type""").to_dataframe()
    d["sub_opp"] = c.query(f"""
        SELECT sm.state_cd, sm.submarket, i.cms_specialty, i.plan_type,
          SUM(COALESCE(i.cms_medicare_providers,0)) - SUM(COALESCE(i.ma_contracted_providers,0)) AS network_gap,
          SUM(COALESCE(i.cms_medicare_providers,0)) AS cms_available,
          SUM(COALESCE(i.ma_contracted_providers,0)) AS aetna_contracted,
          SUM(COALESCE(f.required_provider_count,0)) AS required_count,
          SUM(COALESCE(f.actual_count,0)) AS actual_count,
          COUNT(DISTINCT i.county_name) AS county_count
        FROM `{INV}` i
        LEFT JOIN `{FACT}` f ON i.state_cd=f.state_cd AND i.county_name=f.county_name AND i.cms_specialty=f.cms_specialty AND i.plan_type=f.plan_type
        JOIN ({SM}) sm ON i.state_cd=sm.state_cd AND i.county_name=sm.county_name
        WHERE COALESCE(i.cms_medicare_providers,0) > 0
        GROUP BY sm.state_cd, sm.submarket, i.cms_specialty, i.plan_type
        ORDER BY network_gap DESC, sm.state_cd, sm.submarket, i.cms_specialty""").to_dataframe()
    return d


# ---------- tab 1: Overview ----------
def build_overview(wb, df):
    ws = wb.create_sheet("1. Project Overview")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Medicare Supply Demand",
          f"Network Adequacy & Capacity Modeling  |  {scope_line()}  |  Plan Year 2026")
    r = 4
    r = section_header(ws, r, 2, 8, "OBJECTIVE")
    r = kv(ws, r, "Objective",
           "Determine whether the Aetna MA network has the right capacity, specialties, and "
           "geographic distribution across the scope states, and identify where to add, remove, "
           "or reconfigure providers per CMS 42 CFR 422.116.", h=44)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "SCOPE")
    per_state = df.groupby("state_cd")["county_fips"].nunique().to_dict()
    r = kv(ws, r, "Geography",
           f"{scope_line()} — {df['county_fips'].nunique()} member counties "
           "(" + ", ".join(f"{s} {per_state.get(s,0)}" for s in cfg.STATE_ABBRS) + ").")
    r = kv(ws, r, "Plan Types", ", ".join(sorted(df["plan_type"].unique())))
    r = kv(ws, r, "CMS Specialties", f"{df['cms_specialty'].nunique()} provider + facility types per 42 CFR 422.116")
    r = kv(ws, r, "Regulatory File", "CMS 2026 HSD Reference File (published December 17, 2025)")
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "COMPLIANCE RESULT")
    total, comp = len(df), int((df["compliance_status"] == "COMPLIANT").sum())
    r = kv(ws, r, "Rows (county x specialty x plan)", f"{total:,}")
    r = kv(ws, r, "COMPLIANT", f"{comp:,}  ({100*comp/total:.1f}%)" if total else "0")
    r = kv(ws, r, "NON-COMPLIANT", f"{total-comp:,}  ({100*(total-comp)/total:.1f}%)" if total else "0")
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "KEY ASSUMPTIONS")
    for lab, val in [
        ("Required Count", "From the CMS 2026 HSD Reference File (95th-percentile MA enrollment)."),
        ("Access Threshold", "90% Large Metro/Metro; 85% Micro/Rural/CEAC (422.116(d)(4))."),
        ("Distance", "County-specific max time/distance from the HSD Time & Distance tabs; straight-line."),
        ("Provider County", "Derived from the provider's practice-location zip (not the registered county)."),
        ("Telehealth Credit", "NOT applied (no telehealth flag available)."),
    ]:
        r = kv(ws, r, lab, val, h=28)
    ws.freeze_panes = "A3"


# ---------- tab 4: Compliance Report (core) ----------
COMPLIANCE_COLS = [
    ("state_cd", "State", 8, None, "left"), ("county_name", "County", 18, None, "left"),
    ("county_type", "County Type", 13, None, "left"), ("cms_specialty", "CMS Specialty", 26, None, "left"),
    ("plan_type", "Plan Type", 10, None, "left"),
    ("county_total_beneficiaries", "Total MA Benes", 15, "#,##0", "right"),
    ("ratio_95th_percentile", "95th Pct Ratio", 12, "0.0000", "right"),
    ("beneficiaries_required_to_cover", "Benes Required", 13, "#,##0", "right"),
    ("min_ratio_per_1000", "Min Ratio /1k", 11, "0.00", "right"),
    ("required_provider_count", "Required Count", 12, "#,##0", "right"),
    ("compliance_threshold", "Access Thresh", 11, "0%", "right"),
    ("max_distance_miles", "Max Dist (mi)", 11, "#,##0", "right"),
    ("total_county_population", "County Pop", 14, "#,##0", "right"),
    ("population_with_access", "Pop w/ Access", 13, "#,##0", "right"),
    ("pct_covered", "% w/ Access", 12, "0.0%", "right"),
    ("actual_count", "Contracted", 12, "#,##0", "right"),
    ("total_contracted_beds", "Beds", 10, "#,##0", "right"),
    ("provider_gap", "Gap", 9, "#,##0", "right"),
    ("access_compliant", "Access Met", 10, None, "center"),
    ("count_compliant", "Count Met", 10, None, "center"),
    ("compliance_status", "Compliance Status", 16, None, "left"),
]
COMPLIANCE_BANDS = [(1, 5, "IDENTIFIERS", DARK_GREY), (6, 12, "CMS RULES", MID_BLUE),
                    (13, 17, "POPULATION & NETWORK", "BF8F00"), (18, 21, "COMPLIANCE", DARK_BLUE)]


def build_compliance(wb, df):
    ws = wb.create_sheet("4. Compliance Report")
    n = len(COMPLIANCE_COLS)
    title(ws, "Compliance Report",
          f"County x Specialty x Plan  |  {scope_line()}  |  both tests must pass", ncols=n)
    for lo, hi, label, bg in COMPLIANCE_BANDS:
        ws.merge_cells(f"{get_column_letter(lo)}3:{get_column_letter(hi)}3")
        cell(ws, f"{get_column_letter(lo)}3", label, bold=True, color=WHITE, bg=bg, size=9, h_align="center")
    for i, (_, hdr, w, _, _) in enumerate(COMPLIANCE_COLS):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}4", hdr, bold=True, color=WHITE, bg=DARK_BLUE, size=9, h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[4].height = 30
    for ridx, (_, row) in enumerate(df.iterrows(), start=5):
        rbg = LIGHT_GREEN if row["compliance_status"] == "COMPLIANT" else LIGHT_RED
        for i, (key, _, _, num, align) in enumerate(COMPLIANCE_COLS):
            v = row.get(key)
            if key in ("access_compliant", "count_compliant"):
                v = "Yes" if bool(v) else "No"
            elif num and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i+1)}{ridx}", v, bg=rbg, size=9, bdr=True, num=num, h_align=align)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(n)}{len(df)+4}"


# ---------- tabs 5/6: summaries ----------
def _summary(wb, sheet, df, group_cols, headers, first_w):
    ws = wb.create_sheet(sheet)
    grp = (df.groupby(group_cols)
             .agg(compliant=("compliance_status", lambda s: (s == "COMPLIANT").sum()),
                  non_compliant=("compliance_status", lambda s: (s == "NON-COMPLIANT").sum()),
                  total=("compliance_status", "size"),
                  access_fail=("access_compliant", lambda s: (~s.astype(bool)).sum()),
                  count_fail=("count_compliant", lambda s: (~s.astype(bool)).sum()))
             .reset_index())
    grp["pct_compliant"] = (grp["compliant"] / grp["total"]).round(4)
    grp = grp.sort_values(["pct_compliant"] + group_cols)
    n = len(headers)
    title(ws, sheet.split(". ", 1)[-1], f"{scope_line()}  |  sorted by % compliant ascending", ncols=n)
    for i, h in enumerate(headers):
        cell(ws, f"{get_column_letter(i+1)}3", h, bold=True, color=WHITE, bg=DARK_BLUE, size=9, h_align="center", bdr=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = first_w if i == 0 else (26 if i == 1 else 14)
    ws.row_dimensions[3].height = 28
    out = group_cols + ["compliant", "non_compliant", "total", "pct_compliant", "access_fail", "count_fail"]
    for ridx, (_, row) in enumerate(grp.iterrows(), start=4):
        for i, key in enumerate(out):
            num = "0.0%" if key == "pct_compliant" else ("#,##0" if i >= len(group_cols) else None)
            v = _float(row[key]) if key == "pct_compliant" else (_int(row[key]) if i >= len(group_cols) else row[key])
            cell(ws, f"{get_column_letter(i+1)}{ridx}", v, size=9, bdr=True, num=num,
                 h_align="left" if i < len(group_cols) else "right")
    pc = get_column_letter(out.index("pct_compliant") + 1)
    ws.conditional_formatting.add(f"{pc}4:{pc}{len(grp)+3}",
        ColorScaleRule(start_type="num", start_value=0, start_color="C00000",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="375623"))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(grp)+3}"


def build_summary_specialty(wb, df):
    _summary(wb, "5. Summary by Specialty", df, ["state_cd", "cms_specialty", "plan_type"],
             ["State", "CMS Specialty", "Plan Type", "Compliant", "Non-Compliant",
              "Total Counties", "% Compliant", "Access Fails", "Count Fails"], 8)


def build_summary_county(wb, df):
    _summary(wb, "6. Summary by County", df, ["state_cd", "county_name", "plan_type"],
             ["State", "County", "Plan Type", "Compliant", "Non-Compliant",
              "Total Specialties", "% Compliant", "Access Fails", "Count Fails"], 8)


# ---------- tab 8: CMS Rules ----------
def build_cms_rules(wb, td):
    ws = wb.create_sheet("8. CMS Rules")
    title(ws, "CMS Rules — 42 CFR 422.116 Time & Distance",
          "Base standard per specialty x county_type (min time / max miles). Counties may be relaxed.", ncols=6)
    types = ["Large Metro", "Metro", "Micro", "Rural", "CEAC"]
    for i, h in enumerate(["CMS Specialty"] + types):
        cell(ws, f"{get_column_letter(i+1)}3", h, bold=True, color=WHITE, bg=MID_BLUE, size=10, h_align="center", bdr=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = 28 if i == 0 else 14
    dist = td.pivot_table(index="cms_specialty", columns="county_type", values="base_dist", aggfunc="min")
    tim = td.pivot_table(index="cms_specialty", columns="county_type", values="base_time", aggfunc="min")
    for ridx, spec in enumerate(sorted(dist.index), start=4):
        bg = GREY if ridx % 2 == 0 else WHITE
        cell(ws, f"A{ridx}", spec, bg=bg, size=9, bdr=True)
        for j, ct in enumerate(types, start=2):
            d, t = dist.loc[spec].get(ct), tim.loc[spec].get(ct)
            cell(ws, f"{get_column_letter(j)}{ridx}", "—" if pd.isna(d) else f"{_int(t)} min / {_int(d)} mi",
                 bg=bg, size=9, bdr=True, h_align="center")
    ws.freeze_panes = "A4"


# ---------- tab 7: Data Dictionary ----------
DICT_ROWS = [
    ("State / County / County Type", "ms_ref_county", "Scope state, county, and CMS county designation."),
    ("CMS Specialty / Plan Type", "ms_fact_gap_analysis", "One of 43 CMS specialties; MA-HMO or MA-PPO (evaluated separately)."),
    ("Required Count", "ms_ref_hsd_required_counts", "CMS-precalculated minimum providers (beds for Acute Inpatient)."),
    ("Access Threshold", "ms_ref_county", "0.90 (Large Metro/Metro) or 0.85 (Micro/Rural/CEAC)."),
    ("Max Distance (mi)", "ms_ref_time_distance", "County-specific max distance (member county) per 422.116."),
    ("% Pop w/ Access", "ms_fact_gap_analysis", "population_with_access / total_county_population."),
    ("Contracted (actual)", "ms_fact_gap_analysis", "COUNT(DISTINCT provider) within distance; SUM(beds) for Acute Inpatient."),
    ("Gap", "ms_fact_gap_analysis", "required_count - actual (negative = surplus)."),
    ("Access / Count Met", "ms_fact_gap_analysis", "Test 1 (pct >= threshold) / Test 2 (actual >= required)."),
    ("Compliance Status", "ms_fact_gap_analysis", "COMPLIANT iff both tests pass, else NON-COMPLIANT."),
]


def build_data_dictionary(wb):
    dd = pd.DataFrame(DICT_ROWS, columns=["Column", "Source Table", "Description"])
    simple_table(wb, "7. Data Dictionary", dd,
                 [("Column", "Column", 32, None, "left"), ("Source Table", "Source Table", 26, None, "left"),
                  ("Description", "Description", 72, None, "left")],
                 "Definitions for the Compliance Report columns")


# ---------- tab 9: Methodology ----------
METHOD = [
    ("SUPPLY SIDE", [
        ("Providers", "ms_stg_providers_multi_specialty -- provider x specialty x plan, exploded from the RPDB network; "
                      "county_fips from the practice-location zip."),
        ("Supply source", "ms_mbr_with_all_zips (mbr_with_zip + mdcr_base_provider_mdcr_ntwk), scope states only."),
        ("Participation", "ms_provider_par_flag -- Aetna claims (2024-2025) + CMS Original Medicare via NPI crosswalk."),
    ]),
    ("DEMAND SIDE", [
        ("Beneficiaries", "ms_stg_beneficiaries -- zip population (ACS 2018) + Medicare eligibles (CMS penetration)."),
        ("Required counts", "ms_ref_hsd_required_counts -- long form of the CMS 2026 HSD file, all scope states."),
    ]),
    ("GEOGRAPHY", [
        ("Zip -> county", "ms_ref_zip_reference -- spatial intersection (largest overlap); border zips kept."),
        ("Time & distance", "ms_ref_time_distance -- PER-COUNTY values from the HSD Time & Distance tabs."),
    ]),
    ("KEY DECISIONS", [
        ("Distance test", "Straight-line ST_DISTANCE between zip centroids; CMS uses drive time (rural understated)."),
        ("Provider county", "Practice-location zip (differs from registered county for multi-location providers)."),
        ("Acute Inpatient", "actual = SUM(hosp_list_cmi beds); a state with 0 beds means hosp_list_cmi lacks it."),
        ("Telehealth", "Not credited (no telehealth flag)."),
    ]),
]


def build_methodology(wb):
    ws = wb.create_sheet("9. Methodology")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Methodology & Data Sourcing", f"{scope_line()}  |  multi-state pipeline (expanded_scope)")
    r = 4
    for sec, items in METHOD:
        r = section_header(ws, r, 2, 8, sec)
        for lab, val in items:
            r = kv(ws, r, lab, val, h=30)
        r = blank(ws, r)
    ws.freeze_panes = "A3"


# ---------- assemble ----------
def build(d):
    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb, d["fact"])
    simple_table(wb, "2. County Mapping", d["county_map"],
                 [("state_cd", "State", 8, None, "left"), ("hsd_name", "HSD / CMS Name", 22, None, "left"),
                  ("census_name", "Census Name", 22, None, "left"), ("aetna_names", "Aetna County Name(s)", 40, None, "left")],
                 f"{scope_line()}  |  county names across sources (joined by FIPS)")
    simple_table(wb, "3. County Type Validation", d["county_val"],
                 [("state_cd", "State", 8, None, "left"), ("county_name", "County", 20, None, "left"),
                  ("population", "Population", 14, "#,##0", "right"), ("area_sq_miles", "Area (sq mi)", 12, "#,##0.0", "right"),
                  ("pop_density", "Density", 11, "#,##0.0", "right"), ("census_derived_type", "Census-Derived", 15, None, "left"),
                  ("hsd_official_type", "HSD Official", 15, None, "left"), ("status", "Status", 11, None, "center")],
                 f"{scope_line()}  |  Census-derived vs CMS HSD county_type", status_key="status")
    build_compliance(wb, d["fact"])
    build_summary_specialty(wb, d["fact"])
    build_summary_county(wb, d["fact"])
    build_data_dictionary(wb)
    build_cms_rules(wb, d["td"])
    build_methodology(wb)
    simple_table(wb, "10. W3 Data Inventory", d["w3_inv"],
                 [("state_cd", "State", 8, None, "left"), ("cms_specialty", "CMS Specialty", 26, None, "left"),
                  ("plan_type", "Plan Type", 10, None, "left"), ("county_name", "County", 18, None, "left"),
                  ("ma_contracted_providers", "MA Contracted", 13, "#,##0", "right"),
                  ("aetna_participating_providers", "Aetna Participating", 16, "#,##0", "right"),
                  ("cms_medicare_providers", "CMS Medicare", 13, "#,##0", "right")],
                 f"{scope_line()}  |  county-level counts (do NOT sum across counties)")
    simple_table(wb, "11. W3 Par Flags", d["par"],
                 [("state_cd", "State", 8, None, "left"), ("county_name", "County", 18, None, "left"),
                  ("cms_specialty", "CMS Specialty", 24, None, "left"), ("plan_type", "Plan Type", 10, None, "left"),
                  ("contracted_total", "Contracted", 12, "#,##0", "right"), ("aetna_par", "Aetna Par", 11, "#,##0", "right"),
                  ("cms_medicare", "CMS Medicare", 12, "#,##0", "right"), ("both_par", "Both", 9, "#,##0", "right"),
                  ("aetna_total_claims", "Aetna Claims", 12, "#,##0", "right"),
                  ("aetna_total_allowed", "Aetna Allowed $", 14, "#,##0.00", "right"),
                  ("cms_total_benes", "CMS Benes", 12, "#,##0", "right")],
                 f"{scope_line()}  |  participation summary (county-level, do NOT sum)")

    _SM_SUB = "Submarket is an Aetna internal business grouping -- NOT a CMS compliance unit."
    simple_table(wb, "12. Submarket Compliance", d["sub_compliance"],
                 [("state_cd", "State", 8, None, "left"), ("submarket", "Submarket", 18, None, "left"),
                  ("county_type", "County Type", 12, None, "left"), ("cms_specialty", "CMS Specialty", 24, None, "left"),
                  ("plan_type", "Plan Type", 10, None, "left"),
                  ("county_total_beneficiaries", "Total Benes", 13, "#,##0", "right"),
                  ("required_provider_count", "Required", 11, "#,##0", "right"),
                  ("actual_count", "Actual", 10, "#,##0", "right"), ("pct_covered", "% Covered", 11, "0.0%", "right"),
                  ("compliant_counties", "Compliant", 11, "#,##0", "right"),
                  ("non_compliant_counties", "Non-Compliant", 13, "#,##0", "right"),
                  ("total_counties", "Counties", 10, "#,##0", "right"),
                  ("submarket_status", "Status", 16, None, "left")],
                 _SM_SUB, status_key="submarket_status", good=("ALL COMPLIANT",), bad=("ALL NON-COMPLIANT",))
    simple_table(wb, "13. Submarket Summary", d["sub_summary"],
                 [("state_cd", "State", 8, None, "left"), ("submarket", "Submarket", 18, None, "left"),
                  ("plan_type", "Plan Type", 10, None, "left"), ("total_counties", "Counties", 10, "#,##0", "right"),
                  ("compliant_specialty_counties", "Compliant", 11, "#,##0", "right"),
                  ("non_compliant_specialty_counties", "Non-Compliant", 13, "#,##0", "right"),
                  ("total_specialty_counties", "Total", 9, "#,##0", "right"),
                  ("pct_compliant", "% Compliant", 12, "0.0%", "right"),
                  ("access_failures", "Access Fails", 12, "#,##0", "right"),
                  ("count_failures", "Count Fails", 11, "#,##0", "right")], _SM_SUB)
    simple_table(wb, "14. Submarket Inventory", d["sub_inventory"],
                 [("state_cd", "State", 8, None, "left"), ("submarket", "Submarket", 18, None, "left"),
                  ("cms_specialty", "CMS Specialty", 24, None, "left"), ("plan_type", "Plan Type", 10, None, "left"),
                  ("cms_available", "CMS Available", 13, "#,##0", "right"),
                  ("aetna_contracted", "Aetna Contracted", 15, "#,##0", "right"),
                  ("aetna_active", "Aetna Active", 12, "#,##0", "right"),
                  ("county_count", "Counties", 10, "#,##0", "right")], _SM_SUB)
    simple_table(wb, "15. Submarket Par Flags", d["sub_par"],
                 [("state_cd", "State", 8, None, "left"), ("submarket", "Submarket", 18, None, "left"),
                  ("cms_specialty", "CMS Specialty", 24, None, "left"), ("plan_type", "Plan Type", 10, None, "left"),
                  ("county_count", "Counties", 9, "#,##0", "right"), ("contracted_total", "Contracted", 11, "#,##0", "right"),
                  ("aetna_par", "Aetna Par", 10, "#,##0", "right"), ("cms_medicare", "CMS Medicare", 12, "#,##0", "right"),
                  ("both_par", "Both", 8, "#,##0", "right"), ("aetna_total_claims", "Aetna Claims", 12, "#,##0", "right"),
                  ("aetna_total_allowed", "Aetna Allowed $", 14, "#,##0.00", "right"),
                  ("cms_total_benes", "CMS Benes", 11, "#,##0", "right")], _SM_SUB)
    simple_table(wb, "16. Submarket Opportunity", d["sub_opp"],
                 [("state_cd", "State", 8, None, "left"), ("submarket", "Submarket", 18, None, "left"),
                  ("cms_specialty", "CMS Specialty", 24, None, "left"), ("plan_type", "Plan Type", 10, None, "left"),
                  ("network_gap", "Network Gap", 12, "#,##0", "right"), ("cms_available", "CMS Available", 13, "#,##0", "right"),
                  ("aetna_contracted", "Aetna Contracted", 15, "#,##0", "right"),
                  ("required_count", "Required", 10, "#,##0", "right"), ("actual_count", "Actual", 9, "#,##0", "right"),
                  ("county_count", "Counties", 9, "#,##0", "right")],
                 _SM_SUB + "  Network gap = CMS-available minus Aetna-contracted (recruitment target).")
    return wb


def main():
    d = load()
    if d["fact"].empty:
        raise SystemExit("ms_fact_gap_analysis returned 0 rows -- run 01-12 first.")
    wb = build(d)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}  ({datetime.datetime.now():%Y-%m-%d %H:%M})")
    print(f"  tabs: {wb.sheetnames}")
    print(f"  compliance rows: {len(d['fact'])}  | counties: {d['fact']['county_fips'].nunique()}")


if __name__ == "__main__":
    main()
