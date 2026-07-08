"""
38 - Demand vs Capacity workbook.   [PYTHON / pandas + openpyxl]   *** DEMAND/CAPACITY EXTENSION ***

Builds ONE workbook from the dc_ tables (M1-M8): demand method (CCIR chronic
conditions) and county demand, capacity method and county capacity, the gap report
with deserts and the compliant-but-strained headline, the forecast example, book
utilization, worked examples, and the data dictionary / methodology appendix.

Every worked example queries the real inputs, recomputes the output in Python, and
check()s it against the pipeline's stored value BEFORE writing the memo. If any
recomputation disagrees, the build fails — that is intended. Chosen rows come from
deterministic queries, never hardcoded county names.

INPUT : ms_dc_rate, ms_dc_county_population, ms_dc_demand, ms_dc_capacity,
        ms_dc_provider_capacity, ms_dc_gap, ms_dc_forecast_example,
        ms_dc_book_utilization, ms_dc_member_dim, ms_dc_ref_ccir
OUTPUT: medicare_demand_capacity_ms.xlsx  (repo root)
Run   : python expanded_scope/38_dc_report.py    (needs: pip install db-dtypes openpyxl)
"""

import datetime
import pandas as pd
import config as cfg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule

OUT_XLSX = cfg.repo_path("medicare_demand_capacity_ms.xlsx")

RATE   = cfg.table("dc_rate")
POP    = cfg.table("dc_county_population")
DEM    = cfg.table("dc_demand")
CAP    = cfg.table("dc_capacity")
PCAP   = cfg.table("dc_provider_capacity")
GAP    = cfg.table("dc_gap")
FC     = cfg.table("dc_forecast_example")
BOOK   = cfg.table("dc_book_utilization")
MEMDIM = cfg.table("dc_member_dim")
CCIR   = cfg.table("dc_ref_ccir")
CLAIMS = "anbc-hcb-dev.provider_ds_netconf_data_hcb_dev.A870800_medicare_analysis_2025_claims"

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"
LIGHT_GREEN, LIGHT_GOLD = "E2EFDA", "FFF2CC"
LIGHT_RED = "F8CBAD"

FACILITY_LIST = ('Acute Inpatient Hospitals', 'Outpatient Infusion/Chemo', 'Mammography',
                 'Physical Therapy', 'Occupational Therapy', 'Speech Therapy',
                 'Outpatient Dialysis', 'Skilled Nursing Facilities',
                 'Inpatient Psychiatric Facility Services', 'Outpatient Behavioral Health',
                 'Cardiac Surgery Program', 'Cardiac Catheterization Services',
                 'Diagnostic Radiology')
FACILITY_SQL = "(" + ", ".join(f"'{s}'" for s in FACILITY_LIST) + ")"

VISIT_KEY = "CONCAT(c.member_id,'|',CAST(c.srv_prvdr_id AS STRING),'|',CAST(c.srv_start_dt AS STRING))"


# ---------- styling helpers (from 13_build_report.py) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def section_header(ws, row, c0, c1, text, bg=MID_BLUE):
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell(ws, f"{get_column_letter(c0)}{row}", text, bold=True, color=WHITE, bg=bg, size=11)
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def title(ws, text, sub=None, ncols=8):
    span = f"A1:{get_column_letter(ncols)}1"
    ws.merge_cells(span)
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        cell(ws, "A2", sub, italic=True, color=DARK_BLUE, size=10, bg=LIGHT_BLUE)
        ws.row_dimensions[2].height = 18


def _int(v, d=0):
    try:
        return d if v is None or pd.isna(v) else int(float(v))
    except (TypeError, ValueError):
        return d


def _float(v, d=0.0):
    try:
        return d if v is None or pd.isna(v) else float(v)
    except (TypeError, ValueError):
        return d


# ---------- the worked-example engine ----------
WORKED_EXAMPLES = []


def memo(ws, row, heading, lines):
    WORKED_EXAMPLES.append((heading, list(lines)))
    row = section_header(ws, row, 2, 9, "How to read this - worked example: " + heading,
                         LIGHT_GOLD)
    for line in lines:
        c = ws.cell(row=row, column=2, value=line)
        c.font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        row += 1
    return row + 1


def check(label, computed, stored, tol=0.01):
    stored = float(stored)
    computed = float(computed)
    if abs(computed - stored) > tol * max(1.0, abs(stored)):
        raise ValueError(f"worked example mismatch [{label}]: computed {computed} vs stored {stored}")


# ---------- table + chart writers ----------
def write_table(ws, df, cols, r0, c0=1, filters=False, set_width=True):
    """cols = list of (df_key, header, width, number_format, align)."""
    for i, (_, hdr, w, _, _) in enumerate(cols):
        col = get_column_letter(c0 + i)
        cell(ws, f"{col}{r0}", hdr, bold=True, color=WHITE, bg=DARK_BLUE, size=9,
             h_align="center", bdr=True)
        if set_width and w:
            ws.column_dimensions[col].width = w
    ws.row_dimensions[r0].height = 24
    for ridx, (_, row) in enumerate(df.iterrows(), start=r0 + 1):
        bg = GREY if ridx % 2 == 0 else WHITE
        for i, (key, _, _, num, align) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if num and v is not None and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif v is None or pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(c0 + i)}{ridx}", v, bg=bg, size=9, bdr=True,
                 num=num, h_align=align)
    if filters:
        ws.auto_filter.ref = (f"{get_column_letter(c0)}{r0}:"
                              f"{get_column_letter(c0 + len(cols) - 1)}{r0 + len(df)}")
    return r0 + len(df) + 1


def add_bar(ws, chart_title, r_head, r_last, c_cat, c_val0, c_val1, anchor):
    ch = BarChart()
    ch.type = "col"
    ch.title = chart_title
    ch.height, ch.width = 7, 15
    data = Reference(ws, min_col=c_val0, max_col=c_val1, min_row=r_head, max_row=r_last)
    cats = Reference(ws, min_col=c_cat, max_col=c_cat, min_row=r_head + 1, max_row=r_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)


def add_line(ws, chart_title, r_head, r_last, c_cat, c_val0, c_val1, anchor):
    ch = LineChart()
    ch.title = chart_title
    ch.height, ch.width = 7, 15
    data = Reference(ws, min_col=c_val0, max_col=c_val1, min_row=r_head, max_row=r_last)
    cats = Reference(ws, min_col=c_cat, max_col=c_cat, min_row=r_head + 1, max_row=r_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)


# ---------- data ----------
def load():
    client = cfg.client()
    q = lambda sql: client.query(sql).result().to_dataframe()
    d = {}

    top = q(f"SELECT specialty_ctg_cd, ANY_VALUE(specialty_desc) AS specialty_desc "
            f"FROM `{RATE}` WHERE state_cd='FL' GROUP BY 1 "
            f"ORDER BY SUM(ma_visits) DESC LIMIT 1")
    d["topspec_cd"] = str(top["specialty_ctg_cd"].iloc[0])
    d["topspec_desc"] = str(top["specialty_desc"].iloc[0])
    TOPSPEC = d["topspec_cd"]

    # Demand Method tab (CCIR chronic conditions)
    d["ccir_overview"] = q(f"SELECT chronic_label, COUNT(*) AS icd_codes FROM `{CCIR}` "
                           f"GROUP BY 1 ORDER BY 2 DESC")
    d["top_chronic"] = q(f"""
        SELECT r.icd_code,
               ANY_VALUE(r.icd_description) AS icd_description,
               COUNT(DISTINCT {VISIT_KEY}) AS visits,
               COUNT(DISTINCT c.member_id) AS members
        FROM `{CLAIMS}` c
        JOIN `{CCIR}` r ON REPLACE(c.pri_icd9_dx_cd,'.','') = r.icd_code
        WHERE r.chronic_indicator = 1
        GROUP BY r.icd_code
        ORDER BY visits DESC
        LIMIT 20""")
    d["top_chronic_state"] = q(f"""
        WITH top20 AS (
          SELECT r.icd_code
          FROM `{CLAIMS}` c
          JOIN `{CCIR}` r ON REPLACE(c.pri_icd9_dx_cd,'.','') = r.icd_code
          WHERE r.chronic_indicator = 1
          GROUP BY r.icd_code
          ORDER BY COUNT(DISTINCT {VISIT_KEY}) DESC
          LIMIT 20
        )
        SELECT UPPER(LEFT(c.prvdr_submarket,2)) AS state_cd,
               r.icd_code,
               ANY_VALUE(r.icd_description) AS icd_description,
               COUNT(DISTINCT {VISIT_KEY}) AS visits
        FROM `{CLAIMS}` c
        JOIN `{CCIR}` r ON REPLACE(c.pri_icd9_dx_cd,'.','') = r.icd_code
        WHERE r.chronic_indicator = 1
          AND r.icd_code IN (SELECT icd_code FROM top20)
          AND UPPER(LEFT(c.prvdr_submarket,2)) IN ('FL','OH','AZ','IL')
        GROUP BY state_cd, r.icd_code
        ORDER BY r.icd_code, state_cd""")
    d["visit_freq"] = q(f"""
        SELECT m.morbidity_level,
               COUNT(DISTINCT {VISIT_KEY}) AS visits,
               COUNT(DISTINCT c.member_id) AS members
        FROM `{CLAIMS}` c
        JOIN `{MEMDIM}` m ON c.member_id = m.member_id
        WHERE m.age_band != 'UNDER_60'
        GROUP BY m.morbidity_level
        ORDER BY visits DESC""")
    d["age_morb"] = q(f"SELECT age_band, morbidity_level, COUNT(*) AS members "
                      f"FROM `{MEMDIM}` WHERE age_band != 'UNDER_60' "
                      f"GROUP BY age_band, morbidity_level ORDER BY age_band, morbidity_level")
    d["we_rate"] = q(f"SELECT ma_visits, cell_n, rate_ma_proxy FROM `{RATE}` "
                     f"WHERE state_cd='FL' AND specialty_ctg_cd='{TOPSPEC}' "
                     f"AND age_band='70-74' AND morbidity_level='CHRONIC'")

    d["dem"] = q(f"SELECT state_cd, county_fips, county_name, specialty_ctg_cd, specialty_desc, "
                 f"ROUND(total_demand_visits) AS total_demand_visits, "
                 f"ROUND(ma_demand_visits) AS ma_demand_visits "
                 f"FROM `{DEM}` ORDER BY 1, 3, 4")
    d["state"] = q(f"SELECT state_cd, ROUND(SUM(total_demand_visits)) AS total_demand, "
                   f"ROUND(SUM(ma_demand_visits)) AS ma FROM `{DEM}` GROUP BY 1 ORDER BY 1")

    we_cty = q(f"SELECT county_fips, county_name FROM `{POP}` WHERE state_cd='FL' "
               f"GROUP BY 1, 2 ORDER BY MAX(county_eligibles_total) DESC LIMIT 1")
    d["we_cty_fips"] = str(we_cty["county_fips"].iloc[0])
    d["we_cty_name"] = str(we_cty["county_name"].iloc[0])
    WE_CTY = d["we_cty_fips"]
    d["we_inputs"] = q(f"""
        SELECT p.age_band, p.county_eligibles_total, p.age_share, p.eligibles_in_band,
               mix.morbidity_level, mix.mix_share, r.rate_ma_proxy
        FROM `{POP}` p
        JOIN (SELECT state_cd, age_band, morbidity_level,
                     COUNT(*) / SUM(COUNT(*)) OVER (PARTITION BY state_cd, age_band) AS mix_share
              FROM `{MEMDIM}` WHERE age_band != 'UNDER_60' GROUP BY 1, 2, 3) mix
          ON mix.state_cd='FL' AND mix.age_band = p.age_band
        JOIN `{RATE}` r
          ON r.state_cd='FL' AND r.specialty_ctg_cd='{TOPSPEC}' AND r.age_band=p.age_band
          AND r.morbidity_level=mix.morbidity_level
        WHERE p.county_fips='{WE_CTY}'
        ORDER BY p.age_band, mix.morbidity_level""")
    d["we_dem_stored"] = q(f"SELECT total_demand_visits FROM `{DEM}` "
                           f"WHERE county_fips='{WE_CTY}' AND specialty_ctg_cd='{TOPSPEC}'")

    d["p75"] = q(f"SELECT cms_specialty, COUNT(DISTINCT provider_id) AS providers, "
                 f"ANY_VALUE(typical_annual_capacity) AS typ_cap FROM `{PCAP}` "
                 f"WHERE state_cd='FL' GROUP BY 1 ORDER BY providers DESC LIMIT 10")
    d["funnel"] = q(f"SELECT state_cd, COUNT(DISTINCT provider_id) AS contracted, "
                    f"COUNT(DISTINCT CASE WHEN active_flag=1 THEN provider_id END) AS active, "
                    f"ROUND(SUM(provider_slots)) AS slot_equivalents "
                    f"FROM `{PCAP}` GROUP BY 1 ORDER BY 1")
    d["imp"] = q(f"SELECT state_cd, ROUND(COUNTIF(saturation_imputed)/COUNT(*),3) AS imputed_share "
                 f"FROM `{PCAP}` GROUP BY 1 ORDER BY 1")
    d["we_prov"] = q(f"SELECT provider_id, cms_specialty, typical_annual_capacity, "
                     f"senior_saturation, active_flag, provider_slots FROM `{PCAP}` "
                     f"WHERE state_cd='FL' AND active_flag=1 AND saturation_imputed = FALSE "
                     f"AND senior_saturation BETWEEN 0.6 AND 0.9 ORDER BY provider_id LIMIT 1")

    d["cap"] = q(f"SELECT state_cd, county_fips, county_name, cms_specialty, plan_type, "
                 f"contracted_providers, active_providers, "
                 f"ROUND(capacity_visits) AS capacity_visits FROM `{CAP}` ORDER BY 1, 3, 4, 5")

    d["gap"] = q(f"SELECT state_cd, county_fips, county_name, county_type, cms_specialty, "
                 f"plan_type, compliance_status, required_provider_count, actual_count, "
                 f"ROUND(ma_demand_visits) AS ma_demand_visits, "
                 f"ROUND(capacity_visits) AS capacity_visits, "
                 f"ROUND(demand_capacity_gap) AS demand_capacity_gap, "
                 f"ROUND(market_opportunity_ratio,3) AS market_opportunity_ratio, "
                 f"gap_status, risk_flag FROM `{GAP}` ORDER BY 1, 3, 5, 6")
    d["we_gap"] = q(f"SELECT state_cd, county_name, cms_specialty, plan_type, ma_demand_visits, "
                    f"capacity_visits, demand_capacity_gap, market_opportunity_ratio, "
                    f"compliance_status, risk_flag FROM `{GAP}` WHERE risk_flag "
                    f"ORDER BY demand_capacity_gap DESC LIMIT 1")

    d["quad"] = q(f"SELECT state_cd, compliance_status='COMPLIANT' AS compliant, "
                  f"capacity_visits >= ma_demand_visits AS capacity_ok, COUNT(*) AS cells "
                  f"FROM `{GAP}` WHERE ma_demand_visits IS NOT NULL "
                  f"GROUP BY 1, 2, 3 ORDER BY 1, 2, 3")
    d["strain"] = q(f"SELECT state_cd, COUNTIF(risk_flag) AS compliant_but_strained, "
                    f"COUNT(*) AS mapped_cells FROM `{GAP}` WHERE ma_demand_visits IS NOT NULL "
                    f"GROUP BY 1 ORDER BY 1")
    d["desert_prac"] = q(f"SELECT state_cd, county_name, cms_specialty, plan_type, "
                         f"ROUND(demand_capacity_gap) AS gap FROM `{GAP}` "
                         f"WHERE gap_status='DESERT' AND cms_specialty NOT IN {FACILITY_SQL} "
                         f"ORDER BY demand_capacity_gap DESC LIMIT 15")
    d["desert_fac"] = q(f"SELECT state_cd, county_name, cms_specialty, plan_type, "
                        f"ROUND(demand_capacity_gap) AS gap FROM `{GAP}` "
                        f"WHERE gap_status='DESERT' AND cms_specialty IN {FACILITY_SQL} "
                        f"ORDER BY demand_capacity_gap DESC LIMIT 10")
    d["we_quad"] = q(f"SELECT COUNTIF(compliance_status='COMPLIANT' AND demand_capacity_gap > 0) "
                     f"AS strained, COUNTIF(compliance_status='COMPLIANT') AS compliant, "
                     f"COUNT(*) AS mapped FROM `{GAP}` "
                     f"WHERE state_cd='FL' AND ma_demand_visits IS NOT NULL")

    d["fc"] = q(f"SELECT state_cd, county_name, cms_specialty, plan_type, forecast_month, "
                f"seasonality_index, growth_factor, "
                f"ROUND(projected_demand_visits) AS projected_demand, "
                f"ROUND(monthly_capacity) AS monthly_capacity, crossover_flag "
                f"FROM `{FC}` ORDER BY 1, forecast_month")
    if len(d["fc"]):
        st0 = sorted(d["fc"]["state_cd"].dropna().unique())[0]
        row6 = d["fc"][(d["fc"]["state_cd"] == st0) & (d["fc"]["forecast_month"] == 6)].iloc[0]
        d["we_fc_row"] = row6
        d["we_fc_ma"] = q(f"SELECT ma_demand_visits FROM `{GAP}` "
                          f"WHERE state_cd='{st0}' AND county_name=\"{row6['county_name']}\" "
                          f"AND cms_specialty='{row6['cms_specialty']}' "
                          f"AND plan_type='{row6['plan_type']}'")

    d["book"] = q(f"SELECT state_cd, prvdr_county, specialty_ctg_cd, specialty_desc, lob, "
                  f"age_band, visits FROM `{BOOK}` ORDER BY 1, 2, 3, 5, 6")
    d["we_book_total"] = q(f"SELECT SUM(visits) AS v FROM `{BOOK}` WHERE lob='TOTAL'")
    top_cty = q(f"SELECT prvdr_county FROM `{BOOK}` WHERE state_cd='FL' AND lob='TOTAL' "
                f"GROUP BY 1 ORDER BY SUM(visits) DESC LIMIT 1")
    d["we_book_cty_name"] = str(top_cty["prvdr_county"].iloc[0])
    d["we_book_cty"] = q(f"SELECT SUM(visits) AS v FROM `{BOOK}` WHERE lob='TOTAL' "
                         f"AND state_cd='FL' AND prvdr_county=\"{d['we_book_cty_name']}\"")
    d["we_book_spec"] = q(f"SELECT SUM(visits) AS v FROM `{BOOK}` WHERE lob='TOTAL' "
                          f"AND state_cd='FL' AND prvdr_county=\"{d['we_book_cty_name']}\" "
                          f"AND specialty_ctg_cd='{TOPSPEC}'")
    return d


# ---------- tab 1: Overview ----------
def build_overview(wb):
    ws = wb.create_sheet("Overview")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Medicare Demand vs Capacity — FL OH AZ IL",
          "Demand, capacity, and gap per county x specialty x plan | Aetna MA | claims year 2025")
    r = 4
    r = section_header(ws, r, 2, 8, "THE ASK")
    r = kv(ws, r, "Directive",
           "Estimate utilization driven by Medicare demographics and morbidity; estimate "
           "effective MA capacity; surface where beneficiaries are at risk.", h=30)
    r = kv(ws, r, "What compliance does not measure",
           "CMS 42 CFR 422.116 is a floor: enough providers, close enough. It does not ask "
           "how much care the population will pull, nor whether contracted providers can "
           "absorb new members.", h=42)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "THREE NUMBERS PER COUNTY x SPECIALTY x PLAN")
    r = kv(ws, r, "Demand",
           "Expected visits: population x utilization rate, built per age band and morbidity "
           "level — total demand (population-projected) vs MA demand (Aetna members).", h=30)
    r = kv(ws, r, "Capacity",
           "Deliverable visits: per-provider typical annual volume x active flag x room left "
           "after the provider's existing senior panel, summed to the county.", h=30)
    r = kv(ws, r, "Gap",
           "ma_demand_visits minus capacity_visits; DESERT / BALANCED / OVERSUPPLY status "
           "plus the compliant-but-strained risk flag.", h=30)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "LOCKED RULES")
    r = kv(ws, r, "Visit", "One distinct member x provider x day.")
    r = kv(ws, r, "LOB", "business_ln_cd CP + ME; Medicare = ME.")
    r = kv(ws, r, "Rates", "State-pooled by age band x morbidity level.")
    r = kv(ws, r, "Morbidity",
           "morbidity_level from chronic_condition_count (AHRQ CCIR): CHRONIC if at least one "
           "chronic condition, else NON_CHRONIC.", h=28)
    r = kv(ws, r, "Demand population", "All county Medicare eligibles (CMS penetration file).")
    r = kv(ws, r, "Gap", "ma_demand minus capacity (like-for-like).")
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "READ THIS FIRST")
    r = kv(ws, r, "Estimation method",
           "Rates are built from Aetna ME claims and applied to the whole population; a "
           "total-Medicare rate source is a v2 upgrade.", h=30)
    r = kv(ws, r, "County morbidity",
           "State morbidity mix applied to every county pending the CMS county risk score load.", h=24)
    r = kv(ws, r, "Worked examples",
           "Every chart is followed by a hand calculation of one real row, recomputed live at "
           "build time; Tab 10 collects them.", h=28)
    ws.freeze_panes = "A3"
    return 0


# ---------- tab 2: Demand Method ----------
def build_demand_method(wb, d):
    ws = wb.create_sheet("Demand Method")
    for col, w in {"A": 3, "B": 22, "C": 40, "D": 30, "E": 16, "F": 16, "G": 18, "H": 18, "I": 18}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Demand Method",
          "population x rate, per age band x morbidity level (AHRQ CCIR chronic conditions)",
          ncols=9)
    r = 4

    # B0 - how to read this
    r = section_header(ws, r, 2, 8, "HOW TO READ THIS")
    r = kv(ws, r, "What this page shows",
           "How the demand estimate is built: population times a per-member visit rate, "
           "split by age_band and by morbidity_level.", h=28)
    r = kv(ws, r, "Visit definition",
           "One visit = one member seeing one provider on one day (member x provider x "
           "service date). Every visit count in this workbook uses this definition.", h=28)
    r = kv(ws, r, "Estimation method",
           "Rates are built from Aetna Medicare (ME) claims and applied to the population. "
           "total_demand_visits projects the county's eligible population; ma_demand_visits "
           "applies the same rates to Aetna members.", h=36)
    r = kv(ws, r, "Morbidity",
           "morbidity_level is set from chronic_condition_count (from AHRQ CCIR): a member "
           "with at least one chronic condition is CHRONIC, otherwise NON_CHRONIC.", h=28)
    r = blank(ws, r)

    # B1 - why HCC was not used
    r = section_header(ws, r, 2, 8, "WHY HCC WAS NOT USED FOR CHRONIC CONDITIONS")
    r = kv(ws, r, "HCC labels",
           "Of 79 payment HCC categories, only 8 contain the word 'chronic' and only 5 "
           "contain 'acute' in their label.", h=28)
    r = kv(ws, r, "The gap",
           "Conditions such as Chronic Obstructive Pulmonary Disease (HCC 111) carry neither "
           "word; whether an HCC is chronic is not determinable from its label.", h=28)
    r = kv(ws, r, "Conclusion",
           "HCC counts measure condition burden, not chronic prevalence. AHRQ CCIR classifies "
           "every diagnosis code directly, so it is used instead.", h=28)
    r = blank(ws, r)

    # B2 - CCIR overview
    r = section_header(ws, r, 2, 8, "AHRQ CCIR OVERVIEW")
    r = write_table(ws, d["ccir_overview"],
                    [("chronic_label", "Chronic Label", 22, None, "left"),
                     ("icd_codes", "ICD Codes", 12, "#,##0", "right")],
                    r, c0=2, set_width=False)
    r = kv(ws, r, "Source", "AHRQ Chronic Condition Indicator Refined (CCIR), v2026.1.")
    r = kv(ws, r, "How codes are marked",
           "Each ICD-10-CM code is flagged 0 = NOT_CHRONIC, 1 = CHRONIC, or "
           "9 = NO_DETERMINATION.", h=24)
    r = kv(ws, r, "Coverage",
           "Joined to claims on diagnosis code; matched over 99% of claim lines.")
    r = blank(ws, r)

    # B3 - most common chronic conditions (overall)
    r = section_header(ws, r, 2, 8, "MOST COMMON CHRONIC CONDITIONS")
    t0 = r
    r = write_table(ws, d["top_chronic"],
                    [("icd_code", "ICD Code", 12, None, "left"),
                     ("icd_description", "Description", 40, None, "left"),
                     ("visits", "Visits", 12, "#,##0", "right"),
                     ("members", "Members", 12, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Most common chronic conditions by visits",
            t0, t0 + len(d["top_chronic"]), 2, 4, 4, f"G{t0}")
    r = max(r, t0 + 16) + 1
    r = kv(ws, r, "Note",
           "visits counted at member x provider x service date; state here is the provider's "
           "submarket, not member home.", h=28)
    r = blank(ws, r)

    # B4 - most common chronic conditions by state
    r = section_header(ws, r, 2, 8, "MOST COMMON CHRONIC CONDITIONS BY STATE")
    r = write_table(ws, d["top_chronic_state"],
                    [("state_cd", "State", 8, None, "left"),
                     ("icd_code", "ICD Code", 12, None, "left"),
                     ("icd_description", "Description", 40, None, "left"),
                     ("visits", "Visits", 12, "#,##0", "right")],
                    r, c0=2, set_width=False)
    r = kv(ws, r, "Note", "Same top 20 conditions, split by provider submarket state.")
    r = blank(ws, r)

    # B5 - how members are flagged
    r = section_header(ws, r, 2, 8, "HOW MEMBERS ARE FLAGGED")
    r = kv(ws, r, "Rule",
           "A member with chronic_condition_count of at least 1 is flagged CHRONIC in "
           "morbidity_level; a member with 0 is NON_CHRONIC.", h=28)
    r = kv(ws, r, "chronic_condition_count",
           "The number of distinct CCIR-chronic diagnosis codes on the member's claims.", h=24)
    r = blank(ws, r)

    # B6 - chronic vs non-chronic visit frequency
    r = section_header(ws, r, 2, 8, "CHRONIC VS NON-CHRONIC: VISIT FREQUENCY")
    t0 = r
    r = write_table(ws, d["visit_freq"],
                    [("morbidity_level", "Morbidity Level", 16, None, "left"),
                     ("visits", "Visits", 14, "#,##0", "right"),
                     ("members", "Members", 14, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Visits by morbidity_level", t0, t0 + len(d["visit_freq"]), 2, 3, 3, f"F{t0}")
    r = max(r, t0 + 15) + 1
    r = kv(ws, r, "Read",
           "Compares total visits and distinct members for CHRONIC vs NON_CHRONIC.", h=24)
    r = blank(ws, r)

    # B7 - member frequency by age band and morbidity (wide table drives the grouped chart)
    r = section_header(ws, r, 2, 8, "MEMBER FREQUENCY BY AGE BAND AND MORBIDITY")
    am = d["age_morb"].pivot_table(index="age_band", columns="morbidity_level",
                                   values="members", aggfunc="sum", fill_value=0)
    for lvl in ["CHRONIC", "NON_CHRONIC"]:
        if lvl not in am.columns:
            am[lvl] = 0
    am = am[["CHRONIC", "NON_CHRONIC"]].reset_index()
    t0 = r
    r = write_table(ws, am,
                    [("age_band", "Age Band", 12, None, "left"),
                     ("CHRONIC", "CHRONIC", 14, "#,##0", "right"),
                     ("NON_CHRONIC", "NON_CHRONIC", 14, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Members by age band and morbidity_level",
            t0, t0 + len(am), 2, 3, 4, f"F{t0}")
    r = max(r, t0 + 15) + 1

    # worked example: one rate cell (recomputed live, checked before writing)
    we = d["we_rate"].iloc[0]
    ma_visits, cell_n = float(we["ma_visits"]), float(we["cell_n"])
    computed = ma_visits / cell_n
    check("one rate cell", computed, we["rate_ma_proxy"])
    r = memo(ws, r, "one rate cell", [
        f"Take FL, {d['topspec_desc']}, age 70-74, CHRONIC morbidity (at least one chronic "
        f"condition).",
        f"Our Medicare members in that pool: {cell_n:,.0f}.",
        f"Visits they made in 2025 to this specialty: {ma_visits:,.0f} (visit = one member "
        f"seeing one provider on one day).",
        f"Rate = {ma_visits:,.0f} / {cell_n:,.0f} = {computed:.2f} visits per member per year.",
        "Every rate in this workbook is this same division, one pool at a time.",
    ])
    ws.freeze_panes = "A3"
    return (len(d["ccir_overview"]) + len(d["top_chronic"]) + len(d["top_chronic_state"])
            + len(d["visit_freq"]) + len(am))


# ---------- tab 3: Demand by County ----------
def build_demand_county(wb, d):
    ws = wb.create_sheet("Demand by County")
    title(ws, "Demand by County", "total vs MA demand per county x specialty", ncols=9)
    for col, w in {"A": 3, "B": 26, "C": 12, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    r = 4
    t0 = r
    r = write_table(ws, d["state"],
                    [("state_cd", "State", 8, None, "left"),
                     ("total_demand", "Total Demand", 16, "#,##0", "right"),
                     ("ma", "MA Demand", 14, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Total vs MA demand by state", t0, t0 + len(d["state"]), 2, 3, 4, f"F{t0}")
    r = max(r, t0 + 15) + 1
    r = kv(ws, r, "Caveat",
           "Total demand uses the state morbidity mix for every county.", h=24)
    r = blank(ws, r)

    # worked example: one county's total demand, every input shown
    wi = d["we_inputs"]
    total = 0.0
    for band, grp in wi.groupby("age_band"):
        blended = float((grp["mix_share"] * grp["rate_ma_proxy"]).sum())
        total += float(grp["eligibles_in_band"].iloc[0]) * blended
    check("one county's total demand", total, d["we_dem_stored"]["total_demand_visits"].iloc[0])
    b = wi[wi["age_band"] == "70-74"]
    rate = {row["morbidity_level"]: float(row["rate_ma_proxy"]) for _, row in b.iterrows()}
    mix = {row["morbidity_level"]: float(row["mix_share"]) for _, row in b.iterrows()}
    blend = sum(mix[k] * rate[k] for k in mix)
    elig_band = float(b["eligibles_in_band"].iloc[0])
    band_demand = elig_band * blend
    r = memo(ws, r, "one county's total demand, every input shown", [
        f"{d['we_cty_name']} has {float(b['county_eligibles_total'].iloc[0]):,.0f} Medicare "
        f"eligibles (CMS penetration file).",
        f"Census: {float(b['age_share'].iloc[0]):.0%} of its 60+ population is 70-74, so "
        f"{elig_band:,.0f} eligibles sit in that band.",
        f"FL 70-74 {d['topspec_desc']} rates: CHRONIC {rate.get('CHRONIC', 0):.2f}, "
        f"NON_CHRONIC {rate.get('NON_CHRONIC', 0):.2f} visits/member/yr (from our claims).",
        f"FL 70-74 morbidity mix: {mix.get('CHRONIC', 0):.0%} CHRONIC, "
        f"{mix.get('NON_CHRONIC', 0):.0%} NON_CHRONIC -> blended rate = {blend:.2f}.",
        f"This band alone: {elig_band:,.0f} x {blend:.2f} = {band_demand:,.0f} visits.",
        f"Same arithmetic for the other four bands, summed = {total:,.0f} — the row you can "
        f"find in the table below.",
        "No member data enters this number: population and age mix are public, only the rates "
        "come from our claims.",
    ])
    r = blank(ws, r)
    write_table(ws, d["dem"],
                [("state_cd", "State", 8, None, "left"),
                 ("county_fips", "FIPS", 8, None, "left"),
                 ("county_name", "County", 18, None, "left"),
                 ("specialty_ctg_cd", "Spec Code", 10, None, "left"),
                 ("specialty_desc", "Specialty", 26, None, "left"),
                 ("total_demand_visits", "Total Demand", 15, "#,##0", "right"),
                 ("ma_demand_visits", "MA Demand", 13, "#,##0", "right")],
                r, filters=True)
    return len(d["dem"])


# ---------- tab 4: Capacity Method ----------
def build_capacity_method(wb, d):
    ws = wb.create_sheet("Capacity Method")
    for col, w in {"A": 3, "B": 26, "C": 12, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Capacity Method",
          "provider_slots = typical capacity x active x (1 - saturation)", ncols=9)
    r = 4
    r = section_header(ws, r, 2, 8, "THE FORMULA")
    r = kv(ws, r, "provider_slots",
           "typical_annual_capacity x active_flag x (1 - senior_saturation)", h=22)
    r = kv(ws, r, "typical_annual_capacity",
           "p75 of observed 2025 Aetna ME visits per provider within state x specialty; "
           "4-state pooled p75 where under 20 providers", h=30)
    r = kv(ws, r, "active_flag", "paid Aetna claim in 2024-2025")
    r = kv(ws, r, "senior_saturation",
           "provider percentile of CMS FFS total Medicare panel within state x specialty; "
           "missing FFS match receives the median, flagged", h=30)
    r = blank(ws, r)

    r = section_header(ws, r, 2, 8, "TYPICAL CAPACITY PER SPECIALTY (FL)")
    t0 = r
    r = write_table(ws, d["p75"],
                    [("cms_specialty", "CMS Specialty", 26, None, "left"),
                     ("providers", "Providers", 11, "#,##0", "right"),
                     ("typ_cap", "Typical Cap (p75)", 16, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Typical capacity per specialty (FL, p75)",
            t0, t0 + len(d["p75"]), 2, 4, 4, f"F{t0}")
    r = max(r, t0 + 15) + 1

    # worked example: one provider, three knobs
    we = d["we_prov"].iloc[0]
    typ, sat = float(we["typical_annual_capacity"]), float(we["senior_saturation"])
    computed = typ * float(we["active_flag"]) * (1 - sat)
    check("one provider, three knobs", computed, we["provider_slots"])
    r = memo(ws, r, "one provider, three knobs", [
        f"Take one FL {we['cms_specialty']} provider (id withheld to one row for audit).",
        f"A busy provider of this specialty handles {typ:,.0f} visits/yr (the p75 of what FL "
        f"providers actually delivered).",
        "They billed us in the last two years, so active = 1.",
        f"Their CMS Medicare panel puts them at the {sat:.0%} percentile of their peers — that "
        f"share of their year is already spoken for.",
        f"Usable slots = {typ:,.0f} x 1 x (1 - {sat:.2f}) = {computed:,.0f}.",
        "Sum every contracted provider's slots in a county = that county's capacity.",
    ])

    r = section_header(ws, r, 2, 8, "CONTRACTED vs ACTIVE vs EFFECTIVE")
    t0 = r
    r = write_table(ws, d["funnel"],
                    [("state_cd", "State", 8, None, "left"),
                     ("contracted", "Contracted", 12, "#,##0", "right"),
                     ("active", "Active", 10, "#,##0", "right"),
                     ("slot_equivalents", "Slot Equivalents", 15, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Contracted vs active vs effective slots per state",
            t0, t0 + len(d["funnel"]), 2, 3, 5, f"G{t0}")
    r = max(r, t0 + 15) + 1

    r = section_header(ws, r, 2, 8, "IMPUTED SATURATION SHARE")
    r = write_table(ws, d["imp"],
                    [("state_cd", "State", 8, None, "left"),
                     ("imputed_share", "Imputed Share", 14, "0.0%", "right")],
                    r, c0=2, set_width=False)
    r = kv(ws, r, "Model soft spot",
           "26-33 percent of provider rows carry an imputed median saturation where no CMS "
           "FFS match exists.", h=28)
    r = kv(ws, r, "The ruler",
           "Capacity is measured on the Aetna-observed visit ruler; it is comparable to MA "
           "demand, not to total demand — see Gap tab.", h=28)
    ws.freeze_panes = "A3"
    return len(d["p75"]) + len(d["funnel"]) + len(d["imp"])


# ---------- tab 5: Capacity by County ----------
def build_capacity_county(wb, d):
    ws = wb.create_sheet("Capacity by County")
    title(ws, "Capacity by County", "effective visit capacity per county x specialty x plan")
    write_table(ws, d["cap"],
                [("state_cd", "State", 8, None, "left"),
                 ("county_fips", "FIPS", 8, None, "left"),
                 ("county_name", "County", 18, None, "left"),
                 ("cms_specialty", "CMS Specialty", 26, None, "left"),
                 ("plan_type", "Plan Type", 10, None, "left"),
                 ("contracted_providers", "Contracted", 12, "#,##0", "right"),
                 ("active_providers", "Active", 10, "#,##0", "right"),
                 ("capacity_visits", "Capacity (visits)", 15, "#,##0", "right")],
                4, filters=True)
    ws.freeze_panes = "A5"
    return len(d["cap"])


# ---------- tab 6: Gap Report ----------
GAP_COLS = [
    ("state_cd", "State", 8, None, "left"),
    ("county_fips", "FIPS", 8, None, "left"),
    ("county_name", "County", 16, None, "left"),
    ("county_type", "County Type", 12, None, "left"),
    ("cms_specialty", "CMS Specialty", 24, None, "left"),
    ("plan_type", "Plan", 9, None, "left"),
    ("compliance_status", "Compliance", 14, None, "left"),
    ("required_provider_count", "Required", 10, "#,##0", "right"),
    ("actual_count", "Actual", 9, "#,##0", "right"),
    ("ma_demand_visits", "MA Demand", 12, "#,##0", "right"),
    ("capacity_visits", "Capacity", 11, "#,##0", "right"),
    ("demand_capacity_gap", "Gap", 11, "#,##0", "right"),
    ("market_opportunity_ratio", "Mkt Opp Ratio", 12, "0.000", "right"),
    ("gap_status", "Gap Status", 16, None, "left"),
    ("risk_flag", "Risk Flag", 10, None, "center"),
]


def build_gap(wb, d):
    ws = wb.create_sheet("Gap Report")
    n = len(GAP_COLS)
    title(ws, "Gap Report", "demand vs capacity on the compliance grid", ncols=n)
    r = 4
    r = kv(ws, r, "Gap definition",
           "ma_demand_visits minus capacity_visits — like-for-like on the Aetna-observed "
           "ruler. market_opportunity_ratio = capacity / total demand, context only.", h=28)
    r = kv(ws, r, "Thresholds",
           "DESERT when gap exceeds 20 percent of MA demand; OVERSUPPLY when capacity exceeds "
           "150 percent — first-pass cuts, rank by gap size rather than treating the flag as "
           "a verdict.", h=36)
    r = blank(ws, r)

    # worked example: one gap row, subtraction in the open
    we = d["we_gap"].iloc[0]
    ma, cap = float(we["ma_demand_visits"]), float(we["capacity_visits"])
    gap_val = ma - cap
    check("one gap row", gap_val, we["demand_capacity_gap"])
    ratio = float(we["market_opportunity_ratio"]) if pd.notna(we["market_opportunity_ratio"]) else 0.0
    r = memo(ws, r, "one gap row, subtraction in the open", [
        f"{we['county_name']} x {we['cms_specialty']} x {we['plan_type']}.",
        f"Our members there will need about {ma:,.0f} visits this year (Demand tab math).",
        f"Our contracted network can deliver about {cap:,.0f} (Capacity tab math).",
        f"Gap = {ma:,.0f} - {cap:,.0f} = {gap_val:,.0f} visits short.",
        f"CMS status: {we['compliance_status']} — the county passes the floor while its "
        f"members face a {gap_val:,.0f}-visit shortfall. That combination is the risk_flag, "
        f"and it is the headline of this report.",
        f"For scale: our capacity there could touch {ratio:.0%} of the county's total "
        f"Medicare need.",
    ])
    r = blank(ws, r)
    hdr = r
    write_table(ws, d["gap"], GAP_COLS, hdr, filters=True)
    status_col = get_column_letter([k for k, *_ in GAP_COLS].index("gap_status") + 1)
    rng = f"{status_col}{hdr + 1}:{status_col}{hdr + len(d['gap'])}"
    for value, hx in (("DESERT", LIGHT_RED), ("OVERSUPPLY", LIGHT_GOLD), ("BALANCED", LIGHT_GREEN)):
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill(hx)))
    return len(d["gap"])


# ---------- tab 7: Deserts & Risk ----------
def build_deserts(wb, d):
    ws = wb.create_sheet("Deserts & Risk")
    for col, w in {"A": 3, "B": 26, "C": 12, "D": 20, "E": 22, "F": 22, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Deserts & Risk",
          "where demand exceeds capacity, and who is compliant but strained", ncols=9)
    r = 4
    r = section_header(ws, r, 2, 8, "THE QUADRANT THAT MATTERS")
    quad_rows = []
    for st in sorted(d["quad"]["state_cd"].dropna().unique()):
        sub = d["quad"][d["quad"]["state_cd"] == st]
        get = lambda comp, ok: int(sub[(sub["compliant"] == comp) & (sub["capacity_ok"] == ok)]["cells"].sum())
        quad_rows.append({"state_cd": st,
                          "comp_ok": get(True, True), "comp_strained": get(True, False),
                          "noncomp_ok": get(False, True), "noncomp_strained": get(False, False)})
    quad = pd.DataFrame(quad_rows)
    r = write_table(ws, quad,
                    [("state_cd", "State", 8, None, "left"),
                     ("comp_ok", "Compliant + Capacity OK", 20, "#,##0", "right"),
                     ("comp_strained", "Compliant + Strained", 19, "#,##0", "right"),
                     ("noncomp_ok", "Non-compliant + Capacity OK", 23, "#,##0", "right"),
                     ("noncomp_strained", "Non-compliant + Strained", 21, "#,##0", "right")],
                    r, c0=2, set_width=False)
    r = blank(ws, r)
    for _, row in d["strain"].iterrows():
        r = kv(ws, r, f"{row['state_cd']} compliant-but-strained",
               f"{_int(row['compliant_but_strained']):,} of {_int(row['mapped_cells']):,} mapped cells")
    t0 = r + 1
    r = write_table(ws, d["strain"],
                    [("state_cd", "State", 8, None, "left"),
                     ("compliant_but_strained", "Compliant but Strained", 20, "#,##0", "right"),
                     ("mapped_cells", "Mapped Cells", 13, "#,##0", "right")],
                    t0, c0=2, set_width=False)
    add_bar(ws, "Compliant but strained cells per state",
            t0, t0 + len(d["strain"]), 2, 3, 3, f"F{t0}")
    r = max(r, t0 + 15) + 1

    # worked example: how one state's headline count is built
    wq = d["we_quad"].iloc[0]
    strained, compliant, mapped = int(wq["strained"]), int(wq["compliant"]), int(wq["mapped"])
    fl_stored = d["strain"][d["strain"]["state_cd"] == "FL"]["compliant_but_strained"].iloc[0]
    check("FL headline count", strained, fl_stored, tol=0)
    r = memo(ws, r, "how one state's headline count is built", [
        f"FL has {mapped:,.0f} county x specialty x plan cells with demand mapped.",
        f"{compliant:,.0f} of them pass the CMS floor.",
        f"Of those, {strained:,.0f} still show more member need than network capacity (gap > 0).",
        f"{strained:,.0f} is the FL bar in the chart above: legal on paper, short in practice.",
    ])

    r = section_header(ws, r, 2, 8, "TOP PRACTITIONER DESERTS")
    t0 = r
    r = write_table(ws, d["desert_prac"],
                    [("state_cd", "State", 8, None, "left"),
                     ("county_name", "County", 16, None, "left"),
                     ("cms_specialty", "CMS Specialty", 24, None, "left"),
                     ("plan_type", "Plan", 9, None, "left"),
                     ("gap", "Gap (visits)", 12, "#,##0", "right")],
                    r, c0=2, set_width=False)
    add_bar(ws, "Top practitioner deserts by gap",
            t0, t0 + len(d["desert_prac"]), 3, 6, 6, f"H{t0}")
    r = max(r, t0 + 17) + 1

    r = section_header(ws, r, 2, 8, "FACILITY AND ANCILLARY ROWS (CAPACITY MODEL DOES NOT APPLY)")
    r = write_table(ws, d["desert_fac"],
                    [("state_cd", "State", 8, None, "left"),
                     ("county_name", "County", 16, None, "left"),
                     ("cms_specialty", "CMS Specialty", 24, None, "left"),
                     ("plan_type", "Plan", 9, None, "left"),
                     ("gap", "Gap (visits)", 12, "#,##0", "right")],
                    r, c0=2, set_width=False)
    r = kv(ws, r, "Note",
           "These specialties are facility or ancillary types; slot capacity from practitioner "
           "visit volume does not describe them. Listed for completeness, not ranked for "
           "recruitment.", h=36)
    ws.freeze_panes = "A3"
    return len(d["desert_prac"]) + len(d["desert_fac"]) + len(d["strain"])


# ---------- tab 8: Forecast Example ----------
def build_forecast(wb, d):
    ws = wb.create_sheet("Forecast Example")
    for col, w in {"A": 3, "B": 14, "C": 16, "D": 16, "E": 12, "F": 20, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Forecast Example",
          "12-month projection for the near-tipping cell per state", ncols=9)
    r = 4
    r = kv(ws, r, "What this is",
           "A one-time illustrative projection: 3 percent annual eligible growth placeholder "
           "applied monthly, seasonality from the 12 observed 2025 months, capacity held flat. "
           "Not a validated forecast; one year of claims allows no holdout.", h=42)
    r = blank(ws, r)

    # worked example: one projected month (first state alphabetically, month 6)
    if "we_fc_row" in d and len(d["we_fc_ma"]):
        row6 = d["we_fc_row"]
        ma = float(d["we_fc_ma"]["ma_demand_visits"].iloc[0])
        season, growth = float(row6["seasonality_index"]), float(row6["growth_factor"])
        proj = ma * season * growth
        check("one projected month", proj, row6["projected_demand"])
        cap_m = float(row6["monthly_capacity"])
        crossover = bool(row6["crossover_flag"])
        verdict = ("June exceeds capacity — a crossover month" if crossover
                   else "June still fits under capacity")
        r = memo(ws, r, "one projected month", [
            f"{row6['county_name']} {row6['cms_specialty']}: annual member demand today = "
            f"{ma:,.0f} visits.",
            f"June's share of the year, from our observed 2025 months: {season:.1%}.",
            f"Growth by month 6 at 3 percent/yr: x{growth:.3f}.",
            f"Projected June demand = {ma:,.0f} x {season:.3f} x {growth:.3f} = {proj:,.0f} visits.",
            f"Monthly capacity, held flat: {cap_m:,.0f}. {verdict}.",
        ])

    for st in sorted(d["fc"]["state_cd"].dropna().unique()):
        sub = d["fc"][d["fc"]["state_cd"] == st].sort_values("forecast_month")
        county = sub["county_name"].iloc[0]
        spec = sub["cms_specialty"].iloc[0]
        n_cross = int(sub["crossover_flag"].astype(bool).sum())
        r = section_header(ws, r, 2, 8, f"{st} — {county} — {spec}")
        t0 = r
        r = write_table(ws, sub,
                        [("forecast_month", "Month", 8, "#,##0", "right"),
                         ("projected_demand", "Projected Demand", 16, "#,##0", "right"),
                         ("monthly_capacity", "Monthly Capacity", 16, "#,##0", "right"),
                         ("crossover_flag", "Crossover", 10, None, "center")],
                        r, c0=2, set_width=False)
        add_line(ws, f"{st} {county} {spec}: crossover months = {n_cross}",
                 t0, t0 + len(sub), 2, 3, 4, f"G{t0}")
        r = max(r, t0 + 16) + 1
    ws.freeze_panes = "A3"
    return len(d["fc"])


# ---------- tab 9: Book Utilization ----------
def build_book(wb, d):
    ws = wb.create_sheet("Book Utilization")
    title(ws, "Book Utilization",
          "delivered visits by provider county x specialty x lob x age band", ncols=9)
    for col, w in {"A": 3, "B": 26, "C": 12, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    r = 4
    r = kv(ws, r, "Geography note",
           "Visits are keyed to the PROVIDER's county — where care was delivered. Member home "
           "county is not attributable in the claims. TOTAL = CP + ME.", h=30)
    r = blank(ws, r)

    # worked example: reading the funnel (pure reads, no check needed)
    total = float(d["we_book_total"]["v"].iloc[0])
    cty = float(d["we_book_cty"]["v"].iloc[0])
    spec = float(d["we_book_spec"]["v"].iloc[0]) if len(d["we_book_spec"]) and pd.notna(
        d["we_book_spec"]["v"].iloc[0]) else 0.0
    cty_share = cty / total if total else 0.0
    spec_share = spec / cty if cty else 0.0
    r = memo(ws, r, "reading the funnel", [
        f"All delivered visits on our book, all four states, CP+ME: {total:,.0f}.",
        f"Delivered inside {d['we_book_cty_name']} (the largest delivery county): {cty:,.0f} "
        f"— {cty_share:.1%} of the book.",
        f"Of those, {d['topspec_desc']}: {spec:,.0f} — {spec_share:.1%} of that county's "
        f"delivery.",
        "Every row in the table below is this same slicing at a different depth.",
    ])
    r = blank(ws, r)
    write_table(ws, d["book"],
                [("state_cd", "State", 8, None, "left"),
                 ("prvdr_county", "Provider County", 18, None, "left"),
                 ("specialty_ctg_cd", "Spec Code", 10, None, "left"),
                 ("specialty_desc", "Specialty", 26, None, "left"),
                 ("lob", "LOB", 8, None, "center"),
                 ("age_band", "Age Band", 10, None, "center"),
                 ("visits", "Visits", 11, "#,##0", "right")],
                r, filters=True)
    return len(d["book"])


# ---------- tab 10: Worked Examples ----------
def build_worked_examples(wb):
    ws = wb.create_sheet("Worked Examples")
    for col, w in {"A": 3, "B": 26, "C": 12, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20, "I": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Worked Examples", "every derived number, hand-calculated from real rows", ncols=9)
    r = 4
    r = kv(ws, r, "Purpose",
           "Every derived number in this workbook, hand-calculated once from real rows, "
           "recomputed live at build time. If any recomputation had disagreed with the "
           "pipeline, this file would not have built.", h=42)
    r = blank(ws, r)
    for heading, lines in WORKED_EXAMPLES:
        r = section_header(ws, r, 2, 9, "Worked example: " + heading, LIGHT_GOLD)
        for line in lines:
            c = ws.cell(row=r, column=2, value=line)
            c.font = Font(name="Arial", size=10)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            r += 1
        r = blank(ws, r)
    ws.freeze_panes = "A3"
    return len(WORKED_EXAMPLES)


# ---------- tab 11: Data Dictionary ----------
DICT_ROWS = [
    ("state_cd", "dc_gap", "Scope state (FL, OH, AZ, IL)."),
    ("county_fips / county_name / county_type", "dc_gap", "County key, name, and CMS designation from the compliance fact."),
    ("cms_specialty / plan_type", "dc_gap", "CMS specialty and MA-HMO / MA-PPO from the compliance grid."),
    ("compliance_status / access_compliant / count_compliant", "dc_gap", "42 CFR 422.116 verdict and its two tests."),
    ("required_provider_count / actual_count / provider_gap", "dc_gap", "CMS minimum, contracted count within distance, and the difference."),
    ("total_demand_visits", "dc_gap (from dc_demand)", "Expected visits from ALL county Medicare eligibles."),
    ("ma_demand_visits", "dc_gap (from dc_demand)", "Expected visits from Aetna MA members at resolved cell rates."),
    ("contracted_providers / active_providers", "dc_gap (from dc_capacity)", "Distinct providers on the row; active = paid Aetna claim 2024-2025."),
    ("capacity_visits", "dc_gap (from dc_capacity)", "Sum of provider_slots for the county x specialty x plan."),
    ("demand_capacity_gap", "dc_gap", "ma_demand_visits - capacity_visits (positive = strain)."),
    ("market_opportunity_ratio", "dc_gap", "capacity_visits / total_demand_visits, context only."),
    ("gap_status", "dc_gap", "NO_DEMAND_MAPPING / DESERT / OVERSUPPLY / BALANCED."),
    ("risk_flag", "dc_gap", "TRUE when COMPLIANT and ma_demand exceeds capacity."),
    ("specialty_ctg_cd / specialty_desc", "dc_demand", "Aetna claims specialty category and description."),
    ("pct_cells_thin", "dc_demand", "Share of rate cells behind the row that used the pooled ALL fallback."),
    ("chronic_condition_count", "dc_member_dim", "Distinct CCIR-chronic diagnosis codes on the member's claims."),
    ("morbidity_level", "dc_member_dim", "CHRONIC if chronic_condition_count >= 1, else NON_CHRONIC (AHRQ CCIR)."),
    ("icd_code / icd_description / chronic_indicator / chronic_label", "dc_ref_ccir", "AHRQ CCIR v2026.1: each ICD-10-CM code flagged 0 / 1 / 9."),
    ("provider_id", "dc_provider_capacity", "PIN (par flag grain: provider x plan x specialty x county)."),
    ("active_flag", "dc_provider_capacity", "aetna_par_flag: paid Aetna claim in 2024-2025."),
    ("tot_benes", "dc_provider_capacity", "CMS FFS total Medicare beneficiaries for the provider (via NPI crosswalk)."),
    ("me_visits_2025", "dc_provider_capacity", "Observed 2025 Aetna ME visits for the provider."),
    ("typical_annual_capacity", "dc_provider_capacity", "p75 of me_visits_2025 within state x specialty (pooled when under 20 providers)."),
    ("senior_saturation / saturation_imputed", "dc_provider_capacity", "Percentile of tot_benes within state x specialty; median-imputed when tot_benes = 0."),
    ("provider_slots", "dc_provider_capacity", "typical_annual_capacity x active_flag x (1 - senior_saturation)."),
    ("forecast_month / growth_factor / seasonality_index", "dc_forecast_example", "Month 1-12, 1.03^(month/12), and the observed 2025 month share."),
    ("projected_demand_visits / monthly_capacity / crossover_flag", "dc_forecast_example", "Projected monthly demand, flat capacity/12, and demand > capacity."),
    ("prvdr_county", "dc_book_utilization", "PROVIDER county (delivery location, not member home)."),
    ("lob", "dc_book_utilization", "CP, ME, or TOTAL (= CP + ME)."),
    ("visits", "dc_book_utilization", "Distinct member x provider x day count."),
    ("age_band", "all dc_ tables", "60-64 / 65-69 / 70-74 / 75-79 / 80+ (UNDER_60 excluded from demand)."),
    ("eligibles_in_band / age_share", "dc_county_population", "County Medicare eligibles split by the census 60+ age shares."),
    ("aetna_ma_members", "dc_county_population", "Aetna members in the county x band from dc_member_dim."),
    ("county_morbidity_index", "dc_county_population", "NULL pending the CMS Geographic Variation county risk score load."),
]


def build_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")
    title(ws, "Data Dictionary", "every reported column, its table, and its meaning", ncols=3)
    dd = pd.DataFrame(DICT_ROWS, columns=["column", "source", "meaning"])
    write_table(ws, dd,
                [("column", "Column", 44, None, "left"),
                 ("source", "Source Table", 26, None, "left"),
                 ("meaning", "Meaning", 80, None, "left")],
                4, filters=True)
    ws.freeze_panes = "A5"
    return len(dd)


# ---------- tab 12: Methodology ----------
def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Methodology", "locked rules, verbatim model sentences, and the v2 parking lot")
    r = 4
    r = section_header(ws, r, 2, 8, "LOCKED RULES")
    r = kv(ws, r, "Visit definition",
           "One visit = one distinct member x provider x day "
           "(COUNT(DISTINCT member_id | srv_prvdr_id | srv_start_dt)).", h=28)
    r = kv(ws, r, "LOB rule", "business_ln_cd CP + ME; Medicare = ME; TOTAL = CP + ME.")
    r = kv(ws, r, "Morbidity",
           "morbidity_level from chronic_condition_count (AHRQ CCIR, primary diagnosis, "
           "dot-insensitive join): CHRONIC if at least one chronic condition, else NON_CHRONIC.", h=28)
    r = kv(ws, r, "Membership dedup",
           "Latest 2025 membership row per member (ROW_NUMBER by eff_dt DESC), scope states "
           "FL OH AZ IL.", h=28)
    r = kv(ws, r, "Rate pooling / thin cells",
           "Rates at state x specialty x age band x morbidity; cells with fewer than 30 "
           "members are is_thin_cell and resolve to the 4-state pooled ALL rate.", h=28)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "MODEL SENTENCES (VERBATIM FROM THE MODULES)")
    r = kv(ws, r, "Capacity (35_dc_capacity.py)",
           "Capacity knobs (approved): typical_annual_capacity = p75 of observed 2025 Aetna ME "
           "visit volume per provider within state x specialty (4-state pooled p75 where a "
           "state x specialty has fewer than 20 providers); senior_saturation = provider "
           "percentile of tot_benes within state x cms_specialty among providers with "
           "tot_benes > 0; providers with tot_benes = 0 receive the state x specialty median "
           "saturation, flagged saturation_imputed. provider_slots = typical_annual_capacity "
           "x active_flag x (1 - senior_saturation).", h=96)
    r = kv(ws, r, "Gap (36_dc_gap.py)",
           "Gap is like-for-like: ma_demand_visits minus capacity_visits (both built from the "
           "Aetna member population and observed-visit ruler). total_demand_visits appears as "
           "context only via market_opportunity_ratio; a subtraction against market demand is "
           "not meaningful because capacity is measured on the Aetna-observed ruler. Demand is "
           "bridged to CMS specialties via ref_specialty_crosswalk (specialty_ctg_cd grain); "
           "cells whose category is absent from the crosswalk remain NO_DEMAND_MAPPING.", h=84)
    r = kv(ws, r, "Forecast (37_dc_forecast_example.py)",
           "Illustrative one-time projection, not a validated forecast and not a refresh "
           "pipeline: trend is demographic (flat 3 percent annual eligible growth applied "
           "monthly as 1.03^(1/12), a stated placeholder until penetration YoY is loaded), "
           "shape is the observed 2025 within-year seasonality at state x specialty level, "
           "capacity is held flat. One year of claims allows no holdout validation.", h=72)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "STANDING CAVEATS")
    r = kv(ws, r, "Estimation method",
           "Every rate is built from Aetna ME claims and applied to the whole population; "
           "a total-Medicare rate source is a v2 upgrade.", h=28)
    r = kv(ws, r, "County morbidity",
           "county_morbidity_index is NULL pending the CMS Geographic Variation county risk "
           "score load; the state morbidity mix is applied to every county.", h=28)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "PARKED FOR V2")
    r = kv(ws, r, "1", "CMS county risk score load (county_morbidity_index).")
    r = kv(ws, r, "2", "Total-Medicare rate via FFS apportionment.")
    r = kv(ws, r, "3", "Gap threshold calibration (DESERT 20 percent / OVERSUPPLY 150 percent are first-pass).")
    r = kv(ws, r, "4", "TIN-vs-individual provider grain check.")
    ws.freeze_panes = "A3"
    return 0


# ---------- assemble ----------
def main():
    WORKED_EXAMPLES.clear()
    d = load()
    wb = Workbook()
    wb.remove(wb.active)
    counts = {}
    counts["Overview"] = build_overview(wb)
    counts["Demand Method"] = build_demand_method(wb, d)
    counts["Demand by County"] = build_demand_county(wb, d)
    counts["Capacity Method"] = build_capacity_method(wb, d)
    counts["Capacity by County"] = build_capacity_county(wb, d)
    counts["Gap Report"] = build_gap(wb, d)
    counts["Deserts & Risk"] = build_deserts(wb, d)
    counts["Forecast Example"] = build_forecast(wb, d)
    counts["Book Utilization"] = build_book(wb, d)
    counts["Worked Examples"] = build_worked_examples(wb)
    counts["Data Dictionary"] = build_dictionary(wb)
    counts["Methodology"] = build_methodology(wb)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}  ({datetime.datetime.now():%Y-%m-%d %H:%M})")
    for tab, n in counts.items():
        print(f"  {tab}: {n} rows")


if __name__ == "__main__":
    main()
