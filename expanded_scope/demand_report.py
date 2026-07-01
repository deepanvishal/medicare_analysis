"""
demand_report.py -- Demand / Utilization report (extension).   [PYTHON / pandas + openpyxl]

Separate from the compliance workbook, same formatting style. Reads the demand tables
you built in BigQuery and writes ONE workbook with the pages we discussed:

  1. Overview           scope, locked rules, definitions
  2. Utilization        COUNT(DISTINCT claim_line_id) at provider x specialty, spread
                        across age bands (60-64..80+), split Medicare vs Commercial
  3. Util by Specialty  same, rolled up to state x specialty x payer (manager view)
  4. HCC Morbidity      same utilization rolled to HCC_v24 (via primary dx), by payer

RULES (locked):
  payer        = member present in mdcr_base_membership -> Medicare, else Commercial
  utilization  = COUNT(DISTINCT claim_line_id)
  HCC          = HCC_v24
  diagnosis    = pri_icd9_dx_cd (primary only)
  population   = age_nbr >= 60, banded 60-64 / 65-69 / 70-74 / 75-79 / 80+
  state        = first 2 chars of prvdr_submarket ("FL Central" -> FL)

INPUT (anbc-hcb-dev.provider_ds_netconf_data_hcb_dev.):
  A870800_medicare_analysis_2025_claims, mdcr_base_membership, HCC_ICD_Mapping_2025
OUTPUT: medicare_demand_utilization.xlsx  (repo root)
Run   : python expanded_scope/demand_report.py    (needs: pip install db-dtypes openpyxl)
"""

import datetime
import pandas as pd
import config as cfg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_XLSX = cfg.repo_path("medicare_demand_utilization.xlsx")

CLAIMS = cfg.src("A870800_medicare_analysis_2025_claims")
MCR    = cfg.src("mdcr_base_membership")
HCC    = cfg.src("HCC_ICD_Mapping_2025")

STATE_NAMES = {"FL": "Florida", "OH": "Ohio", "AZ": "Arizona", "IL": "Illinois"}
AGE_ORDER = ["60-64", "65-69", "70-74", "75-79", "80+"]

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"
LIGHT_GREEN, LIGHT_GOLD = "E2EFDA", "FFF2CC"


# ---------- styling helpers (same as 13_build_report.py) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def section_header(ws, row, c0, c1, text, bg=MID_BLUE):
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell(ws, f"{get_column_letter(c0)}{row}", text, bold=True, color=WHITE, bg=bg, size=11)
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def title(ws, text, sub=None, ncols=8):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        cell(ws, "A2", sub, italic=True, color=DARK_BLUE, size=10, bg=LIGHT_BLUE)
        ws.row_dimensions[2].height = 18


def _int(v, d=0):
    try:
        return d if v is None or pd.isna(v) else int(float(v))
    except (TypeError, ValueError):
        return d


def _float(v, d=0.0):
    try:
        return d if v is None or pd.isna(v) else float(v)
    except (TypeError, ValueError):
        return d


def scope_line(states):
    return " + ".join(f"{STATE_NAMES.get(s, s)} ({s})" for s in states)


# ---------- generic formatted data table (same as 13_build_report.py) ----------
def simple_table(wb, sheet, df, cols, subtitle, payer_key=None):
    """cols = list of (df_key, header, width, number_format, align)."""
    ws = wb.create_sheet(sheet)
    n = len(cols)
    title(ws, sheet.split(". ", 1)[-1], subtitle, ncols=n)
    for i, (_, hdr, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}3", hdr, bold=True, color=WHITE, bg=DARK_BLUE, size=9, h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[3].height = 26
    for ridx, (_, row) in enumerate(df.iterrows(), start=4):
        bg = GREY if ridx % 2 == 0 else WHITE
        if payer_key is not None:
            bg = LIGHT_GREEN if row.get(payer_key) == "Medicare" else LIGHT_GOLD
        for i, (key, _, _, num, align) in enumerate(cols):
            v = row.get(key)
            if num and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i+1)}{ridx}", v, bg=bg, size=9, bdr=True, num=num, h_align=align)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{len(df)+3}"
    return ws


# ---------- data ----------
AGE_CASE = """CASE
      WHEN c.age_nbr BETWEEN 60 AND 64 THEN '60-64'
      WHEN c.age_nbr BETWEEN 65 AND 69 THEN '65-69'
      WHEN c.age_nbr BETWEEN 70 AND 74 THEN '70-74'
      WHEN c.age_nbr BETWEEN 75 AND 79 THEN '75-79'
      WHEN c.age_nbr >= 80             THEN '80+'
    END"""


def load():
    c = cfg.client()
    MM = f"(SELECT DISTINCT member_id FROM `{MCR}`)"           # presence = Medicare member
    PAYER = "IF(mm.member_id IS NOT NULL, 'Medicare', 'Commercial')"
    d = {}

    # long grain: provider x specialty x age band x payer (utilization + members)
    d["util"] = c.query(f"""
        WITH mm AS {MM}
        SELECT
          LEFT(c.prvdr_submarket, 2)                AS state_cd,
          c.prvdr_submarket                         AS submarket,
          c.prvdr_county                            AS county_name,
          c.srv_prvdr_id                            AS pin,
          c.specialty_ctg_cd,
          c.specialty_ctg_cd_desc,
          {AGE_CASE}                                AS age_bucket,
          {PAYER}                                   AS payer,
          COUNT(DISTINCT c.claim_line_id)           AS utilization,
          COUNT(DISTINCT c.member_id)               AS members
        FROM `{CLAIMS}` c
        LEFT JOIN mm ON c.member_id = mm.member_id
        WHERE c.age_nbr >= 60
        GROUP BY state_cd, submarket, county_name, pin,
                 specialty_ctg_cd, specialty_ctg_cd_desc, age_bucket, payer
    """).to_dataframe()

    # distinct members at state x specialty x payer (NOT summable from the long grain --
    # a member seen by several providers would be double counted)
    d["spec_members"] = c.query(f"""
        WITH mm AS {MM}
        SELECT
          LEFT(c.prvdr_submarket, 2)  AS state_cd,
          c.specialty_ctg_cd,
          c.specialty_ctg_cd_desc,
          {PAYER}                     AS payer,
          COUNT(DISTINCT c.member_id) AS members
        FROM `{CLAIMS}` c
        LEFT JOIN mm ON c.member_id = mm.member_id
        WHERE c.age_nbr >= 60
        GROUP BY state_cd, specialty_ctg_cd, specialty_ctg_cd_desc, payer
    """).to_dataframe()

    # morbidity: utilization rolled to HCC_v24 via primary dx
    d["hcc"] = c.query(f"""
        WITH mm AS {MM}
        SELECT
          LEFT(c.prvdr_submarket, 2)        AS state_cd,
          h.HCC_v24                         AS hcc_code,
          {PAYER}                           AS payer,
          COUNT(DISTINCT c.claim_line_id)   AS utilization,
          COUNT(DISTINCT c.member_id)       AS members,
          COUNT(DISTINCT c.pri_icd9_dx_cd)  AS distinct_dx
        FROM `{CLAIMS}` c
        JOIN `{HCC}` h ON TRIM(c.pri_icd9_dx_cd) = TRIM(h.diagnosis_code)
        LEFT JOIN mm ON c.member_id = mm.member_id
        WHERE c.age_nbr >= 60 AND h.HCC_v24 IS NOT NULL
        GROUP BY state_cd, hcc_code, payer
    """).to_dataframe()
    return d


# ---------- tab 1: Overview ----------
def build_overview(wb, d):
    ws = wb.create_sheet("1. Overview")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    util = d["util"]
    states = sorted(x for x in util["state_cd"].dropna().unique())
    title(ws, "Medicare Demand & Utilization",
          f"Age 60+  |  Medicare vs Commercial  |  Service Year 2025  |  {scope_line(states)}")
    r = 4
    r = section_header(ws, r, 2, 8, "OBJECTIVE")
    r = kv(ws, r, "Objective",
           "Measure realized demand (claims utilization) for the 60+ population at the "
           "provider and specialty level, split Medicare vs Commercial, to see where demand "
           "concentrates by geography, age, and condition (HCC).", h=44)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "SCOPE")
    r = kv(ws, r, "Geography", f"{scope_line(states)} — provider submarket / county")
    r = kv(ws, r, "Population", "Members age 60+ (age_nbr). Age bands: " + " / ".join(AGE_ORDER))
    r = kv(ws, r, "Payers", "Medicare and Commercial (both included, reported side by side)")
    r = kv(ws, r, "Service Year", "2025 claims (A870800_medicare_analysis_2025_claims)")
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "HEADLINE")
    tot = int(util["utilization"].sum())
    ma = int(util.loc[util["payer"] == "Medicare", "utilization"].sum())
    comm = tot - ma
    r = kv(ws, r, "Total Utilization (claim lines)", f"{tot:,}")
    r = kv(ws, r, "Medicare", f"{ma:,}  ({100*ma/tot:.1f}%)" if tot else "0")
    r = kv(ws, r, "Commercial", f"{comm:,}  ({100*comm/tot:.1f}%)" if tot else "0")
    r = kv(ws, r, "Providers (PIN)", f"{util['pin'].nunique():,}")
    r = kv(ws, r, "Specialties", f"{util['specialty_ctg_cd'].nunique():,}")
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "LOCKED RULES")
    for lab, val in [
        ("Payer split", "Member present in mdcr_base_membership = Medicare; otherwise Commercial."),
        ("Utilization", "COUNT(DISTINCT claim_line_id) -- one count per distinct claim line."),
        ("Morbidity (HCC)", "HCC_v24, mapped from the primary diagnosis (pri_icd9_dx_cd) via HCC_ICD_Mapping_2025."),
        ("Diagnosis", "Primary diagnosis only (pri_icd9_dx_cd). Secondary dx not counted -- undercounts HCCs."),
        ("State", "First 2 characters of prvdr_submarket."),
        ("HCC labels", "HCC category descriptions attach via a separate crosswalk (pending) -- codes shown for now."),
    ]:
        r = kv(ws, r, lab, val, h=28)
    ws.freeze_panes = "A3"


# ---------- pivot helpers ----------
def _age_wide(df, index_cols):
    """Pivot utilization to one column per age band (+ Total) for the given index."""
    p = df.pivot_table(index=index_cols, columns="age_bucket", values="utilization",
                       aggfunc="sum", fill_value=0)
    for a in AGE_ORDER:
        if a not in p.columns:
            p[a] = 0
    p = p[AGE_ORDER].copy()
    p["total_util"] = p.sum(axis=1)
    return p


AGE_COLS = [(a, a, 9, "#,##0", "right") for a in AGE_ORDER] + \
           [("total_util", "Total Util", 11, "#,##0", "right")]


# ---------- tab 2: Utilization detail (provider x specialty) ----------
def build_util_detail(wb, util):
    idx = ["state_cd", "submarket", "county_name", "pin",
           "specialty_ctg_cd", "specialty_ctg_cd_desc", "payer"]
    out = _age_wide(util, idx)
    out["members"] = util.groupby(idx)["members"].sum()      # valid: one age band per member here
    out = out.reset_index().sort_values(
        ["state_cd", "submarket", "pin", "specialty_ctg_cd", "payer"])
    cols = [
        ("state_cd", "State", 8, None, "left"),
        ("submarket", "Submarket", 18, None, "left"),
        ("county_name", "County", 16, None, "left"),
        ("pin", "PIN", 14, None, "left"),
        ("specialty_ctg_cd", "Spec Code", 10, None, "left"),
        ("specialty_ctg_cd_desc", "Specialty", 26, None, "left"),
        ("payer", "Payer", 11, None, "center"),
    ] + AGE_COLS + [("members", "Members", 10, "#,##0", "right")]
    simple_table(wb, "2. Utilization Detail", out, cols,
                 "Utilization = COUNT(DISTINCT claim_line_id), age 60+  |  provider x specialty x payer, by age band",
                 payer_key="payer")


# ---------- tab 3: Utilization by specialty (rollup) ----------
def build_util_specialty(wb, util, spec_members):
    idx = ["state_cd", "specialty_ctg_cd", "specialty_ctg_cd_desc", "payer"]
    out = _age_wide(util, idx)
    out["providers"] = util.groupby(idx)["pin"].nunique()
    out = out.reset_index().merge(spec_members, on=idx, how="left")
    out = out.sort_values(["state_cd", "total_util"], ascending=[True, False])
    cols = [
        ("state_cd", "State", 8, None, "left"),
        ("specialty_ctg_cd", "Spec Code", 10, None, "left"),
        ("specialty_ctg_cd_desc", "Specialty", 30, None, "left"),
        ("payer", "Payer", 11, None, "center"),
    ] + AGE_COLS + [
        ("providers", "Providers", 10, "#,##0", "right"),
        ("members", "Members", 10, "#,##0", "right"),
    ]
    simple_table(wb, "3. Util by Specialty", out, cols,
                 "Rolled up to state x specialty x payer  |  members are distinct (not summed from detail)",
                 payer_key="payer")


# ---------- tab 4: HCC morbidity ----------
def build_hcc(wb, hcc):
    out = hcc.sort_values(["state_cd", "utilization"], ascending=[True, False])
    cols = [
        ("state_cd", "State", 8, None, "left"),
        ("hcc_code", "HCC v24", 12, None, "left"),
        ("payer", "Payer", 11, None, "center"),
        ("utilization", "Utilization", 12, "#,##0", "right"),
        ("members", "Members", 11, "#,##0", "right"),
        ("distinct_dx", "Distinct Dx", 11, "#,##0", "right"),
    ]
    simple_table(wb, "4. HCC Morbidity", out, cols,
                 "HCC_v24 from primary dx (pri_icd9_dx_cd -> HCC_ICD_Mapping_2025). "
                 "HCC category labels attach via crosswalk (pending).",
                 payer_key="payer")


# ---------- assemble ----------
def build(d):
    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb, d)
    build_util_detail(wb, d["util"])
    build_util_specialty(wb, d["util"], d["spec_members"])
    build_hcc(wb, d["hcc"])
    return wb


def main():
    d = load()
    if d["util"].empty:
        raise SystemExit("claims query returned 0 rows -- check A870800_medicare_analysis_2025_claims.")
    wb = build(d)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}  ({datetime.datetime.now():%Y-%m-%d %H:%M})")
    print(f"  tabs: {wb.sheetnames}")
    print(f"  utilization rows: {len(d['util'])}  | total claim lines: {int(d['util']['utilization'].sum()):,}")
    print(f"  HCC rows: {len(d['hcc'])}")


if __name__ == "__main__":
    main()
