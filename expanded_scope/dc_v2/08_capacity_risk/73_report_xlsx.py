"""
73 - capacity report workbook (export lane)   [PYTHON / pandas + openpyxl]

WHAT  : Frozen-scenario Excel deliverable built from the 74a export tables
        plus the cap_* pipeline tables. Reads ONLY BigQuery tables - the
        fill is NEVER recomputed here. Scenario = measured enrollment
        growth (G_BASE); G_MINUS2/G_PLUS2 appear as sensitivity columns
        only, so rankings do not change within the file. House style
        copied from 56_final_report.py (Arial, DARK_BLUE/MID_BLUE scheme,
        cell/fill/thin helpers); every tab carries the frozen-scenario
        banner and its own caveat line.
TABS  : 1 README, 2 Demand, 3 County Ranking, 4 County Deep-Dive,
        5 Specialty View, 6 Providers, 7 Zero-Utilization, 8 Methodology,
        9 Data Quality.
GRAIN : per tab - see each builder's header comment.
INPUTS: cap_growth_measured, cap_scenario_input, cap_scenario_results,
        cap_county_drivers, cap_action_lists, dem_segment_split,
        cap_provider_year, cap_willing, cap_cohort_bench, cap_hours_annual,
        cap_daily_capped, cap_observed_detail, cap_provider_segment,
        cap_params, ms_ref_county, ref_specialty_crosswalk (cfg.base),
        00_docs/capacity_methodology_v2.md (limitations text, read at
        build time so the report stays verbatim with the doc)
OUTPUT: expanded_scope/dc_v2/08_capacity_risk/outputs/capacity_report.xlsx
Run   : python expanded_scope/dc_v2/08_capacity_risk/73_report_xlsx.py
"""

# ASSUMPTION [1]: "rank" on the County Ranking tab = within-state rank by
#   unplaced_cnt (ties by unplaced_pct), matching module 71's within-state
#   ranking direction (1 = worst). Unattributed rows (NULL cms_specialty or
#   NULL segment) are excluded from the table and counted in the footer.
# ASSUMPTION [2]: the deep-dive prints charts for the TOP 30 counties by
#   G_BASE unplaced (footprint-wide, attributed cells); every county still
#   gets its full tables. Action lists display at most 15 rows each per
#   county ("top 15 of N" noted); the full lists live in cap_action_lists
#   and the Providers tab carries the flags.
# ASSUMPTION [3]: demand-tab "specialty" = bridged cms_specialty (MIN-dedup
#   per aetna_cd, 69 A5); unbridged codes display as 'Unattributed
#   (unbridged)'. Baseline visits = SUM(segment_demand) per county x
#   specialty, which re-sums to the module 50 anchor by V5.
# ASSUMPTION [4]: the Methodology tab's demand flow = the dashboard's
#   FLOW_STAGES 2-7 wording verbatim, renumbered 1-6: the dashboard's
#   stage 1 (the growth slider) is replaced in this EXPORT lane by the
#   measured enrollment growth note. The capacity flow = the dashboard's
#   8-step CAP_FLOW_STAGES wording (DASH-1) verbatim.
# ASSUMPTION [5]: provider/action county keys are provider county NAME +
#   state; demand counties are fips codes. The two meet through
#   ms_ref_county name+state <-> fips (the 69 A2 / 74a A7 pattern).
# ASSUMPTION [6]: Data Quality "sample/full-mode note" - run mode is not
#   recorded in the tables, so each table carries the producing module and
#   a note to check the prompt-pack STATUS line for the run state at build
#   time.

import datetime
import os
import re
import sys
import time


def _expanded_scope_dir():
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.dirname(os.path.dirname(here))
    except NameError:
        probe = os.getcwd()
        for _ in range(6):
            if os.path.isfile(os.path.join(probe, "config.py")):
                return probe
            cand = os.path.join(probe, "expanded_scope")
            if os.path.isfile(os.path.join(cand, "config.py")):
                return cand
            probe = os.path.dirname(probe)
    raise FileNotFoundError(
        "config.py not found - run from the repo root or any folder inside it")


sys.path.insert(0, _expanded_scope_dir())
import config as cfg

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RUN_MODE = "sample"   # no claims scan; reads pipeline tables only

GROWTH = cfg.table("cap_growth_measured")
IN_T   = cfg.table("cap_scenario_input")
RES    = cfg.table("cap_scenario_results")
DRV    = cfg.table("cap_county_drivers")
ACT    = cfg.table("cap_action_lists")
DEM    = cfg.table("dem_segment_split")
PY     = cfg.table("cap_provider_year")
WILL   = cfg.table("cap_willing")
BENCH  = cfg.table("cap_cohort_bench")
ANNUAL = cfg.table("cap_hours_annual")
CAPPED = cfg.table("cap_daily_capped")
OBS    = cfg.table("cap_observed_detail")
MATRIX = cfg.table("cap_provider_segment")
PARAMS = cfg.table("cap_params")
CTY    = cfg.table("ref_county")
XWALK  = cfg.base("ref_specialty_crosswalk")

OUT_XLSX = cfg.repo_path("expanded_scope", "dc_v2", "08_capacity_risk",
                         "outputs", "capacity_report.xlsx")
METHOD_MD = cfg.repo_path("expanded_scope", "dc_v2", "00_docs",
                          "capacity_methodology_v2.md")

RUN_DATE = datetime.date.today().isoformat()
BANNER = (f"Frozen scenario: measured enrollment growth (G_BASE), run "
          f"{RUN_DATE}. Sensitivity = G_MINUS2/G_PLUS2 columns. Rankings do "
          f"not change within this file.")

TOP_CHART_COUNTIES = 30
LIST_ROWS_SHOWN = 15
UNBRIDGED = "Unattributed (unbridged)"
SEGMENTS_8 = ["NEW_CHR_60_74", "NEW_CHR_75P", "NEW_NONCHR_60_74",
              "NEW_NONCHR_75P", "RET_CHR_60_74", "RET_CHR_75P",
              "RET_NONCHR_60_74", "RET_NONCHR_75P"]

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"
LIGHT_GOLD, LIGHT_RED, LIGHT_GREEN = "FFF2CC", "FFE0E0", "E2EFDA"

# ---- wording copied verbatim from whatif_dashboard_v2.py (A4) ----
# Demand flow: FLOW_STAGES 2-7 renumbered 1-6 (stage 1 = slider, replaced
# by the frozen measured-growth note below).
DEMAND_FLOW = [
    ("New member counts per age group.",
     "Real enrollment: members per county per age group, counted from "
     "membership records (Dec 2025). The slider applies to these counts. "
     "Example: 1,000 members aged 85+, slider +10% = 100 new members."),
    ("Condition mix — people, not visits yet.",
     "From 2025 claims we counted, per county and age group, the share of "
     "members with each condition — the prevalence. New members x "
     "prevalence = new members with each condition. Example: 100 x 30% = "
     "30 new CKD members. This is the condition table."),
    ("Visits per specialty — what each condition adds.",
     "The one fitted model. One row per member (1.58M, 2025): 56 yes/no "
     "condition flags -> visit counts per specialty (a visit = one member, "
     "one provider, one day). A regression per specialty gives the base "
     "visits everyone makes plus a coefficient per condition — those "
     "coefficients are the visit rates. Example: 30 CKD members x 2.8 = "
     "84 nephrology visits — the model's job ends at per-member visits; "
     "everything after is addition and a sharing rule."),
    ("Local care pattern — anchor to this county's reality.",
     "The formula uses national patterns; counties have local habits. Per "
     "county and specialty: actual 2025 visits divided by the formula's "
     "2025 number = the factor. Baseline matches reality exactly, and "
     "every slider change is resized to local behavior. Example: factor "
     "0.9 -> 84 becomes 76."),
    ("County demand per specialty — add everyone up.",
     "Each member's predicted visits are summed to the county total per "
     "specialty. No model here — just addition. This is the number on "
     "the demand charts."),
    ("Providers — who absorbs the change.",
     "A sharing rule, not a model: each provider takes a share of the new "
     "demand equal to their share of last year's new patients in that "
     "county and specialty (intake weight). That load is checked against "
     "their ceiling — currently v0: busiest observed month x 12. Modeling "
     "who patients actually choose is a future improvement, listed in "
     "assumptions."),
]
DEMAND_FLOW_NOTE = (
    "In the dashboard, stage 1 is the growth slider (your assumption). In "
    "this frozen export the growth assumption is MEASURED: distinct "
    "enrolled members per state, 2024 vs 2025, g = members_2025 / "
    "members_2024 - 1, applied at G_BASE with +/-2pt sensitivity.")

CAP_FLOW = [
    ("Observed work — every visit we can see.",
     "Aetna MA claims (day grain) plus the CMS FFS public file (annual, "
     "one row per provider) — module 61."),
    ("Minutes per service.",
     "The CMS physician work-time file gives intra-service minutes per "
     "procedure code; unmatched codes get zero minutes by design — "
     "modules 60/62."),
    ("Calibration — every tuning number solved, none hard-coded.",
     "Deflation, the daily cap, benchmark percentile, cohort sizes and "
     "the blending constant are solved and stored in one table (cap_params) "
     "— module 63."),
    ("The feasible day.",
     "Hours are deflated, capped at the calibrated daily maximum; "
     "over-cap hours are kept separately as team uplift — module 64."),
    ("Ceiling — what a full year could hold.",
     "Peer benchmark rate x the provider's own working days = ceiling; "
     "spare = ceiling minus observed — module 65."),
    ("Patient-type matrix.",
     "8 patient types (new/returning x chronic x age). Each provider's "
     "intake per type, credibility-blended with peers; OWN vs BORROWED "
     "tagged — modules 66/67."),
    ("Two-lane fill.",
     "Growth demand splits by patient type; facilities keep their share; "
     "NEW patients deal against intake caps, RETURNING against remaining "
     "room — modules 68/69."),
    ("County risk — who cannot be placed.",
     "Whatever no provider could absorb = unplaced = the risk number, by "
     "county, specialty and patient type — modules 70/71, validated by 72."),
]

GLOSSARY = [
    ("segment / patient type", "One of 8 cells: new/returning x "
     "chronic/non-chronic x age band (60-74 / 75+), defined in ref_segment."),
    ("ceiling", "A provider's feasible annual clinical hours (low/high "
     "range); ceiling_low is used for risk - conservative."),
    ("spare hours", "ceiling_low minus observed capped hours, floor 0."),
    ("team uplift", "Hours trimmed by the daily cap - team-billing signal; "
     "absorbing capacity for the fill = spare + uplift (CD-23)."),
    ("open door", "Provider with nonzero blended intake for at least one "
     "segment."),
    ("unplaced demand", "Growth patients no open-door provider with room "
     "could absorb = the county's risk number."),
    ("facility absorbed", "Growth kept by facility/org billers at their "
     "historical share - no ceiling construct applies (CD-24)."),
    ("OWN / BORROWED", "Whether a provider's intake rate is mostly their "
     "own signal or borrowed from the peer cohort (credibility blend)."),
    ("zero-utilization / paper network", "Contracted provider with zero "
     "Aetna MA claims in the window (CD-07); zero claims may be new "
     "contracts."),
    ("measured growth", "Distinct enrolled members per state 2025 vs 2024 "
     "from the membership extract; the export lane's frozen growth rate."),
    ("G_MINUS2 / G_BASE / G_PLUS2", "The frozen scenarios: measured rate "
     "minus 2 points / as measured / plus 2 points, floor 0."),
]

TAB_GUIDE = [
    ("README", "What this file is, the measured growth rates, vintages, "
     "and the 15 methodology limitations verbatim."),
    ("Demand", "Baseline visits and G_BASE growth per county x specialty, "
     "with per-state KPI blocks."),
    ("County Ranking", "One row per county: growth, placed, unplaced, "
     "rank, sensitivity, dominant driver, mix percentages, provider "
     "counts."),
    ("County Deep-Dive", "Per-county blocks: headline, drivers, specialty "
     "table, 8 patient-type buckets, action lists; charts for the top "
     f"{TOP_CHART_COUNTIES} risk counties."),
    ("Specialty View", "Specialty x state: unplaced, counties hit, worst "
     "county, chronic/new mix of the gap."),
    ("Providers", "Provider x county roster: hours, ceiling, spare, "
     "uplift, utilization, Aetna share, flags."),
    ("Zero-Utilization", "Contracted zero-claim providers by county x "
     "specialty (the paper network)."),
    ("Methodology", "The demand flow (6 stages) and capacity flow (8 "
     "steps), assumptions, glossary."),
    ("Data Quality", "Source row counts, match rates, impossible-day and "
     "borrowed shares, run-mode notes."),
]


# ---------- styling helpers (house style, from 56_final_report.py) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size,
                  italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center",
                            wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def section_header(ws, row, c0, c1, text, bg=MID_BLUE):
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell(ws, f"{get_column_letter(c0)}{row}", text, bold=True, color=WHITE,
         bg=bg, size=11)
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def tab_top(ws, text, caveat, ncols=8):
    """Title row + the frozen-scenario banner + the tab caveat (spec)."""
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    cell(ws, "A2", BANNER, bold=True, size=10, bg=LIGHT_GOLD)
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
    cell(ws, "A3", f"Caveat: {caveat}", italic=True, size=9,
         color=DARK_GREY, bg=GREY)
    ws.row_dimensions[3].height = 24
    return 5


def header_row(ws, r, cols, light=False):
    for i, (label, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}{r}", label, bold=True,
             color=(DARK_BLUE if light else WHITE),
             bg=(LIGHT_BLUE if light else DARK_BLUE), size=9,
             h_align="center", bdr=True)
        if not light:
            ws.column_dimensions[col].width = w
    ws.row_dimensions[r].height = 24
    return r + 1


def data_table(ws, df, cols, r0, filters=True, zebra=True, freeze=True):
    """cols = (header, width, df_key, number_format). Styled house table."""
    hdr = r0
    header_row(ws, hdr, cols)
    for ridx, (_, row) in enumerate(df.iterrows(), start=hdr + 1):
        bg = GREY if (zebra and ridx % 2 == 0) else WHITE
        for i, (_, _, key, num) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i + 1)}{ridx}", v, bg=bg, size=9,
                 bdr=True, num=num,
                 h_align=("right" if num else "left"))
    if filters and len(df):
        ws.auto_filter.ref = (f"A{hdr}:{get_column_letter(len(cols))}"
                              f"{hdr + len(df)}")
    if freeze:
        ws.freeze_panes = f"A{hdr + 1}"
    return hdr + len(df)


def plain_table(ws, df, cols, r0):
    """Light writer for big tabs (Providers): headers styled, data plain."""
    hdr = r0
    header_row(ws, hdr, cols)
    r = hdr
    for _, row in df.iterrows():
        r += 1
        for i, (_, _, key, num) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if v is None or pd.isna(v):
                v = None
            c = ws.cell(row=r, column=i + 1, value=v)
            c.font = Font(name="Arial", size=9)
            if num:
                c.number_format = num
    if len(df):
        ws.auto_filter.ref = (f"A{hdr}:{get_column_letter(len(cols))}"
                              f"{hdr + len(df)}")
    ws.freeze_panes = f"A{hdr + 1}"
    return r


def limitations_from_doc():
    with open(METHOD_MD, "r", encoding="utf-8") as f:
        text = f.read()
    m = re.search(r"## 10\. Limitations.*?\n(.*?)(?:\n## )", text, re.S)
    if not m:
        raise SystemExit(f"STOP -- limitations section not found in {METHOD_MD}")
    items = re.findall(r"^\d+\.\s.*$", m.group(1), re.M)
    if len(items) != 15:
        raise SystemExit(f"STOP -- expected 15 limitations, found {len(items)}")
    return items


# ---------- data ----------
def _q(client, label, sql):
    t0 = time.time()
    df = client.query(sql).result().to_dataframe()
    print(f"[{label}] {len(df):,} rows, {time.time() - t0:.1f}s")
    return df


def load(client):
    d = {}
    d["growth"] = _q(client, "growth", f"""
        SELECT state_cd, members_2024, members_2025, g_state
        FROM `{GROWTH}` ORDER BY state_cd""")

    d["demand"] = _q(client, "demand cells", f"""
        WITH x AS (SELECT aetna_cd, MIN(cms_specialty) AS cms_specialty
                   FROM `{XWALK}` GROUP BY 1)
        SELECT d.mbr_state_cd, d.mbr_county_cd, d.specialty_ctg_cd,
               x.cms_specialty,
               SUM(d.segment_demand) AS baseline_visits,
               SUM(d.growth_demand)  AS growth_visits
        FROM `{IN_T}` d LEFT JOIN x ON d.specialty_ctg_cd = x.aetna_cd
        WHERE d.scenario_cd = 'G_BASE'
        GROUP BY 1, 2, 3, 4""")

    d["scen_county"] = _q(client, "scenario county rollup", f"""
        SELECT scenario_cd, mbr_state_cd, mbr_county_cd,
               (cms_specialty IS NOT NULL AND segment_cd IS NOT NULL)
                 AS attributed,
               SUM(growth_demand)         AS growth,
               SUM(placed_cnt)            AS placed,
               SUM(facility_absorbed_cnt) AS facility,
               SUM(unplaced_cnt)          AS unplaced
        FROM `{RES}` WHERE row_type_cd = 'CELL'
        GROUP BY 1, 2, 3, 4""")

    d["cells"] = _q(client, "G_BASE cells", f"""
        SELECT mbr_state_cd, mbr_county_cd, cms_specialty, segment_cd,
               SUM(growth_demand)         AS growth,
               SUM(placed_cnt)            AS placed,
               SUM(facility_absorbed_cnt) AS facility,
               SUM(unplaced_cnt)          AS unplaced
        FROM `{RES}` WHERE row_type_cd = 'CELL' AND scenario_cd = 'G_BASE'
        GROUP BY 1, 2, 3, 4""")

    d["open"] = _q(client, "open providers per county", f"""
        SELECT mbr_state_cd, mbr_county_cd,
               COUNT(DISTINCT IF(seg_market_share > 0,
                     COALESCE(npi, epdb_dw_prvdr_id), NULL)) AS providers_open
        FROM `{RES}` WHERE row_type_cd = 'ALLOC' AND scenario_cd = 'G_BASE'
        GROUP BY 1, 2""")

    d["drivers"] = _q(client, "drivers", f"SELECT * FROM `{DRV}`")
    d["actions"] = _q(client, "action lists", f"SELECT * FROM `{ACT}`")

    d["cty"] = _q(client, "county names", f"""
        SELECT DISTINCT LPAD(TRIM(CAST(county_fips AS STRING)), 5, '0')
                 AS fips, county_name, state_cd
        FROM `{CTY}`""")

    d["providers"] = _q(client, "providers", f"""
        SELECT p.npi, p.epdb_dw_prvdr_id, p.prvdr_state_cd, p.prvdr_county,
               p.specialty_ctg_cd, p.capped_hrs_yr, p.ceiling_low_hrs,
               p.spare_hrs, p.team_uplift_hrs, p.util_ratio,
               w.aetna_share, w.contracted_flag, w.zero_utilization_flag
        FROM `{PY}` p
        LEFT JOIN `{WILL}` w
          ON COALESCE(p.npi, '') = COALESCE(w.npi, '')
          AND COALESCE(p.epdb_dw_prvdr_id, '') = COALESCE(w.epdb_dw_prvdr_id, '')
          AND COALESCE(p.prvdr_county, '(NULL)') = COALESCE(w.prvdr_county, '(NULL)')
        ORDER BY p.prvdr_state_cd, p.prvdr_county, p.util_ratio DESC""")

    d["zero"] = _q(client, "zero-utilization", f"""
        SELECT w.prvdr_state_cd, w.prvdr_county, p.specialty_ctg_cd,
               COUNT(DISTINCT COALESCE(w.npi, w.epdb_dw_prvdr_id)) AS providers
        FROM `{WILL}` w
        LEFT JOIN `{PY}` p
          ON COALESCE(w.npi, '') = COALESCE(p.npi, '')
          AND COALESCE(w.epdb_dw_prvdr_id, '') = COALESCE(p.epdb_dw_prvdr_id, '')
          AND COALESCE(w.prvdr_county, '(NULL)') = COALESCE(p.prvdr_county, '(NULL)')
        WHERE w.zero_utilization_flag = 1
        GROUP BY 1, 2, 3
        ORDER BY 1, 2, providers DESC""")

    d["quality_rows"] = _q(client, "table row counts", " UNION ALL ".join(
        f"SELECT '{name}' AS table_name, '{module}' AS produced_by, "
        f"COUNT(*) AS row_count FROM `{ref}`"
        for name, module, ref in [
            ("cap_observed_detail", "61", OBS),
            ("cap_hours_annual", "62", ANNUAL),
            ("cap_daily_capped", "64", CAPPED),
            ("cap_provider_year", "65", PY),
            ("cap_cohort_bench", "66", BENCH),
            ("cap_provider_segment", "67", MATRIX),
            ("dem_segment_split", "68", DEM),
            ("cap_willing", "70", WILL),
            ("cap_growth_measured", "74a", GROWTH),
            ("cap_scenario_input", "74a", IN_T),
            ("cap_scenario_results", "74a", RES),
            ("cap_county_drivers", "74a", DRV),
            ("cap_action_lists", "74a", ACT),
            ("cap_params", "63", PARAMS),
        ]))

    d["quality_stats"] = _q(client, "quality stats", f"""
        SELECT 'time-match % (internal services with MPFS minutes)' AS metric,
               SAFE_DIVIDE(SUM(mapped_svc_cnt),
                 SUM(mapped_svc_cnt) + SUM(unmapped_svc_cnt)) AS value
        FROM `{ANNUAL}` WHERE src = 'AETNA_MA'
        UNION ALL
        SELECT 'impossible-day % (raw > 24h, pre-cap)',
               AVG(impossible_day_flag) FROM `{CAPPED}`
        UNION ALL
        SELECT 'borrowed % of G_BASE placed volume (CD-14)',
               SAFE_DIVIDE(SUM(IF(signal_src_cd = 'BORROWED', placed_cnt, 0)),
                           SUM(placed_cnt))
        FROM `{RES}` WHERE row_type_cd = 'ALLOC' AND scenario_cd = 'G_BASE'
        UNION ALL
        SELECT 'borrowed % of matrix cells',
               COUNTIF(signal_src_cd = 'BORROWED') / COUNT(*) FROM `{MATRIX}`
        UNION ALL
        SELECT 'force-normalized providers (limitation 15)',
               CAST(COUNT(DISTINCT IF(alloc_forced_flag = 1,
                    COALESCE(npi, epdb_dw_prvdr_id), NULL)) AS FLOAT64)
        FROM `{PY}`
        UNION ALL
        SELECT 'fallback-cohort % (bench rolled to state x specialty)',
               AVG(fallback_flag) FROM `{BENCH}`
        UNION ALL
        SELECT 'CMS rows on cohort avg-minutes fallback %',
               SAFE_DIVIDE(COUNTIF(avg_mins_src_cd = 'COHORT'),
                           COUNTIF(avg_mins_src_cd IS NOT NULL))
        FROM `{ANNUAL}` WHERE src = 'CMS_FFS'""")
    return d


# ---------- shared county assembly ----------
def county_frame(d):
    """One row per county (G_BASE, attributed): all ranking measures (A1)."""
    name_by_fips = {r.fips: (r.county_name, r.state_cd)
                    for r in d["cty"].itertuples()}

    def cname(cty):
        return name_by_fips.get(str(cty).zfill(5), (str(cty), None))[0]

    sc = d["scen_county"]
    base = sc[(sc["scenario_cd"] == "G_BASE") & sc["attributed"]] \
        .set_index(["mbr_state_cd", "mbr_county_cd"])
    minus = sc[(sc["scenario_cd"] == "G_MINUS2") & sc["attributed"]] \
        .set_index(["mbr_state_cd", "mbr_county_cd"])["unplaced"]
    plus = sc[(sc["scenario_cd"] == "G_PLUS2") & sc["attributed"]] \
        .set_index(["mbr_state_cd", "mbr_county_cd"])["unplaced"]

    cells = d["cells"]
    attr = cells[cells["cms_specialty"].notna() & cells["segment_cd"].notna()]
    seg = attr.groupby(["mbr_state_cd", "mbr_county_cd", "segment_cd"])[
        "unplaced"].sum()
    spec = attr.groupby(["mbr_state_cd", "mbr_county_cd", "cms_specialty"])[
        "unplaced"].sum()

    drv = d["drivers"].groupby(["mbr_state_cd", "mbr_county_cd"])[
        ["unplaced_no_providers", "unplaced_doors_closed",
         "unplaced_at_capacity"]].sum()
    paper = d["drivers"].groupby(["mbr_state_cd", "mbr_county_cd"])[
        "paper_network_cnt"].max()

    act = d["actions"]
    act = act.assign(_key=list(zip(act["prvdr_state_cd"],
                                   act["prvdr_county"].astype(str)
                                   .str.upper().str.strip())))
    at_cap = act[act["list_cd"] == "AT_CAPACITY"].groupby("_key").size()
    zero = act[act["list_cd"] == "ZERO_CLAIM"].groupby("_key").size()

    open_p = d["open"].set_index(["mbr_state_cd", "mbr_county_cd"])[
        "providers_open"]

    rows = []
    for (st, cty), r in base.iterrows():
        growth = float(r["growth"] or 0)
        unplaced = float(r["unplaced"] or 0)
        try:
            segs = seg.xs((st, cty), level=[0, 1])
        except KeyError:
            segs = pd.Series(dtype=float)
        try:
            specs = spec.xs((st, cty), level=[0, 1]).sort_values(
                ascending=False)
        except KeyError:
            specs = pd.Series(dtype=float)
        seg_tot = float(segs.sum()) or 0.0
        chr_u = float(segs[[s for s in segs.index if "_CHR_" in s]].sum()) \
            if len(segs) else 0.0
        p75_u = float(segs[[s for s in segs.index if s.endswith("75P")]]
                      .sum()) if len(segs) else 0.0
        new_u = float(segs[[s for s in segs.index if s.startswith("NEW")]]
                      .sum()) if len(segs) else 0.0
        if (st, cty) in drv.index:
            dr = drv.loc[(st, cty)]
            order = [("NO_PROVIDERS", dr["unplaced_no_providers"]),
                     ("DOORS_CLOSED", dr["unplaced_doors_closed"]),
                     ("AT_CAPACITY", dr["unplaced_at_capacity"])]
            dominant = max(order, key=lambda t: float(t[1] or 0))[0] \
                if unplaced > 0 else "-"
        else:
            dominant = "-"
        name = cname(cty)
        key = (st, str(name).upper().strip())
        rows.append({
            "state": st, "fips": str(cty), "county": name,
            "growth": growth, "facility": float(r["facility"] or 0),
            "placed": float(r["placed"] or 0), "unplaced": unplaced,
            "unplaced_pct": (unplaced / growth) if growth > 0 else 0.0,
            "unplaced_minus2": float(minus.get((st, cty), 0) or 0),
            "unplaced_plus2": float(plus.get((st, cty), 0) or 0),
            "dominant_driver": dominant,
            "top_specialty": (specs.index[0] if len(specs) else "-"),
            "top_spec_unplaced": (float(specs.iloc[0]) if len(specs) else 0.0),
            "pct_chronic": (chr_u / seg_tot) if seg_tot else 0.0,
            "pct_75p": (p75_u / seg_tot) if seg_tot else 0.0,
            "pct_new": (new_u / seg_tot) if seg_tot else 0.0,
            "providers_open": int(open_p.get((st, cty), 0) or 0),
            "providers_at_capacity": int(at_cap.get(key, 0)),
            "providers_zero_claim": int(zero.get(key, 0)),
            "paper_network": int(paper.get((st, cty), 0) or 0),
        })
    df = pd.DataFrame(rows)
    if df.empty:
        raise SystemExit("STOP -- no attributed G_BASE county rows found; "
                         "run 74a before 73")
    df["rank"] = df.groupby("state")["unplaced"] \
        .rank(method="first", ascending=False).astype(int)
    return df.sort_values(["state", "rank"]).reset_index(drop=True)


# ---------- tabs ----------
def build_readme(wb, d, limitations):
    ws = wb.create_sheet("README")
    for col, w in {"A": 3, "B": 24, "C": 10, "D": 22, "E": 22, "F": 22,
                   "G": 22, "H": 22}.items():
        ws.column_dimensions[col].width = w
    r = tab_top(ws, "Capacity Risk Report — Frozen Export",
                "This file is a frozen snapshot; the dashboard remains the "
                "interactive (slider) path. Nothing here reacts to input.")
    r = section_header(ws, r, 2, 8, "WHAT EACH TAB ANSWERS")
    for name, what in TAB_GUIDE:
        r = kv(ws, r, name, what, h=26)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8,
                       "MEASURED ENROLLMENT GROWTH (cap_growth_measured)")
    r = kv(ws, r, "How measured",
           "Distinct enrolled members per state, 2024 vs 2025, from the "
           "membership extract (member x month, DD-08) - never claims "
           "utilizers. g = members_2025 / members_2024 - 1. Scenarios: "
           "G_MINUS2 = g - 2pts, G_BASE = g, G_PLUS2 = g + 2pts, floor 0.",
           h=44)
    for row in d["growth"].itertuples():
        r = kv(ws, r, str(row.state_cd),
               f"members 2024: {row.members_2024:,} | members 2025: "
               f"{row.members_2025:,} | g = {float(row.g_state or 0):+.2%}",
               h=18)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "VINTAGES")
    r = kv(ws, r, "Internal claims", "Aetna MA (CP/ME, 60+), 2024-2025; "
           "capacity base year 2025.", h=18)
    r = kv(ws, r, "CMS FFS", "By-provider summary file, calendar 2023 "
           "(annual, no service dates, suppressed cells).", h=18)
    r = kv(ws, r, "Membership", "EMIS extract 2023-2025; growth measured "
           "on 2024 vs 2025.", h=18)
    r = kv(ws, r, "MPFS time file", "CMS-1807-F Work Time, CY2025.", h=18)
    r = kv(ws, r, "Demand anchor", "Module 50 forecast, feature month "
           "2025-12 scored for the next 12 months.", h=18)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8,
                       "LIMITATIONS (verbatim from capacity_methodology_v2.md)")
    for item in limitations:
        r = kv(ws, r, item.split(".")[0] + ".", item.split(". ", 1)[1]
               if ". " in item else item, h=30)
    ws.freeze_panes = "A4"


def build_demand(wb, d):
    # county x specialty demand, per-state KPI block at top
    ws = wb.create_sheet("Demand")
    r = tab_top(ws, "Demand — baseline and measured growth (G_BASE)",
                "Demand = realized visits (module 50 anchor); unmet demand "
                "is out of scope. The split invents no demand (CD-17).",
                ncols=8)
    r = section_header(ws, r, 1, 8, "PER-STATE KPI SUMMARY")
    df = d["demand"].copy()
    df["mbr_state_cd"] = df["mbr_state_cd"].fillna("(none)")
    kpi = df.groupby("mbr_state_cd")[["baseline_visits", "growth_visits"]] \
        .sum().reset_index()
    g_map = {r_.state_cd: r_ for r_ in d["growth"].itertuples()}
    kpi_cols = [("state", 10, "state", None),
                ("members 2024", 14, "m24", "#,##0"),
                ("members 2025", 14, "m25", "#,##0"),
                ("measured g", 11, "g", "+0.00%"),
                ("baseline visits", 15, "base", "#,##0"),
                ("growth visits (G_BASE)", 18, "grow", "#,##0")]
    kpi_rows = []
    for row in kpi.itertuples():
        g = g_map.get(row.mbr_state_cd)
        kpi_rows.append({
            "state": row.mbr_state_cd,
            "m24": (int(g.members_2024) if g is not None else None),
            "m25": (int(g.members_2025) if g is not None else None),
            "g": (float(g.g_state or 0) if g is not None else None),
            "base": float(row.baseline_visits or 0),
            "grow": float(row.growth_visits or 0)})
    r = data_table(ws, pd.DataFrame(kpi_rows), kpi_cols, r, filters=False,
                   freeze=False) + 2
    r = section_header(ws, r, 1, 8, "COUNTY x SPECIALTY")
    name_by_fips = {x.fips: x.county_name for x in d["cty"].itertuples()}
    df["county"] = df["mbr_county_cd"].map(
        lambda c: name_by_fips.get(str(c).zfill(5), str(c)))
    df["specialty"] = df["cms_specialty"].fillna(UNBRIDGED)
    df["after"] = df["baseline_visits"].fillna(0) + df["growth_visits"].fillna(0)
    df["delta_pct"] = df.apply(
        lambda x: (x["growth_visits"] / x["baseline_visits"])
        if x["baseline_visits"] else None, axis=1)
    df = df.sort_values(["mbr_state_cd", "county", "specialty"])
    cols = [("state", 8, "mbr_state_cd", None),
            ("county", 20, "county", None),
            ("specialty", 26, "specialty", None),
            ("baseline visits", 14, "baseline_visits", "#,##0"),
            ("growth (G_BASE)", 14, "growth_visits", "#,##0"),
            ("after growth", 14, "after", "#,##0"),
            ("delta %", 10, "delta_pct", "+0.0%")]
    data_table(ws, df, cols, r)
    ws.freeze_panes = f"A{r + 1}"


def build_ranking(wb, d, cf, excl_note):
    ws = wb.create_sheet("County Ranking")
    r = tab_top(ws, "County Ranking (G_BASE)",
                "Same-county fill (limitation 14); Medicare-visible only "
                "(limitation 1); ceiling_low used - risk is conservative "
                "(limitation 9). Rank = within state, 1 = worst.",
                ncols=19)
    cols = [("state", 7, "state", None),
            ("county", 18, "county", None),
            ("growth", 10, "growth", "#,##0"),
            ("facility absorbed", 11, "facility", "#,##0"),
            ("placed", 10, "placed", "#,##0"),
            ("unplaced", 10, "unplaced", "#,##0"),
            ("unplaced %", 10, "unplaced_pct", "0.0%"),
            ("rank", 6, "rank", "0"),
            ("unplaced G_MINUS2", 11, "unplaced_minus2", "#,##0"),
            ("unplaced G_PLUS2", 11, "unplaced_plus2", "#,##0"),
            ("dominant driver", 14, "dominant_driver", None),
            ("top specialty by unplaced", 20, "top_specialty", None),
            ("its unplaced", 10, "top_spec_unplaced", "#,##0"),
            ("% chronic of gap", 10, "pct_chronic", "0%"),
            ("% 75+ of gap", 10, "pct_75p", "0%"),
            ("% new of gap", 10, "pct_new", "0%"),
            ("providers open", 10, "providers_open", "#,##0"),
            ("at capacity", 10, "providers_at_capacity", "#,##0"),
            ("zero-claim", 10, "providers_zero_claim", "#,##0")]
    last = data_table(ws, cf, cols, r)
    cell(ws, f"A{last + 2}", excl_note, italic=True, size=9, color=DARK_GREY)
    ws.freeze_panes = f"A{r + 1}"


def _mini_header(ws, r, headers, widths=None, c0=1):
    for i, h in enumerate(headers):
        col = get_column_letter(c0 + i)
        cell(ws, f"{col}{r}", h, bold=True, color=WHITE, bg=MID_BLUE,
             size=9, bdr=True, h_align="center")
    return r + 1


def _mini_row(ws, r, values, nums=None, c0=1):
    for i, v in enumerate(values):
        num = nums[i] if nums else None
        cell(ws, f"{get_column_letter(c0 + i)}{r}", v, size=9, bdr=True,
             num=num, h_align=("right" if num else "left"))
    return r + 1


def build_deepdive(wb, d, cf):
    ws = wb.create_sheet("County Deep-Dive")
    for col, w in {"A": 26, "B": 13, "C": 13, "D": 13, "E": 13, "F": 13,
                   "G": 13}.items():
        ws.column_dimensions[col].width = w
    r = tab_top(ws, "County Deep-Dive (G_BASE)",
                "Per-county blocks, ordered worst-first within state. "
                f"Charts on the top {TOP_CHART_COUNTIES} risk counties "
                f"only; action lists show up to {LIST_ROWS_SHOWN} rows "
                "each (full lists in cap_action_lists).", ncols=7)

    cells = d["cells"]
    attr = cells[cells["cms_specialty"].notna() & cells["segment_cd"].notna()]
    drv = d["drivers"]
    act = d["actions"].copy()
    act["_key"] = list(zip(act["prvdr_state_cd"],
                           act["prvdr_county"].astype(str).str.upper()
                           .str.strip()))

    top_fips = set(cf.sort_values("unplaced", ascending=False)
                   .head(TOP_CHART_COUNTIES)["fips"])

    # index
    r = section_header(ws, r, 1, 7, "INDEX (worst first within state)")
    idx_row = r
    r += (len(cf) // 4) + 2
    anchors = {}

    order = cf.sort_values(["state", "rank"])
    blocks_start = r + 1
    r = blocks_start
    for row in order.itertuples():
        anchors[row.fips] = r
        r = section_header(
            ws, r, 1, 7,
            f"{row.state} — {row.county} ({row.fips}) | rank {row.rank} in "
            f"state | unplaced {row.unplaced:,.0f}", bg=DARK_BLUE)
        cell(ws, f"A{r}",
             f"Headline: growth {row.growth:,.0f} -> facility absorbed "
             f"{row.facility:,.0f} -> placed {row.placed:,.0f} -> unplaced "
             f"{row.unplaced:,.0f} ({row.unplaced_pct:.1%} of growth)",
             bold=True, size=10, bg=LIGHT_GOLD)
        ws.merge_cells(f"A{r}:G{r}")
        r += 2

        # driver table (3 causes + paper-network context = the 4 causes)
        r = _mini_header(ws, r, ["driver", "unplaced"])
        sub = drv[(drv["mbr_state_cd"] == row.state)
                  & (drv["mbr_county_cd"].astype(str) == row.fips)]
        sums = sub[["unplaced_no_providers", "unplaced_doors_closed",
                    "unplaced_at_capacity"]].sum()
        r = _mini_row(ws, r, ["NO_PROVIDERS",
                              float(sums["unplaced_no_providers"])],
                      [None, "#,##0"])
        r = _mini_row(ws, r, ["DOORS_CLOSED",
                              float(sums["unplaced_doors_closed"])],
                      [None, "#,##0"])
        r = _mini_row(ws, r, ["AT_CAPACITY",
                              float(sums["unplaced_at_capacity"])],
                      [None, "#,##0"])
        r = _mini_row(ws, r, ["PAPER_NETWORK (context: contracted "
                              "zero-claim providers)", row.paper_network],
                      [None, "#,##0"])
        r += 1

        # specialty table, worst first
        spec = sub.groupby("cms_specialty", dropna=False)[
            ["growth_demand", "unplaced_cnt"]].sum() \
            .sort_values("unplaced_cnt", ascending=False).reset_index()
        spec_start = r
        r = _mini_header(ws, r, ["specialty", "growth", "unplaced"])
        for s in spec.itertuples():
            nm = s.cms_specialty if pd.notna(s.cms_specialty) else UNBRIDGED
            r = _mini_row(ws, r, [nm, float(s.growth_demand or 0),
                                  float(s.unplaced_cnt or 0)],
                          [None, "#,##0", "#,##0"])
        spec_end = r - 1
        r += 1

        # 8-bucket table
        seg = attr[(attr["mbr_state_cd"] == row.state)
                   & (attr["mbr_county_cd"].astype(str) == row.fips)] \
            .groupby("segment_cd")[["growth", "placed", "unplaced"]].sum()
        seg_start = r
        r = _mini_header(ws, r, ["patient type", "growth", "placed",
                                 "unplaced"])
        for s in SEGMENTS_8:
            v = seg.loc[s] if s in seg.index else None
            r = _mini_row(ws, r, [
                s,
                float(v["growth"]) if v is not None else 0.0,
                float(v["placed"]) if v is not None else 0.0,
                float(v["unplaced"]) if v is not None else 0.0],
                [None, "#,##0", "#,##0", "#,##0"])
        seg_end = r - 1
        r += 1

        # native charts, top counties only (A2)
        if row.fips in top_fips:
            n_spec = min(spec_end - spec_start, 10)
            if n_spec > 0:
                ch = BarChart()
                ch.type = "bar"
                ch.title = f"{row.county}: unplaced by specialty (top {n_spec})"
                data = Reference(ws, min_col=3, min_row=spec_start,
                                 max_row=spec_start + n_spec)
                cats = Reference(ws, min_col=1, min_row=spec_start + 1,
                                 max_row=spec_start + n_spec)
                ch.add_data(data, titles_from_data=True)
                ch.set_categories(cats)
                ch.height, ch.width = 7.5, 11
                ch.legend = None
                ws.add_chart(ch, f"I{spec_start}")
            ch2 = BarChart()
            ch2.type = "col"
            ch2.title = f"{row.county}: patient-type buckets"
            data = Reference(ws, min_col=2, max_col=4, min_row=seg_start,
                             max_row=seg_end)
            cats = Reference(ws, min_col=1, min_row=seg_start + 1,
                             max_row=seg_end)
            ch2.add_data(data, titles_from_data=True)
            ch2.set_categories(cats)
            ch2.height, ch2.width = 7.5, 11
            ws.add_chart(ch2, f"I{seg_start + 16}")
            r = max(r, seg_start + 34)

        # action lists (A2: top LIST_ROWS_SHOWN of N)
        key = (row.state, str(row.county).upper().strip())
        for list_cd, label, sort_col, asc in [
                ("TOP_ROOM", "Top providers by remaining room",
                 "remaining_room_hrs", False),
                ("AT_CAPACITY", "At-capacity providers",
                 "used_hrs_g_base", False),
                ("ZERO_CLAIM", "Contracted zero-claim providers",
                 "absorbing_hrs", False)]:
            sub_a = act[(act["list_cd"] == list_cd) & (act["_key"] == key)]
            sub_a = sub_a.sort_values(sort_col, ascending=asc,
                                      na_position="last")
            shown = sub_a.head(LIST_ROWS_SHOWN)
            cell(ws, f"A{r}",
                 f"{label} — showing {len(shown)} of {len(sub_a)}",
                 bold=True, size=9, bg=LIGHT_BLUE)
            ws.merge_cells(f"A{r}:G{r}")
            r += 1
            r = _mini_header(ws, r, ["npi", "epdb id", "specialty",
                                     "absorbing hrs", "used hrs (G_BASE)",
                                     "remaining hrs"])
            for a in shown.itertuples():
                r = _mini_row(ws, r, [
                    a.npi, a.epdb_dw_prvdr_id, a.specialty_ctg_cd,
                    float(a.absorbing_hrs) if pd.notna(a.absorbing_hrs) else None,
                    float(a.used_hrs_g_base) if pd.notna(a.used_hrs_g_base) else None,
                    float(a.remaining_room_hrs)
                    if pd.notna(a.remaining_room_hrs) else None],
                    [None, None, None, "#,##0", "#,##0", "#,##0"])
            r += 1
        r += 1

    # fill the index with hyperlinks (4 per row)
    ir, ic = idx_row, 0
    for row in order.itertuples():
        ref = f"{get_column_letter(1 + (ic % 4) * 2)}{ir + ic // 4}"
        c = cell(ws, ref, f"{row.state} {row.county}", size=9,
                 color=MID_BLUE)
        c.hyperlink = f"#'County Deep-Dive'!A{anchors[row.fips]}"
        ic += 1


def build_specialty(wb, d):
    ws = wb.create_sheet("Specialty View")
    r = tab_top(ws, "Specialty View (G_BASE)",
                "Specialty = bridged cms_specialty (one per Aetna code, "
                "MIN-dedup); Unattributed rows kept as their own group so "
                "no unplaced volume hides.", ncols=8)
    cells = d["cells"]
    name_by_fips = {x.fips: x.county_name for x in d["cty"].itertuples()}
    df = cells.copy()
    df["specialty"] = df["cms_specialty"].fillna(UNBRIDGED)
    grp = df.groupby(["specialty", "mbr_state_cd"], dropna=False)
    rows = []
    for (spec_nm, st), g in grp:
        seg = g[g["segment_cd"].notna()]
        seg_tot = float(seg["unplaced"].sum()) or 0.0
        chr_u = float(seg[seg["segment_cd"].str.contains("_CHR_", na=False)]
                      ["unplaced"].sum())
        new_u = float(seg[seg["segment_cd"].str.startswith("NEW", na=False)]
                      ["unplaced"].sum())
        by_cty = g.groupby("mbr_county_cd")["unplaced"].sum() \
            .sort_values(ascending=False)
        worst = (name_by_fips.get(str(by_cty.index[0]).zfill(5),
                                  str(by_cty.index[0]))
                 if len(by_cty) and by_cty.iloc[0] > 0 else "-")
        rows.append({
            "specialty": spec_nm, "state": st,
            "unplaced": float(g["unplaced"].sum()),
            "counties_hit": int((by_cty > 0).sum()),
            "worst_county": worst,
            "pct_chronic": (chr_u / seg_tot) if seg_tot else 0.0,
            "pct_new": (new_u / seg_tot) if seg_tot else 0.0})
    out = pd.DataFrame(rows).sort_values(["unplaced"], ascending=False)
    cols = [("specialty", 26, "specialty", None),
            ("state", 8, "state", None),
            ("unplaced", 12, "unplaced", "#,##0"),
            ("counties hit", 11, "counties_hit", "#,##0"),
            ("worst county", 18, "worst_county", None),
            ("% chronic of gap", 13, "pct_chronic", "0%"),
            ("% new of gap", 13, "pct_new", "0%")]
    data_table(ws, out, cols, r)
    ws.freeze_panes = f"A{r + 1}"


def build_providers(wb, d):
    ws = wb.create_sheet("Providers")
    r = tab_top(ws, "Providers — hours, ceilings, willingness",
                "Medicare-visible workload only (limitation 1); Aetna share "
                "is a bounded proxy (CD-06); ceilings are a low/high range "
                "and ceiling_low is shown (limitation 9).", ncols=14)
    df = d["providers"].copy()
    act = d["actions"]
    at_cap_keys = set(
        (str(a.npi), str(a.epdb_dw_prvdr_id), str(a.prvdr_county))
        for a in act[act["list_cd"] == "AT_CAPACITY"].itertuples())
    df["at_capacity_flag"] = [
        1 if (str(x.npi), str(x.epdb_dw_prvdr_id), str(x.prvdr_county))
        in at_cap_keys else 0 for x in df.itertuples()]
    cols = [("npi", 13, "npi", None),
            ("epdb id", 12, "epdb_dw_prvdr_id", None),
            ("state", 7, "prvdr_state_cd", None),
            ("county", 16, "prvdr_county", None),
            ("specialty", 12, "specialty_ctg_cd", None),
            ("observed hrs", 11, "capped_hrs_yr", "#,##0"),
            ("ceiling (low)", 11, "ceiling_low_hrs", "#,##0"),
            ("spare hrs", 10, "spare_hrs", "#,##0"),
            ("team uplift", 10, "team_uplift_hrs", "#,##0"),
            ("utilization", 10, "util_ratio", "0%"),
            ("Aetna share", 10, "aetna_share", "0%"),
            ("contracted", 9, "contracted_flag", "0"),
            ("zero-claim", 9, "zero_utilization_flag", "0"),
            ("at capacity (G_BASE)", 11, "at_capacity_flag", "0")]
    plain_table(ws, df, cols, r)


def build_zero(wb, d):
    ws = wb.create_sheet("Zero-Utilization")
    r = tab_top(ws, "Zero-Utilization (paper network)",
                "Contracted, zero Aetna MA claims in 2024-2025 (CD-07). "
                "Zero claims may be new contracts - neutral wording by "
                "design.", ncols=4)
    df = d["zero"].copy()
    cols = [("state", 8, "prvdr_state_cd", None),
            ("county", 20, "prvdr_county", None),
            ("specialty", 14, "specialty_ctg_cd", None),
            ("providers", 11, "providers", "#,##0")]
    data_table(ws, df, cols, r)
    ws.freeze_panes = f"A{r + 1}"


def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    for col, w in {"A": 3, "B": 30, "C": 10, "D": 24, "E": 24, "F": 24,
                   "G": 24, "H": 24}.items():
        ws.column_dimensions[col].width = w
    r = tab_top(ws, "Methodology",
                "Flow wording mirrors the dashboard (DASH-1); the frozen "
                "export replaces the growth slider with the measured "
                "enrollment rate.")
    r = section_header(ws, r, 2, 8, "HOW THE DEMAND NUMBERS FLOW (6 STAGES)")
    r = kv(ws, r, "Note", DEMAND_FLOW_NOTE, h=40)
    for i, (t, s) in enumerate(DEMAND_FLOW, start=1):
        r = kv(ws, r, f"{i}. {t}", s, h=52)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8,
                       "HOW THE CAPACITY NUMBERS FLOW (8 STEPS, v2)")
    for i, (t, s) in enumerate(CAP_FLOW, start=1):
        r = kv(ws, r, f"{i}. {t}", s, h=34)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "ASSUMPTIONS")
    r = kv(ws, r, "Frozen scenarios",
           "Growth = measured enrollment rate per state (floor 0 after the "
           "+/-2pt shifts); segment mix, intake rates, Aetna shares and "
           "local patterns are frozen at their observed values.", h=40)
    r = kv(ws, r, "Sticky shares",
           "New patients distribute like existing ones (closed doors "
           "excluded); defensible for a 1-year horizon only.", h=28)
    r = kv(ws, r, "Same-county fill",
           "Patients are placed within their own county only; cross-county "
           "access is understated - conservative.", h=28)
    r = kv(ws, r, "Aetna share",
           "Applied once, at the end (CD-06); never inside layers.", h=18)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "GLOSSARY")
    for term, text in GLOSSARY:
        r = kv(ws, r, term, text, h=26)
    ws.freeze_panes = "A4"


def build_quality(wb, d):
    ws = wb.create_sheet("Data Quality")
    r = tab_top(ws, "Data Quality",
                "Run mode is not stamped in the tables (A6): check the "
                "prompt-pack STATUS line; sample-mode numbers are 1% of "
                "members on every claims-derived table.", ncols=4)
    r = section_header(ws, r, 1, 4, "SOURCE ROW COUNTS")
    df = d["quality_rows"].copy()
    df["note"] = df["produced_by"].map(
        lambda m: f"module {m}; claims-derived tables reflect that "
                  f"module's RUN_MODE at its last run")
    cols = [("table", 26, "table_name", None),
            ("produced by", 11, "produced_by", None),
            ("rows", 12, "row_count", "#,##0"),
            ("run-mode note", 52, "note", None)]
    r = data_table(ws, df.sort_values("table_name"), cols, r,
                   freeze=False) + 2
    r = section_header(ws, r, 1, 4, "QUALITY METRICS")
    qs = d["quality_stats"].copy()
    cols = [("metric", 52, "metric", None),
            ("value", 16, "value", "0.0000")]
    data_table(ws, qs, cols, r, filters=False, freeze=False)


# ---------- main ----------
def main():
    print(f"RUN_MODE = {RUN_MODE} (reads tables only; mode governed by "
          f"upstream runs)")
    limitations = limitations_from_doc()
    client = cfg.client()
    d = load(client)

    if d["growth"].empty:
        raise SystemExit("STOP -- cap_growth_measured is empty; run 74a first")

    cf = county_frame(d)
    sc = d["scen_county"]
    unattr = sc[(sc["scenario_cd"] == "G_BASE") & (~sc["attributed"])]
    excl_note = (f"Unattributed rows excluded from this table (NULL "
                 f"specialty or patient type): {len(unattr):,} county "
                 f"cells, {float(unattr['unplaced'].sum() or 0):,.0f} "
                 f"unplaced visits. They appear on the Demand and "
                 f"Specialty View tabs as '{UNBRIDGED}'.")

    wb = Workbook()
    wb.remove(wb.active)
    t0 = time.time()
    build_readme(wb, d, limitations)
    build_demand(wb, d)
    build_ranking(wb, d, cf, excl_note)
    build_deepdive(wb, d, cf)
    build_specialty(wb, d)
    build_providers(wb, d)
    build_zero(wb, d)
    build_methodology(wb)
    build_quality(wb, d)
    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"workbook written: {OUT_XLSX} ({time.time() - t0:.1f}s build)")
    print(f"counties ranked: {len(cf):,}; charts on top "
          f"{TOP_CHART_COUNTIES}; providers tab rows: "
          f"{len(d['providers']):,}")


if __name__ == "__main__":
    main()


# REVIEW
# Reviewer 1 LOGIC:
#  - No fill recomputation anywhere: every number is an aggregate of
#    cap_scenario_results CELL/ALLOC rows or the 74a driver/action tables;
#    conservation was gated in 74a before these tables exist.
#  - County joins: demand counties are fips (mbr_county_cd + mbr_state_cd);
#    provider/action counties are name + state; the bridge is
#    ms_ref_county, the same pattern as 69 A2 / 74a A7 (rule 12 respected -
#    state always rides along in keys and columns).
#  - Sensitivity columns come from the same frozen tables (G_MINUS2 /
#    G_PLUS2 CELL rollups); nothing on any tab depends on scenario choice
#    except those two columns, so rankings cannot shift within the file.
# Reviewer 2 SPEC:
#  - Nine tabs as specified; banner on every tab via tab_top(); caveats on
#    every tab surface; limitations parsed verbatim from the methodology
#    doc with a hard count check (15). Deviations = six ASSUMPTION blocks
#    (A1 rank definition, A2 chart/list caps, A4 six-stage mapping).
#  - House style: helpers copied from 56_final_report.py; openpyxl only;
#    output under 08_capacity_risk/outputs/.
# Reviewer 3 EFFICIENCY:
#  - Zero claims scans (R1 trivially satisfied); ~12 BQ reads, all
#    aggregates except cap_provider_year/cap_willing (roster tab) and
#    cap_action_lists. Providers tab uses the plain writer to keep file
#    size and build time down. Deep-dive loops are pandas group lookups,
#    not per-county queries.
