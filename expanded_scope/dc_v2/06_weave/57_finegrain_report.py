"""
57 - fine-grain predictions report   [PYTHON / pandas + openpyxl]

WHAT  : Predictions at the lowest honest grain, each carrying an
        expected-error column measured from actual past performance (2025
        validation months the models never trained on) - not claimed,
        measured. Error fallback: exact cell -> county -> error band
        (counties grouped by their own measured error level: A <= 0.25,
        B <= 0.50, C otherwise or unmeasured), recorded in error_basis.
        House style copied from 13_build_report.py.
GRAIN : demand at mbr_county_cd x specialty_ctg_cd; capacity at
        prvdr_county x specialty_ctg_cd; provider tab at epdb_dw_prvdr_id x
        specialty_ctg_cd (top 5,000 by provider_pred_next_12m).
INPUTS: dc2_demand_predictions, dc2_capacity_predictions,
        dc2_capacity_provider_future, dc2_capacity_provider
OUTPUT: medicare_demand_capacity_finegrain.xlsx (repo root)
Run   : python expanded_scope/dc_v2/06_weave/57_finegrain_report.py
"""

import os
import sys


def _expanded_scope_dir():
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.dirname(os.path.dirname(here))
    except NameError:
        probe = os.getcwd()
        for _ in range(6):
            if os.path.isfile(os.path.join(probe, "config.py")):
                return probe
            cand = os.path.join(probe, "expanded_scope")
            if os.path.isfile(os.path.join(cand, "config.py")):
                return cand
            probe = os.path.dirname(probe)
        raise FileNotFoundError(
            "config.py not found - run from the repo root or any folder inside it")


sys.path.insert(0, _expanded_scope_dir())
import config as cfg

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_XLSX = cfg.repo_path("medicare_demand_capacity_finegrain.xlsx")

DEM_PRED = cfg.src("dc2_demand_predictions")
CAP_PRED = cfg.src("dc2_capacity_predictions")
PROV_FUT = cfg.src("dc2_capacity_provider_future")
CAP_PROV = cfg.src("dc2_capacity_provider")

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"


# ---------- styling helpers (from 13_build_report.py) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def title(ws, text, sub=None, ncols=8):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        cell(ws, "A2", sub, italic=True, color=DARK_BLUE, size=10, bg=LIGHT_BLUE)
        ws.row_dimensions[2].height = 18


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def _int(v, d=0):
    try:
        return d if v is None or pd.isna(v) else int(float(v))
    except (TypeError, ValueError):
        return d


def _float(v, d=0.0):
    try:
        return d if v is None or pd.isna(v) else float(v)
    except (TypeError, ValueError):
        return d


def derived_table(ws, df, cols, r0, filters=True):
    """cols = (df_key, derivation_sentence, width, number_format, align);
    headers are the column names verbatim, derivation row above."""
    for i, (key, derivation, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}{r0}", derivation, italic=True, size=8, color=DARK_GREY,
             bg=GREY, h_align="center", bdr=True)
        cell(ws, f"{col}{r0 + 1}", key, bold=True, color=WHITE, bg=DARK_BLUE, size=9,
             h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[r0].height = 42
    ws.row_dimensions[r0 + 1].height = 24
    hdr = r0 + 1
    for ridx, (_, row) in enumerate(df.iterrows(), start=hdr + 1):
        bg = GREY if ridx % 2 == 0 else WHITE
        for i, (key, _, _, num, align) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if num and v is not None and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif v is None or pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i + 1)}{ridx}", v, bg=bg, size=9, bdr=True,
                 num=num, h_align=align)
    if filters:
        ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{hdr + len(df)}"
    ws.freeze_panes = f"A{hdr + 1}"
    return hdr


AS_STORED = "as stored"


def error_band(mape):
    if pd.isna(mape):
        return "C"
    if mape <= 0.25:
        return "A"
    if mape <= 0.50:
        return "B"
    return "C"


def build_error_tables(val, county_col, pred_col, actual_col):
    """cell / county / error-band error stats from validation rows."""
    val = val.copy()
    val["abs_err"] = (val[pred_col] - val[actual_col]).abs()
    val["ape"] = np.where(val[actual_col] >= 10,
                          val["abs_err"] / val[actual_col], np.nan)
    keys = [county_col, "specialty_ctg_cd"]
    cell_stats = val.groupby(keys, as_index=False).agg(
        mae_cell=("abs_err", "mean"),
        mape_cell=("ape", "mean"),
        n_q_cell=("ape", "count"))
    county_stats = val.groupby(county_col, as_index=False).agg(
        mae_county=("abs_err", "mean"),
        mape_county=("ape", "mean"),
        n_q_county=("ape", "count"))
    county_stats["error_band"] = county_stats["mape_county"].map(error_band)
    band_stats = (county_stats[county_stats["mape_county"].notna()]
                  .groupby("error_band", as_index=False)
                  .agg(mae_band=("mae_county", "mean"),
                       mape_band=("mape_county", "mean")))
    return cell_stats, county_stats, band_stats


def attach_errors(fut, county_col, cell_stats, county_stats, band_stats):
    keys = [county_col, "specialty_ctg_cd"]
    df = fut.merge(cell_stats, on=keys, how="left")
    df = df.merge(county_stats, on=county_col, how="left")
    df["error_band"] = df["error_band"].fillna("C")
    df = df.merge(band_stats, on="error_band", how="left")
    use_cell = df["n_q_cell"].fillna(0) > 0
    use_county = ~use_cell & (df["n_q_county"].fillna(0) > 0)
    df["error_basis"] = np.where(use_cell, "cell",
                                 np.where(use_county, "county", "band"))
    df["mape_hist"] = np.where(use_cell, df["mape_cell"],
                               np.where(use_county, df["mape_county"], df["mape_band"]))
    df["mae_hist"] = np.where(use_cell, df["mae_cell"],
                              np.where(use_county, df["mae_county"], df["mae_band"]))
    return df


def _normalize_provider_keys(df, name):
    raw = df["epdb_dw_prvdr_id"].copy()
    df["epdb_dw_prvdr_id"] = pd.to_numeric(df["epdb_dw_prvdr_id"], errors="coerce").astype("Int64")
    n_null = int(df["epdb_dw_prvdr_id"].isna().sum())
    print(f"{name}: epdb_dw_prvdr_id null after cast: {n_null} (must be 0)")
    if n_null:
        print(f"{name}: sample raw values before cast: "
              f"{raw[df['epdb_dw_prvdr_id'].isna()].head(5).tolist()}")
    df["specialty_ctg_cd"] = df["specialty_ctg_cd"].astype(str)
    return df


def load():
    client = cfg.client()
    q = lambda sql: client.query(sql).result().to_dataframe()
    d = {}
    d["dem_val"] = q(f"SELECT mbr_county_cd, specialty_ctg_cd, actual_next_1m, "
                     f"pred_next_1m_xgb FROM `{DEM_PRED}` "
                     f"WHERE split_label = 'validation' AND actual_next_1m IS NOT NULL")
    d["dem_fut"] = q(f"SELECT mbr_county_cd, specialty_ctg_cd, pred_next_1m_xgb, "
                     f"pred_next_12m_xgb FROM `{DEM_PRED}` WHERE split_label = 'future' "
                     f"ORDER BY mbr_county_cd, specialty_ctg_cd")
    d["cap_val"] = q(f"SELECT prvdr_county, specialty_ctg_cd, actual_next_1m, "
                     f"bottom_up_next_1m FROM `{CAP_PRED}` "
                     f"WHERE split_label = 'validation' AND actual_next_1m IS NOT NULL "
                     f"AND bottom_up_next_1m IS NOT NULL")
    d["cap_fut"] = q(f"SELECT prvdr_county, specialty_ctg_cd, bottom_up_next_1m, "
                     f"bottom_up_next_12m FROM `{CAP_PRED}` WHERE split_label = 'future' "
                     f"ORDER BY prvdr_county, specialty_ctg_cd")
    d["prov_fut"] = q(f"SELECT epdb_dw_prvdr_id, specialty_ctg_cd, prvdr_county, "
                      f"provider_pred_next_1m, provider_pred_next_12m FROM `{PROV_FUT}`")
    d["prov_fut"] = _normalize_provider_keys(d["prov_fut"], "dc2_capacity_provider_future")
    d["prov_inputs"] = q(f"""
        SELECT a.epdb_dw_prvdr_id, a.specialty_ctg_cd,
               a.panel_members, a.pct_new_patients, a.tenure_months,
               b.visits_2025
        FROM (SELECT epdb_dw_prvdr_id, specialty_ctg_cd, panel_members,
                     pct_new_patients, tenure_months
              FROM `{CAP_PROV}` WHERE month = DATE '2025-12-01') a
        LEFT JOIN (SELECT epdb_dw_prvdr_id, specialty_ctg_cd, SUM(visits) AS visits_2025
                   FROM `{CAP_PROV}` WHERE year = 2025 GROUP BY 1, 2) b
        ON a.epdb_dw_prvdr_id = b.epdb_dw_prvdr_id
        AND a.specialty_ctg_cd = b.specialty_ctg_cd""")
    d["prov_inputs"] = _normalize_provider_keys(d["prov_inputs"], "dc2_capacity_provider")
    return d


DEM_COLS = [
    ("mbr_county_cd", AS_STORED, 12, None, "left"),
    ("specialty_ctg_cd", AS_STORED, 12, None, "left"),
    ("pred_next_1m_xgb", "model estimate of visits needed in January 2026", 15, "#,##0", "right"),
    ("pred_next_12m_xgb", "model estimate of visits needed in calendar 2026", 15, "#,##0", "right"),
    ("mae_hist", "average size of this model's miss on 2025 months it did not train on, in visits",
     13, "#,##0.0", "right"),
    ("mape_hist", "average percent miss of this same model on 2025 months it did not train on, "
     "for this county and specialty", 13, "0.0%", "right"),
    ("error_basis", "where the error estimate comes from: this exact cell, this county overall, "
     "or counties with similar error levels", 11, None, "center"),
]

CAP_COLS = [
    ("prvdr_county", AS_STORED, 16, None, "left"),
    ("specialty_ctg_cd", AS_STORED, 12, None, "left"),
    ("bottom_up_next_1m", "summed per-provider model estimates for January 2026", 15, "#,##0", "right"),
    ("bottom_up_next_12m", "summed per-provider model estimates for calendar 2026", 15, "#,##0", "right"),
    ("mae_hist", "average size of this model's miss on 2025 months it did not train on, in visits",
     13, "#,##0.0", "right"),
    ("mape_hist", "average percent miss of this same model on 2025 months it did not train on, "
     "for this county and specialty", 13, "0.0%", "right"),
    ("error_basis", "where the error estimate comes from: this exact cell, this county overall, "
     "or counties with similar error levels", 11, None, "center"),
]

PROV_COLS = [
    ("epdb_dw_prvdr_id", AS_STORED, 14, None, "left"),
    ("specialty_ctg_cd", AS_STORED, 12, None, "left"),
    ("prvdr_county", AS_STORED, 16, None, "left"),
    ("provider_pred_next_1m", "model estimate of visits this provider will deliver next month",
     15, "#,##0", "right"),
    ("provider_pred_next_12m", "model estimate of visits this provider will deliver in calendar 2026",
     15, "#,##0", "right"),
    ("panel_members", "distinct members seen in the trailing 12 months (December 2025)",
     12, "#,##0", "right"),
    ("pct_new_patients", "share of December 2025 visits from members new to this provider",
     12, "0.0%", "right"),
    ("tenure_months", "months since the provider's first claim in the data", 11, "#,##0", "right"),
    ("visits_2025", "actual visits delivered in 2025", 11, "#,##0", "right"),
    ("mape_hist", "county-level percent miss of the capacity model where available; "
     "provider-level error was not individually validated", 13, "0.0%", "right"),
    ("error_basis", "county where the county's own error was measured, otherwise band; "
     "provider-level error was not individually validated", 11, None, "center"),
]


def main():
    d = load()

    dem_cell, dem_cnty, dem_band = build_error_tables(
        d["dem_val"], "mbr_county_cd", "pred_next_1m_xgb", "actual_next_1m")
    dem = attach_errors(d["dem_fut"], "mbr_county_cd", dem_cell, dem_cnty, dem_band)

    cap_cell, cap_cnty, cap_band = build_error_tables(
        d["cap_val"], "prvdr_county", "bottom_up_next_1m", "actual_next_1m")
    cap = attach_errors(d["cap_fut"], "prvdr_county", cap_cell, cap_cnty, cap_band)

    prov = d["prov_fut"].merge(d["prov_inputs"],
                               on=["epdb_dw_prvdr_id", "specialty_ctg_cd"], how="left")
    prov = prov.merge(cap_cnty[["prvdr_county", "mape_county", "n_q_county", "error_band"]],
                      on="prvdr_county", how="left")
    prov["error_band"] = prov["error_band"].fillna("C")
    prov = prov.merge(cap_band, on="error_band", how="left")
    has_cnty = prov["n_q_county"].fillna(0) > 0
    prov["mape_hist"] = np.where(has_cnty, prov["mape_county"], prov["mape_band"])
    prov["error_basis"] = np.where(has_cnty, "county", "band")
    prov = (prov.sort_values("provider_pred_next_12m", ascending=False)
            .head(5000).reset_index(drop=True))

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Demand by County-Specialty")
    title(ws, "Demand by County-Specialty",
          "2026 predictions with measured historical error", ncols=len(DEM_COLS))
    derived_table(ws, dem[[k for k, *_ in DEM_COLS]], DEM_COLS, 4)

    ws = wb.create_sheet("Capacity by County-Specialty")
    title(ws, "Capacity by County-Specialty",
          "2026 predictions with measured historical error", ncols=len(CAP_COLS))
    derived_table(ws, cap[[k for k, *_ in CAP_COLS]], CAP_COLS, 4)

    ws = wb.create_sheet("Provider Level")
    title(ws, "Provider Level",
          "2026 provider predictions; capped at the top 5,000 providers by "
          "provider_pred_next_12m", ncols=len(PROV_COLS))
    derived_table(ws, prov[[k for k, *_ in PROV_COLS]], PROV_COLS, 4)

    ws = wb.create_sheet("How to read the error columns")
    for col, w in {"A": 3, "B": 28, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "How to read the error columns", "measured on months the model never saw")
    r = 4
    r = kv(ws, r, "mae_hist",
           "The average size of the model's miss, in visits. If mae_hist is 50, past "
           "estimates for cells like this one were off by about 50 visits per month on "
           "average.", h=40)
    r = kv(ws, r, "mape_hist",
           "The same miss as a share of the true number. If mape_hist is 0.20, past "
           "estimates were off by about 20 percent on average.", h=32)
    r = kv(ws, r, "Measured, not claimed",
           "Both numbers come from comparing the model's estimates against real 2025 months "
           "the model never trained on. They are the model's actual track record, not a "
           "promise.", h=40)
    r = kv(ws, r, "Thin cells",
           "Thin cells carry county- or band-level error because their own history is too "
           "thin to measure an error rate from; band means counties with similar measured "
           "error levels.", h=40)
    big = dem.sort_values("pred_next_12m_xgb", ascending=False).head(1)
    if len(big):
        b = big.iloc[0]
        pred = float(b["pred_next_12m_xgb"])
        mape = float(b["mape_hist"]) if pd.notna(b["mape_hist"]) else 0.0
        lo, hi = pred * (1 - mape), pred * (1 + mape)
        r = kv(ws, r, "Worked example",
               f"Largest cell: county {b['mbr_county_cd']}, specialty "
               f"{b['specialty_ctg_cd']}. 2026 estimate {pred:,.0f} visits; historical "
               f"percent miss {mape:.1%} ({b['error_basis']} basis); so the plausible range "
               f"is roughly {lo:,.0f} to {hi:,.0f} visits.", h=52)
    ws.freeze_panes = "A3"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"Demand by County-Specialty: {len(dem):,} rows")
    print(f"Capacity by County-Specialty: {len(cap):,} rows")
    print(f"Provider Level: {len(prov):,} rows")
    print("demand error_basis counts:")
    print(dem["error_basis"].value_counts().to_string())
    print("capacity error_basis counts:")
    print(cap["error_basis"].value_counts().to_string())
    print("provider error_basis counts:")
    print(prov["error_basis"].value_counts().to_string())


if __name__ == "__main__":
    main()
