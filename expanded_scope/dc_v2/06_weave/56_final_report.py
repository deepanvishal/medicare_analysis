"""
56 - master report   [PYTHON / pandas + openpyxl]

WHAT  : The dc_v2 decision workbook, built from dc2_weave plus the model
        input tables. House style copied from 13_build_report.py (Arial,
        DARK_BLUE/MID_BLUE/LIGHT_BLUE scheme, cell/fill/thin helpers).
        Every data tab uses the actual table column names as headers, with
        a derivation row (italic, small, grey) above the header explaining
        each column in one plain sentence; stored columns get "as stored".
        Scope: CP and ME members aged 60 and above. Predictions are for
        calendar 2026 from models trained on 2024-2025. Plan-type detail
        lives in the v1 compliance report; this report is county x
        specialty.
GRAIN : Gap 2026 at county_fips x cms_specialty; input tabs per county;
        summaries per specialty ranking in the Answers tab.
INPUTS: dc2_weave, dc2_demand_base, dc2_demand_chronic,
        dc2_capacity_provider, cfg.base("ref_specialty_crosswalk")
OUTPUT: medicare_demand_capacity_dc2.xlsx (repo root)
Run   : python expanded_scope/dc_v2/06_weave/56_final_report.py
"""

import os
import sys


def _expanded_scope_dir():
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.dirname(os.path.dirname(here))
    except NameError:
        probe = os.getcwd()
        for _ in range(6):
            if os.path.isfile(os.path.join(probe, "config.py")):
                return probe
            cand = os.path.join(probe, "expanded_scope")
            if os.path.isfile(os.path.join(cand, "config.py")):
                return cand
            probe = os.path.dirname(probe)
        raise FileNotFoundError(
            "config.py not found - run from the repo root or any folder inside it")


sys.path.insert(0, _expanded_scope_dir())
import config as cfg

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT_XLSX = cfg.repo_path("medicare_demand_capacity_dc2.xlsx")

WEAVE    = cfg.src("dc2_weave")
DEM_BASE = cfg.src("dc2_demand_base")
DEM_CHR  = cfg.src("dc2_demand_chronic")
CAP_PROV = cfg.src("dc2_capacity_provider")
XWALK    = cfg.base("ref_specialty_crosswalk")

DARK_BLUE, MID_BLUE, LIGHT_BLUE = "1F3864", "2E75B6", "D6E4F0"
GREY, DARK_GREY, WHITE = "F2F2F2", "595959", "FFFFFF"
LIGHT_GREEN, LIGHT_RED = "E2EFDA", "FFE0E0"
LIGHT_GOLD = "FFF2CC"
COMPLIANCE_RED = "F8CBAD"


# ---------- styling helpers (from 13_build_report.py) ----------
def fill(hx):
    return PatternFill("solid", fgColor=hx)


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def cell(ws, ref, value, bold=False, color="000000", bg=None, size=10,
         h_align="left", wrap=True, bdr=False, italic=False, num=None):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin()
    if num:
        c.number_format = num
    return c


def section_header(ws, row, c0, c1, text, bg=MID_BLUE):
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell(ws, f"{get_column_letter(c0)}{row}", text, bold=True, color=WHITE, bg=bg, size=11)
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws, row, label, value, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=GREY, bdr=True)
    ws.merge_cells(f"D{row}:H{row}")
    cell(ws, f"D{row}", value, size=10, bg=WHITE, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1


def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def title(ws, text, sub=None, ncols=8):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    cell(ws, "A1", text, bold=True, color=WHITE, bg=DARK_BLUE, size=16)
    ws.row_dimensions[1].height = 34
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        cell(ws, "A2", sub, italic=True, color=DARK_BLUE, size=10, bg=LIGHT_BLUE)
        ws.row_dimensions[2].height = 18


def _int(v, d=0):
    try:
        return d if v is None or pd.isna(v) else int(float(v))
    except (TypeError, ValueError):
        return d


def _float(v, d=0.0):
    try:
        return d if v is None or pd.isna(v) else float(v)
    except (TypeError, ValueError):
        return d


def derived_table(ws, df, cols, r0, filters=True):
    """cols = (df_key, derivation_sentence, width, number_format, align);
    headers are the table column names verbatim; derivation row above."""
    for i, (key, derivation, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}{r0}", derivation, italic=True, size=8, color=DARK_GREY,
             bg=GREY, h_align="center", bdr=True)
        cell(ws, f"{col}{r0 + 1}", key, bold=True, color=WHITE, bg=DARK_BLUE, size=9,
             h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[r0].height = 48
    ws.row_dimensions[r0 + 1].height = 24
    hdr = r0 + 1
    for ridx, (_, row) in enumerate(df.iterrows(), start=hdr + 1):
        bg = GREY if ridx % 2 == 0 else WHITE
        for i, (key, _, _, num, align) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if num and v is not None and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif v is None or pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i + 1)}{ridx}", v, bg=bg, size=9, bdr=True,
                 num=num, h_align=align)
    if filters:
        ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(cols))}{hdr + len(df)}"
    ws.freeze_panes = f"A{hdr + 1}"
    return hdr


AS_STORED = "as stored"


# ---------- data ----------
def load():
    client = cfg.client()
    q = lambda sql: client.query(sql).result().to_dataframe()
    d = {}
    d["weave"] = q(f"SELECT * FROM `{WEAVE}` ORDER BY state_cd, county_fips, cms_specialty")

    d["dem_dec"] = q(f"""
        SELECT mbr_county_cd, MAX(members) AS members,
               MAX(mbr_age_60_64) AS mbr_age_60_64, MAX(mbr_age_65_74) AS mbr_age_65_74,
               MAX(mbr_age_75_84) AS mbr_age_75_84, MAX(mbr_age_85p) AS mbr_age_85p
        FROM `{DEM_BASE}` WHERE month = DATE '2025-12-01' GROUP BY 1""")
    d["dem_yr"] = q(f"""
        SELECT mbr_county_cd, SUM(visits) AS visits,
               AVG(pct_new_patients) AS pct_new_patients
        FROM `{DEM_BASE}` WHERE year = 2025 GROUP BY 1""")
    d["top5_prev"] = q(f"""
        WITH top5 AS (
          SELECT HCC_v24 FROM `{DEM_CHR}` GROUP BY 1
          ORDER BY AVG(prevalence) DESC LIMIT 5
        )
        SELECT mbr_county_cd, HCC_v24, prevalence
        FROM `{DEM_CHR}`
        WHERE month = DATE '2025-12-01' AND HCC_v24 IN (SELECT HCC_v24 FROM top5)""")

    d["cap_inputs"] = q(f"""
        SELECT prvdr_county,
               COUNT(DISTINCT IF(year = 2025, epdb_dw_prvdr_id, NULL)) AS providers,
               SUM(IF(month = DATE '2025-12-01', panel_members, 0)) AS panel_members,
               SUM(IF(month = DATE '2025-12-01', panel_60_64, 0)) AS panel_60_64,
               SUM(IF(month = DATE '2025-12-01', panel_65_74, 0)) AS panel_65_74,
               SUM(IF(month = DATE '2025-12-01', panel_75_84, 0)) AS panel_75_84,
               SUM(IF(month = DATE '2025-12-01', panel_85p, 0)) AS panel_85p,
               SUM(IF(month = DATE '2025-12-01', panel_chronic_members, 0)) AS panel_chronic_members,
               AVG(IF(year = 2025, pct_new_patients, NULL)) AS pct_new_patients,
               APPROX_QUANTILES(IF(month = DATE '2025-12-01', tenure_months, NULL), 100)[OFFSET(50)] AS tenure_months,
               SUM(IF(year = 2025, visits, 0)) AS visits
        FROM `{CAP_PROV}` GROUP BY 1 ORDER BY visits DESC""")

    top_cell = q(f"SELECT mbr_county_cd, specialty_ctg_cd FROM `{DEM_BASE}` "
                 f"WHERE year = 2024 GROUP BY 1, 2 ORDER BY SUM(visits) DESC LIMIT 1")
    d["we_county"] = str(top_cell["mbr_county_cd"].iloc[0])
    d["we_spec"] = str(top_cell["specialty_ctg_cd"].iloc[0])
    d["we_month_row"] = q(f"SELECT month, visits, pct_new_patients FROM `{DEM_BASE}` "
                          f"WHERE mbr_county_cd = '{d['we_county']}' "
                          f"AND specialty_ctg_cd = '{d['we_spec']}' AND year = 2024 "
                          f"ORDER BY month DESC LIMIT 1")
    d["we_prev_row"] = q(f"SELECT HCC_v24, members_with_hcc, members, prevalence "
                         f"FROM `{DEM_CHR}` WHERE mbr_county_cd = '{d['we_county']}' "
                         f"AND month = DATE '2024-12-01' ORDER BY prevalence DESC LIMIT 1")
    we_fips = d["we_county"].zfill(5)
    d["we_gap_row"] = q(f"SELECT county_fips, cms_specialty, demand_next_12m_xgb, "
                        f"capacity_next_12m_bottom_up, gap_model_2026 FROM `{WEAVE}` "
                        f"WHERE county_fips = '{we_fips}' AND cms_specialty IN "
                        f"(SELECT cms_specialty FROM `{XWALK}` WHERE aetna_cd = '{d['we_spec']}') "
                        f"AND gap_model_2026 IS NOT NULL "
                        f"ORDER BY ABS(gap_model_2026) DESC LIMIT 1")
    if d["we_gap_row"].empty:
        d["we_gap_row"] = q(f"SELECT county_fips, cms_specialty, demand_next_12m_xgb, "
                            f"capacity_next_12m_bottom_up, gap_model_2026 FROM `{WEAVE}` "
                            f"WHERE county_fips = '{we_fips}' "
                            f"AND gap_model_2026 IS NOT NULL "
                            f"ORDER BY ABS(gap_model_2026) DESC LIMIT 1")
    return d


# ---------- tab 1: Overview ----------
def build_overview(wb):
    ws = wb.create_sheet("Overview")
    for col, w in {"A": 3, "B": 26, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Medicare Demand vs Capacity - dc_v2",
          "Modeled 2026 estimates | commercial (CP) and Medicare (ME) members aged 60+")
    r = 4
    r = section_header(ws, r, 2, 8, "WHAT THIS WORKBOOK IS")
    r = kv(ws, r, "In plain words",
           "For every county and medical specialty in our footprint, this workbook estimates "
           "how many visits our members aged 60 and above will need in 2026 and how many "
           "visits our contracted providers will deliver to them. The difference is the gap. "
           "The estimates come from models that learned from our own 2024-2025 visit history, "
           "member counts, ages, and health conditions. The current method's numbers are "
           "shown next to the model's so the two can be compared.", h=90)
    r = kv(ws, r, "Scope",
           "CP and ME members aged 60 and above. Predictions are for calendar 2026, from "
           "models trained on 2024-2025 history.", h=28)
    r = kv(ws, r, "Plan types",
           "Plan-type detail lives in the v1 compliance report; this report is county x "
           "specialty.", h=26)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "THREE CAVEATS - READ BEFORE THE NUMBERS")
    r = kv(ws, r, "Caveat 1",
           "Capacity counts only visits delivered to our members, not the provider's whole "
           "practice.", h=26)
    r = kv(ws, r, "Caveat 2",
           "Condition measures use each visit's main diagnosis only, so they undercount.", h=26)
    r = kv(ws, r, "Caveat 3",
           "Rows with expected_error_band C should be read at submarket or state rollup, "
           "not individually.", h=26)
    r = blank(ws, r)
    r = kv(ws, r, "Compliance and the gap",
           "Compliance columns reflect the CMS minimum-provider rules; the gap columns show "
           "whether the network actually has enough visit capacity - a county can be "
           "compliant and still short.", h=40)
    ws.freeze_panes = "A3"
    return 0


# ---------- tab: Coverage ----------
def build_coverage(wb, weave):
    ws = wb.create_sheet("Coverage")
    total = len(weave)
    rows = [{"column": c,
             "populated_count": int(weave[c].notna().sum()),
             "populated_pct": (float(weave[c].notna().sum()) / total if total else 0.0)}
            for c in weave.columns]
    df = pd.DataFrame(rows)
    cols = [("column", "dc2_weave column name, verbatim", 32, None, "left"),
            ("populated_count", "rows where this column has a value", 16, "#,##0", "right"),
            ("populated_pct", "share of all rows with a value; low = a join failed upstream",
             14, "0.0%", "right")]
    title(ws, "Coverage", f"join diagnosis: {total:,} total rows in dc2_weave",
          ncols=len(cols))
    hdr = derived_table(ws, df, cols, 4)
    pct_col = get_column_letter(3)
    rng = f"{pct_col}{hdr + 1}:{pct_col}{hdr + len(df)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.5"],
                                                  fill=fill(LIGHT_RED), stopIfTrue=True))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.9"],
                                                  fill=fill(LIGHT_GOLD)))
    return len(df)


# ---------- tab 2: Gap 2026 (master) ----------
GAP_COLS = [
    ("state_cd", AS_STORED, 8, None, "left"),
    ("county_fips", AS_STORED, 11, None, "left"),
    ("county_name", "county name from ref_county, for display", 16, None, "left"),
    ("cms_specialty", AS_STORED, 24, None, "left"),
    ("demand_visits_2025_actual",
     "visits that actually happened in 2025 - context for the prediction beside it; "
     "late-year claims may still be arriving, so this can slightly undercount",
     14, "#,##0", "right"),
    ("capacity_visits_2025_actual",
     "visits that actually happened in 2025 - context for the prediction beside it; "
     "late-year claims may still be arriving, so this can slightly undercount",
     14, "#,##0", "right"),
    ("gap_2025_actual",
     "real 2025 visits: demand minus capacity; cross-county patient flow makes this nonzero",
     13, "#,##0", "right"),
    ("demand_next_12m_xgb",
     "model estimate of visits members in this county will need in 2026",
     14, "#,##0", "right"),
    ("capacity_next_12m_bottom_up",
     "sum of per-provider model estimates of visits providers here will deliver in 2026",
     15, "#,##0", "right"),
    ("gap_model_2026",
     "demand_next_12m_xgb minus capacity_next_12m_bottom_up; positive = shortage",
     13, "#,##0", "right"),
    ("gap_status",
     "UNDER = members need more visits than providers can deliver; OVER = the reverse; "
     "no middle label - read the gap and ratio columns for degree",
     10, None, "center"),
    ("capacity_to_demand_ratio",
     "capacity divided by demand; 1.0 = exactly enough",
     12, "0.00", "right"),
    ("compliance_status",
     "pass or fail against the CMS minimum-provider rules, from the compliance pipeline; "
     "failed if any plan type fails",
     15, None, "center"),
    ("expected_error_pct",
     "the measured average percent miss behind the band",
     11, "0.0%", "right"),
    ("expected_error_band",
     "A = model missed by 25% or less on months it never saw; B = up to 50%; "
     "C = more - roll up before reading",
     10, None, "center"),
    ("capacity_potential_p75",
     "if every provider here delivered like the busiest quarter of their local peers",
     14, "#,##0", "right"),
    ("pct_medicare_age_members",
     "share of the county's members aged 65 and above, December 2025",
     12, "0.0%", "right"),
    ("demand_rate_estimate",
     "old formula: current members times historical visit rates; tied to no calendar year; "
     "context only",
     14, "#,##0", "right"),
    ("market_max_demand",
     "ceiling - every Medicare-eligible in the county, not just our members; "
     "context only, in no gap",
     14, "#,##0", "right"),
]

GAP_BLOCKS = [
    ("ID", 1, 4, WHITE),
    ("ACTUAL", 5, 7, LIGHT_GREEN),
    ("ESTIMATE 2026", 8, 12, LIGHT_BLUE),
    ("COMPLIANCE", 13, 13, COMPLIANCE_RED),
    ("MODEL QUALITY", 14, 15, GREY),
    ("CONTEXT", 16, 19, LIGHT_GOLD),
]


def build_gap(wb, weave):
    ws = wb.create_sheet("Gap 2026")
    title(ws, "Gap 2026", "one row per county x cms_specialty from dc2_weave",
          ncols=len(GAP_COLS))
    block_row = 4
    for tag, c0, c1, bg in GAP_BLOCKS:
        ws.merge_cells(f"{get_column_letter(c0)}{block_row}:{get_column_letter(c1)}{block_row}")
        cell(ws, f"{get_column_letter(c0)}{block_row}", tag, bold=True, size=9,
             bg=bg, h_align="center", bdr=True)
    ws.row_dimensions[block_row].height = 18
    hdr = derived_table(ws, weave, GAP_COLS, 5)
    keys = [k for k, *_ in GAP_COLS]
    gap_col = get_column_letter(keys.index("gap_model_2026") + 1)
    rng = f"{gap_col}{hdr + 1}:{gap_col}{hdr + len(weave)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
                                                  fill=fill(LIGHT_RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
                                                  fill=fill(LIGHT_GREEN)))
    band_col = get_column_letter(keys.index("expected_error_band") + 1)
    brng = f"{band_col}{hdr + 1}:{band_col}{hdr + len(weave)}"
    for value, hx in (("A", LIGHT_GREEN), ("B", LIGHT_GOLD), ("C", LIGHT_RED)):
        ws.conditional_formatting.add(
            brng, CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill(hx)))
    comp_col = get_column_letter(keys.index("compliance_status") + 1)
    crng = f"{comp_col}{hdr + 1}:{comp_col}{hdr + len(weave)}"
    ws.conditional_formatting.add(
        crng, CellIsRule(operator="equal", formula=['"NON-COMPLIANT"'],
                         font=Font(name="Arial", color="C00000")))
    return len(weave)


# ---------- tab 3: Answers ----------
ANSWER_COLS = [
    ("state_cd", 8, None, "left"),
    ("county_fips", 11, None, "left"),
    ("county_name", 16, None, "left"),
    ("cms_specialty", 24, None, "left"),
    ("demand_next_12m_xgb", 14, "#,##0", "right"),
    ("capacity_next_12m_bottom_up", 15, "#,##0", "right"),
    ("gap_2025_actual", 13, "#,##0", "right"),
    ("gap_model_2026", 13, "#,##0", "right"),
    ("capacity_to_demand_ratio", 12, "0.00", "right"),
    ("expected_error_band", 10, None, "center"),
]


def _answer_table(ws, df, r0, cols=None):
    cols = cols or ANSWER_COLS
    for i, (key, w, _, _) in enumerate(cols):
        col = get_column_letter(i + 1)
        cell(ws, f"{col}{r0}", key, bold=True, color=WHITE, bg=DARK_BLUE, size=9,
             h_align="center", bdr=True)
        ws.column_dimensions[col].width = w
    ws.row_dimensions[r0].height = 24
    for ridx, (_, row) in enumerate(df.iterrows(), start=r0 + 1):
        bg = GREY if ridx % 2 == 0 else WHITE
        for i, (key, _, num, align) in enumerate(cols):
            v = row.get(key)
            if hasattr(v, "item"):
                v = v.item()
            if num and v is not None and pd.notna(v):
                v = _float(v) if ("%" in num or "." in num) else _int(v)
            elif v is None or pd.isna(v):
                v = None
            cell(ws, f"{get_column_letter(i + 1)}{ridx}", v, bg=bg, size=9, bdr=True,
                 num=num, h_align=align)
    return r0 + len(df) + 2


def build_answers(wb, weave):
    ws = wb.create_sheet("Answers")
    title(ws, "Answers", "ranked views of the Gap 2026 tab; bands A and B only",
          ncols=len(ANSWER_COLS))
    ab = weave[weave["expected_error_band"].isin(["A", "B"])]
    r = 4

    r = section_header(ws, r, 1, len(ANSWER_COLS), "WHERE ARE WE SHORT?")
    ws.merge_cells(f"A{r}:{get_column_letter(len(ANSWER_COLS))}{r}")
    cell(ws, f"A{r}", "Top 25 by gap_model_2026, largest shortage first. Band C rows are "
                      "excluded here and appear in the master tab.", italic=True, size=9,
         color=DARK_GREY)
    r += 1
    short = ab[ab["gap_model_2026"].notna()].sort_values(
        "gap_model_2026", ascending=False).head(25)
    r = _answer_table(ws, short, r)

    r = section_header(ws, r, 1, len(ANSWER_COLS), "WHERE DO WE HAVE EXCESS?")
    ws.merge_cells(f"A{r}:{get_column_letter(len(ANSWER_COLS))}{r}")
    cell(ws, f"A{r}", "Top 25 by capacity_to_demand_ratio, most excess first. Band C rows "
                      "are excluded here and appear in the master tab.", italic=True, size=9,
         color=DARK_GREY)
    r += 1
    excess = ab[ab["capacity_to_demand_ratio"].notna()].sort_values(
        "capacity_to_demand_ratio", ascending=False).head(25)
    r = _answer_table(ws, excess, r)

    r = section_header(ws, r, 1, len(ANSWER_COLS), "WATCH LIST")
    ws.merge_cells(f"A{r}:{get_column_letter(len(ANSWER_COLS))}{r}")
    cell(ws, f"A{r}", "Rows where gap_2025_actual and gap_model_2026 disagree in sign: what "
                      "really happened last year and what the model expects point in "
                      "different directions - these need human eyes. Band C rows are "
                      "excluded here and appear in the master tab.",
         italic=True, size=9, color=DARK_GREY)
    r += 1
    watch = ab[ab["gap_model_2026"].notna() & ab["gap_2025_actual"].notna()
               & (np.sign(ab["gap_model_2026"]) != np.sign(ab["gap_2025_actual"]))]
    r = _answer_table(ws, watch, r)

    r = section_header(ws, r, 1, len(ANSWER_COLS), "NON-COMPLIANT AND SHORT")
    ws.merge_cells(f"A{r}:{get_column_letter(len(ANSWER_COLS))}{r}")
    cell(ws, f"A{r}", "Rows failing the CMS minimum-provider rules AND showing a modeled "
                      "2026 shortage, ranked by gap_model_2026. All error bands shown - "
                      "check the band column before trusting an individual number.",
         italic=True, size=9, color=DARK_GREY)
    r += 1
    nc_short = weave[(weave["compliance_status"] == "NON-COMPLIANT")
                     & weave["gap_model_2026"].notna()
                     & (weave["gap_model_2026"] > 0)].sort_values(
        "gap_model_2026", ascending=False)
    r = _answer_table(ws, nc_short, r)
    ws.freeze_panes = "A4"
    return len(short) + len(excess) + len(watch) + len(nc_short)


# ---------- tab 4: Demand Inputs ----------
def build_demand_inputs(wb, d):
    ws = wb.create_sheet("Demand Inputs")
    prev = d["top5_prev"].copy()
    prev["hcc_col"] = prev["HCC_v24"].map(
        lambda c: f"prev_hcc_{int(float(c))}" if str(c).replace('.', '').isdigit()
        else f"prev_hcc_{str(c).strip()}")
    prev_wide = prev.pivot_table(index="mbr_county_cd", columns="hcc_col",
                                 values="prevalence", aggfunc="max").reset_index()
    df = (d["dem_dec"].merge(d["dem_yr"], on="mbr_county_cd", how="outer")
          .merge(prev_wide, on="mbr_county_cd", how="left")
          .sort_values("visits", ascending=False).reset_index(drop=True))
    prev_cols = sorted(c for c in df.columns if c.startswith("prev_hcc_"))
    cols = [
        ("mbr_county_cd", AS_STORED, 12, None, "left"),
        ("members", "distinct members aged 60+ in the county, December 2025", 11, "#,##0", "right"),
        ("mbr_age_60_64", "members 60-64, December 2025", 11, "#,##0", "right"),
        ("mbr_age_65_74", "members 65-74, December 2025", 11, "#,##0", "right"),
        ("mbr_age_75_84", "members 75-84, December 2025", 11, "#,##0", "right"),
        ("mbr_age_85p", "members 85 and over, December 2025", 11, "#,##0", "right"),
        ("pct_new_patients",
         "share of 2025 visits where the member had not seen that provider in the previous 12 months",
         13, "0.0%", "right"),
        ("visits", "total 2025 visits by members of this county", 12, "#,##0", "right"),
    ] + [(c, "share of the county's members with this condition in the trailing 24 months "
             "(December 2025)", 12, "0.0%", "right") for c in prev_cols]
    title(ws, "Demand Inputs", "what the demand model consumed, rolled up per county",
          ncols=len(cols))
    derived_table(ws, df, cols, 4)
    return len(df)


# ---------- tab 5: Capacity Inputs ----------
def build_capacity_inputs(wb, d):
    ws = wb.create_sheet("Capacity Inputs")
    df = d["cap_inputs"]
    cols = [
        ("prvdr_county", AS_STORED, 16, None, "left"),
        ("providers", "distinct providers with 2025 visits in this county", 11, "#,##0", "right"),
        ("panel_members", "sum of provider panel sizes, December 2025", 13, "#,##0", "right"),
        ("panel_60_64", "panel members 60-64, December 2025", 11, "#,##0", "right"),
        ("panel_65_74", "panel members 65-74, December 2025", 11, "#,##0", "right"),
        ("panel_75_84", "panel members 75-84, December 2025", 11, "#,##0", "right"),
        ("panel_85p", "panel members 85 and over, December 2025", 11, "#,##0", "right"),
        ("panel_chronic_members",
         "panel members with at least one chronic condition, December 2025",
         14, "#,##0", "right"),
        ("pct_new_patients",
         "share of 2025 visits where the member had not seen that provider in the previous 12 months",
         13, "0.0%", "right"),
        ("tenure_months", "median months since the provider's first claim, December 2025",
         12, "#,##0", "right"),
        ("visits", "total 2025 visits delivered in this county", 12, "#,##0", "right"),
    ]
    title(ws, "Capacity Inputs", "what the capacity model consumed, rolled up per county",
          ncols=len(cols))
    derived_table(ws, df, cols, 4)
    return len(df)


# ---------- tab 6: Worked Examples ----------
def _example(ws, r, heading, caption, rows):
    r = section_header(ws, r, 2, 8, heading)
    ws.merge_cells(f"B{r}:H{r}")
    cell(ws, f"B{r}", caption, italic=True, size=9, color=DARK_GREY)
    r += 1
    for label, value in rows:
        r = kv(ws, r, label, value)
    return blank(ws, r)


def build_worked_examples(wb, d):
    ws = wb.create_sheet("Worked Examples")
    for col, w in {"A": 3, "B": 30, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Worked Examples",
          f"largest county x specialty cell by 2024 visits: county {d['we_county']}, "
          f"specialty {d['we_spec']} - real numbers from the tables")
    r = 4
    n = 0

    if len(d["we_month_row"]):
        m = d["we_month_row"].iloc[0]
        visits = float(m["visits"])
        pct = float(m["pct_new_patients"]) if pd.notna(m["pct_new_patients"]) else 0.0
        new_visits = round(pct * visits)
        r = _example(ws, r, "EXAMPLE A - one month's pct_new_patients",
                     f"For {m['month']}, the share of visits made by members who had not seen "
                     f"that provider in the previous 12 months.",
                     [("numerator (new visits)", f"{new_visits:,.0f}"),
                      ("denominator (total visits)", f"{visits:,.0f}"),
                      ("result (pct_new_patients)",
                       f"{new_visits:,.0f} / {visits:,.0f} = {pct:.4f}")])
        n += 1

    if len(d["we_prev_row"]):
        p = d["we_prev_row"].iloc[0]
        r = _example(ws, r, "EXAMPLE B - one prev_hcc value",
                     f"For December 2024, condition {p['HCC_v24']}: the share of the county's "
                     f"members with that condition in the trailing 24 months.",
                     [("numerator (members with the condition)",
                       f"{float(p['members_with_hcc']):,.0f}"),
                      ("denominator (county members)", f"{float(p['members']):,.0f}"),
                      ("result (prevalence)",
                       f"{float(p['members_with_hcc']):,.0f} / {float(p['members']):,.0f} "
                       f"= {float(p['prevalence']):.4f}")])
        n += 1

    if len(d["we_gap_row"]):
        g = d["we_gap_row"].iloc[0]
        dem = float(g["demand_next_12m_xgb"]) if pd.notna(g["demand_next_12m_xgb"]) else 0.0
        cap = float(g["capacity_next_12m_bottom_up"]) if pd.notna(
            g["capacity_next_12m_bottom_up"]) else 0.0
        gap = float(g["gap_model_2026"]) if pd.notna(g["gap_model_2026"]) else dem - cap
        r = _example(ws, r, "EXAMPLE C - how gap_model_2026 was computed",
                     f"County {g['county_fips']}, specialty {g['cms_specialty']}: demand "
                     f"estimate minus capacity estimate; positive = shortage.",
                     [("numerator (demand_next_12m_xgb)", f"{dem:,.0f}"),
                      ("denominator (capacity_next_12m_bottom_up)", f"{cap:,.0f}"),
                      ("result (gap_model_2026)", f"{dem:,.0f} - {cap:,.0f} = {gap:,.0f}")])
        n += 1
    ws.freeze_panes = "A3"
    return n


# ---------- tab 7: Data Dictionary ----------
DICT_ROWS = [
    ("state_cd", "state code of the county", "ref_county via normalization", "stored"),
    ("county_fips", "county code; the join key of the table (all sides normalized through ref_county)", "ref_county", "derived"),
    ("county_name", "county name, for display", "ref_county", "stored"),
    ("cms_specialty", "CMS specialty name after the one-time bridge from claims specialty codes", "ref_specialty_crosswalk", "stored"),
    ("demand_visits_2025_actual", "visits that actually happened in 2025, member-county view; late-year claims may still be arriving", "dc2_demand_base", "derived"),
    ("capacity_visits_2025_actual", "visits that actually happened in 2025, provider-county view; late-year claims may still be arriving", "dc2_capacity_county", "derived"),
    ("gap_2025_actual", "real 2025 visits: demand minus capacity; cross-county patient flow makes this nonzero", "computed in 55_weave.py", "derived"),
    ("demand_next_12m_xgb", "model estimate of visits the county's members will need in 2026", "dc2_demand_predictions (future rows)", "derived"),
    ("capacity_next_12m_bottom_up", "sum of per-provider model estimates of visits delivered in 2026", "dc2_capacity_predictions (future rows)", "derived"),
    ("gap_model_2026", "demand_next_12m_xgb minus capacity_next_12m_bottom_up; positive = shortage", "computed in 55_weave.py", "derived"),
    ("gap_status", "UNDER where gap_model_2026 is positive, OVER otherwise; empty where a side is missing", "computed in 55_weave.py", "derived"),
    ("capacity_to_demand_ratio", "capacity divided by demand; 1.0 = exactly enough", "computed in 55_weave.py", "derived"),
    ("compliance_status", "pass or fail against the CMS minimum-provider rules; failed if any plan type fails", "fact_gap_analysis (collapsed over plan types)", "derived"),
    ("expected_error_pct", "measured average percent miss of the demand model on unseen 2025 months, per county", "dc2_demand_predictions (validation rows)", "derived"),
    ("expected_error_band", "A at 25% or less, B up to 50%, C above or unmeasured", "computed in 55_weave.py", "derived"),
    ("capacity_potential_p75", "75th percentile of local per-provider 2026 estimates times the provider count", "dc2_capacity_provider_future", "derived"),
    ("pct_medicare_age_members", "December 2025 members aged 65+ over all members", "A870800_medicare_analysis_membership", "derived"),
    ("demand_rate_estimate", "old formula: current members times historical visit rates; tied to no calendar year; context only", "dc2_baselines (renamed from demand_current_book)", "stored"),
    ("market_max_demand", "eligibles-based demand ceiling, context only, in no gap", "dc2_baselines (from the v1 gap table)", "stored"),
]


def build_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")
    dd = pd.DataFrame(DICT_ROWS, columns=["column", "meaning", "source", "derived_or_stored"])
    cols = [("column", "column name, verbatim", 28, None, "left"),
            ("meaning", "plain-words meaning", 66, None, "left"),
            ("source", "source table", 40, None, "left"),
            ("derived_or_stored", "derived here or stored upstream", 16, None, "center")]
    title(ws, "Data Dictionary", "every dc2_weave column", ncols=len(cols))
    derived_table(ws, dd, cols, 4)
    return len(dd)


# ---------- tab 8: Methodology ----------
def build_methodology(wb):
    ws = wb.create_sheet("Methodology")
    for col, w in {"A": 3, "B": 30, "C": 8, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = w
    title(ws, "Methodology", "plain words, one paragraph per step")
    r = 4
    r = section_header(ws, r, 2, 8, "THE STEPS")
    r = kv(ws, r, "The demand model",
           "One model, trained on every county and specialty together, using 2024-2025 "
           "history: how many members aged 60 and above lived in each county, their ages, "
           "how common each health condition was, the share of first-time patients, and the "
           "calendar. It produces an estimate of visits needed per county and specialty for "
           "2026.", h=64)
    r = kv(ws, r, "The capacity model",
           "One model, trained on every provider together, using each provider's own "
           "history: how many members they saw, the ages and conditions of those members, "
           "how many were new to them, and how long the provider has been in our data. Its "
           "per-provider 2026 estimates are added up to the county level.", h=64)
    r = kv(ws, r, "The old-formula column",
           "demand_rate_estimate carries the earlier method's demand estimate (current "
           "members times historical visit rates, tied to no calendar year). "
           "market_max_demand is the eligibles-based ceiling. Both are context only and "
           "enter no gap.", h=52)
    r = kv(ws, r, "The gap and the ratio",
           "gap_model_2026 = demand estimate minus capacity estimate for the same county and "
           "specialty; positive means members are expected to need more visits than the "
           "network delivers. capacity_to_demand_ratio is the same comparison as a ratio; "
           "1.0 means exactly enough.", h=52)
    r = kv(ws, r, "expected_error_band",
           "Measured on 2025 months the model never trained on: the average percent miss per "
           "county. A means 25 percent or less, B up to 50 percent, C more than that or too "
           "little history to measure. C rows should be read only in rollups.", h=52)
    r = kv(ws, r, "The 2025 actuals",
           "demand_visits_2025_actual and capacity_visits_2025_actual are the visits that "
           "really happened in 2025, shown beside the 2026 estimates so every prediction "
           "can be compared with last year. gap_2025_actual is their difference. Late-year "
           "claims may still be arriving, so these can slightly undercount.", h=52)
    r = kv(ws, r, "Compliance and the gap",
           "Compliance columns reflect the CMS minimum-provider rules; the gap columns show "
           "whether the network actually has enough visit capacity - a county can be "
           "compliant and still short.", h=40)
    r = kv(ws, r, "Model quality, measured",
           "The demand annual model missed by about 655 visits per county-specialty on "
           "average on unseen months. The capacity model missed by about 18 percent in "
           "large counties and about 40 percent in mid-size counties; small-county numbers "
           "are unreliable individually. Both models are gradient boosted trees (XGBoost) - "
           "a method that builds many small decision rules and adds them together.", h=76)
    r = blank(ws, r)
    r = section_header(ws, r, 2, 8, "OTHER MODEL DESIGNS CONSIDERED")
    r = kv(ws, r, "Per-level forecasts with reconciliation",
           "Build numbers at zip, county, and state independently, then force them to agree. "
           "Unnecessary once county became the single modeling level.", h=36)
    r = kv(ws, r, "County-level top-down capacity model",
           "Built and tested; lost to the provider-level model in every county size band on "
           "held-out 2025 data, dropped.", h=36)
    r = kv(ws, r, "Cluster-then-model",
           "Group similar counties or providers into 10-20 clusters and fit one small model "
           "per cluster instead of one pooled model. Planned next step if per-segment "
           "explanation is requested, and the route to the cluster-average third estimate "
           "from the original design.", h=52)
    ws.freeze_panes = "A3"
    return 0


# ---------- assemble ----------
def main():
    d = load()
    wb = Workbook()
    wb.remove(wb.active)
    counts = {}
    counts["Overview"] = build_overview(wb)
    counts["Coverage"] = build_coverage(wb, d["weave"])
    counts["Gap 2026"] = build_gap(wb, d["weave"])
    counts["Answers"] = build_answers(wb, d["weave"])
    counts["Demand Inputs"] = build_demand_inputs(wb, d)
    counts["Capacity Inputs"] = build_capacity_inputs(wb, d)
    counts["Worked Examples"] = build_worked_examples(wb, d)
    counts["Data Dictionary"] = build_dictionary(wb)
    counts["Methodology"] = build_methodology(wb)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    for tab, n in counts.items():
        print(f"  {tab}: {n} rows")


if __name__ == "__main__":
    main()
