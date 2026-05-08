from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── COLORS ──────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
GREY         = "F2F2F2"
DARK_GREY    = "595959"
WHITE        = "FFFFFF"
YELLOW       = "FFF2CC"
LIGHT_YELLOW = "FFFDE7"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, ref, value, bold=False, color="000000", bg=None,
         size=10, h_align="left", wrap=True, bdr=False, italic=False):
    c = ws[ref]
    c.value = value
    c.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if bdr:
        c.border = thin_border()
    return c

def section_header(ws, row, col_start, col_end, text):
    start = get_column_letter(col_start)
    end   = get_column_letter(col_end)
    ws.merge_cells(f"{start}{row}:{end}{row}")
    cell(ws, f"{start}{row}", text,
         bold=True, color=WHITE, bg=MID_BLUE, size=11, h_align="left")
    ws.row_dimensions[row].height = 20
    return row + 1

def kv(ws, row, label, value, ncols=4, label_bg=GREY, value_bg=WHITE, h=18):
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", label, bold=True, size=10, bg=label_bg, bdr=True)
    ws.merge_cells(f"D{row}:G{row}")
    cell(ws, f"D{row}", value, size=10, bg=value_bg, bdr=True, wrap=True)
    ws.row_dimensions[row].height = h
    return row + 1

def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1


def build_tab1(wb):
    ws = wb.create_sheet("1. Project Overview")
    ws.sheet_view.showGridLines = False

    # col widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20

    # ── TITLE ──────────────────────────────────────────────
    ws.merge_cells("B1:G1")
    cell(ws, "B1", "Medicare Supply Demand",
         bold=True, color=WHITE, bg=DARK_BLUE, size=18, h_align="center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("B2:G2")
    cell(ws, "B2", "Network Adequacy & Capacity Modeling  |  Florida Medicare Advantage  |  Plan Year 2026",
         color="AAAAAA", bg=DARK_BLUE, size=10, h_align="center", italic=True)
    ws.row_dimensions[2].height = 20

    row = 4

    # ── OBJECTIVE ──────────────────────────────────────────
    row = section_header(ws, row, 2, 7, "  PROJECT OBJECTIVE")
    row = blank(ws, row)
    ws.merge_cells(f"B{row}:G{row}")
    cell(ws, f"B{row}",
         "Build analytic models to determine whether the Aetna Medicare Advantage provider network "
         "has the right capacity, specialties, and geographic distribution — and to identify where "
         "to add, remove, or reconfigure providers under CMS regulatory requirements.",
         size=10, bg=WHITE, wrap=True)
    ws.row_dimensions[row].height = 40
    row += 2

    # ── DELIVERABLES ───────────────────────────────────────
    row = section_header(ws, row, 2, 7, "  DELIVERABLES")
    row = blank(ws, row)
    for label, value in [
        ("Compliance Table",
         "County × Specialty × Plan Type → COMPLIANT / NON-COMPLIANT per 42 CFR 422.116"),
        ("Access Coverage",
         "% of Medicare beneficiaries per county with at least 1 contracted provider within CMS distance threshold"),
        ("Provider Gap",
         "Contracted provider count vs CMS-required minimum per county per specialty"),
        ("Counties at Risk",
         "Counties failing access % or provider count standard — prioritized by gap size"),
        ("Bed Count Compliance",
         "Acute Inpatient Hospital contracted beds vs CMS required beds per county"),
    ]:
        row = kv(ws, row, f"  • {label}", value)
    row = blank(ws, row)

    # ── SCOPE ──────────────────────────────────────────────
    row = section_header(ws, row, 2, 7, "  SCOPE")
    row = blank(ws, row)
    for label, value in [
        ("Geography",       "Florida — 67 member counties evaluated for compliance"),
        ("Plan Types",      "MA-HMO, MA-PPO"),
        ("CMS Specialties", "43 provider and facility specialty types per 42 CFR 422.116"),
        ("Regulatory Year", "CMS 2026 HSD Reference File (published December 17, 2025)"),
        ("Data Snapshot",   "Most recent available month — CMS MA penetration file"),
    ]:
        row = kv(ws, row, f"  {label}", value)
    row = blank(ws, row)

    # ── GEOGRAPHY & DISTANCE ───────────────────────────────
    row = section_header(ws, row, 2, 7, "  HOW GEOGRAPHY & DISTANCE WORKS")
    row = blank(ws, row)
    for label, value in [
        ("Member Side",
         "All 67 Florida counties. Population sourced at zip code level from ACS 2018 Census. "
         "Compliance is evaluated at the MEMBER county level — not provider county."),
        ("Provider Side",
         "41 of 67 Florida counties have contracted Aetna providers. "
         "26 counties have zero contracted providers and are automatically non-compliant. "
         "Provider location is determined using the zip code centroid (geographic center of the zip)."),
        ("Distance Method",
         "Straight-line distance measured from member zip centroid to provider zip centroid "
         "using BigQuery ST_DISTANCE function, converted from meters to miles. "
         "CMS uses drive time — straight-line is an approximation."),
        ("Threshold",
         "CMS specifies a maximum distance per specialty per county type. "
         "A provider counts toward a member zip ONLY if the distance is within this threshold. "
         "Threshold uses the MEMBER county type — not the provider county type."),
        ("Cross-County Access",
         "A provider in one county can count toward compliance in a neighboring member county "
         "if their zip centroid is within the CMS distance threshold. "
         "Compliance is always measured from the member's perspective."),
        ("Rollup to County",
         "After identifying which member zips have access, population is rolled up to the "
         "member county. % Members With Access = population in zips with access / total county population."),
        ("Zip Uncertainty",
         "A confidence band is calculated using zip radius = SQRT(area_sq_miles / PI()). "
         "Distance lower/upper bound = measured distance ± (member zip radius + provider zip radius). "
         "Borderline cases near the threshold are flagged separately."),
    ]:
        row = kv(ws, row, f"  {label}", value, h=40)
    row = blank(ws, row)

    # ── V2 APPROACH ────────────────────────────────────────
    row = section_header(ws, row, 2, 7, "  V2 APPROACH — SPECIALTY MAPPING")
    row = blank(ws, row)
    for label, value in [
        ("Method",
         "Uses specialty_cd (raw specialty code from RPDB_RPNPRAC network table) "
         "mapped to CMS specialties via Global Lookup Table. "
         "One provider can map to multiple CMS specialties."),
        ("Multi-Specialty",
         "Provider network IDs are exploded from the network_id field in the provider file. "
         "Each provider's full specialty list is retrieved — not just the primary specialty category."),
        ("Difference from V1",
         "V1 used specialty_ctg_cd (primary specialty category code only — single code per provider). "
         "V2 uses specialty_cd (all specialty codes per provider via network join — broader coverage)."),
    ]:
        row = kv(ws, row, f"  {label}", value, h=35)
    row = blank(ws, row)

    # ── SPECIALTY MAPPING TABLE ────────────────────────────
    row = section_header(ws, row, 2, 7, "  CMS SPECIALTY → AETNA CODE MAPPING (43 Specialties)")
    row = blank(ws, row)

    # embedded specialty mapping — code - description per CMS specialty
    SPECIALTY_MAPPING = {
        'Primary Care': '10101 - General Practice, 10201 - Family Practice, 10202 - Geriatric Medicine/Family Prac, 10301 - Internal Medicine, 10308 - Geriatric Medicine/Internal Me, 10336 - Internal Medicine Hospice, 2FP - Family Practice, 2GP - General Practice, 2I - Internal Medicine, 2IM - Internal Medicine, 91151 - Sleep Medicine-Family Practice, 91209 - Config-Primary Care, 91210 - Config-Primary Care Attestatio',
        'Allergy and Immunology': '10326 - Allergy/Immunology, 10501 - Allergy & Immunology, 10603 - Dermatological Immunology/Diag, 2A - Allergy, 2AIM - Immunology, 2ENA - Otolaryngology/Allergy, 90003 - Allergy, 90335 - Immunology, 90386 - Otolaryngology/Allergy',
        'Cardiology': '10302 - Cardiac Electrophysiology, 10303 - Cardiovascular Disease, 10322 - Cardiology, 10332 - Interventional Cardiology, 10339 - Cardiology Adv Heart Failure/, 2C - Cardiology, 2CEP - Cardiac Electrophysiology, 2CI - Cardiology (Invasive), 2CS - Cardiothoracic/Cardiovascular, 40312 - Nuclear Cardiology, 90313 - Cardiology (Invasive), 91046 - Cardiac Valve Replacement, 91205 - Cardiac Monitoring Service',
        'Chiropractor': '2CH - Chiropractics, 91146 - Chiropractics, DC - Chiropractor',
        'Clinical Psychology': '2NPH - Neuropsychology, 2PHGR - Geriatric Psychology, 91018 - Psychological Testing, 91029 - Neuropsych Testing, CP - Clinical Psychologist, NPS - Neuropsychologist',
        'Clinical Social Work': '2MLS - Social Worker Masters Licensed, 2MUS - Social Worker(Masters w/o Lic), 2PYSW - Psychiatric Social Worker, 90371 - Psychiatric Social Worker, 91207 - Certified Social Work, SW - Clinical Social Worker',
        'Dermatology': '10601 - Dermatology, 10602 - Dermatopathology/Dermatology, 2D - Dermatology, 2DP - Dermatopathology, 40207 - Dermatopathology/Pathology',
        'Endocrinology': '10306 - Endocrinology Diabetes & Meta, 10319 - Endocrinology, 20105 - Endocrinology Reproductive, 2E - Endocrinology, 91059 - Endocrine Surgery',
        'ENT/Otolaryngology': '14601 - Neurotology, 2EN - Otolaryngology, 2ENA - Otolaryngology/Allergy, 2ENHN - Otolaryngology(Head&Neck) Su, 2ENN - Neuro-Otology, 30601 - Otolaryngology, 30603 - Otorhinolaryngology & Oro-Faci, 30604 - Otorhinolaryngology/Plastic Su, 30605 - Otology, 30607 - Otorhinolaryngology, 30608 - Otology/Neurotology, 30806 - Surgery Head & Neck, 90386 - Otolaryngology/Allergy, 91061 - ENT Trauma, 91078 - Otolaryngology (ENT) Cancer Su, 91117 - Sleep Medicine (Otolaryngology)',
        'Gastroenterology': '10307 - Gastroenterology, 2G - Gastroenterology, 91045 - Capsule Endoscopy, 91060 - Endoscopic Ultrasound, 91063 - Endoscopic Retrograde Cholangi, 91065 - Esophageal Motility Disorders',
        'General Surgery': '2S - Surgery (General), 30201 - Surgery, Colon & Rectal, 30804 - Surgery General Vascular, 30809 - Surgery Hand',
        'Gynecology OB/GYN': '20104 - Maternal & Fetal Medicine, 20106 - Gynecology, 20107 - Perinatology, 20108 - Obstetrics & Gynecology - CA P, 20109 - Obstetrics/Gynecology Hospice, 20110 - Female Pelvic Medicine & Recon, 20191 - Obstetrics & Gynecology, 2NPOG - Nurse practitioner (ob/gyn), 2OG - Ob/Gyn, 2OGOB - Obstetrics, 2OH - Perinatology, 2PAOG - Physicians Assistant Ob/Gyn, 2UGY - Uro-gynecology, 30807 - Surgery Obstetrics & Gynecolo, 90069 - Perinatology/PF, 90355 - Obstetrics, 90398 - Uro-Gynecology',
        'Infectious Diseases': '10310 - Infectious Disease, 2III - Infectious Disease, 91158 - Infectious Disease Focus',
        'Nephrology': '10312 - Nephrology, 2HD - Hemodialysis, 2N - Nephrology, 91217 - Hemodialysis, 91220 - Kidney Transplant Program, DI - Dialysis Center',
        'Neurology': '10334 - Vascular Neurology, 10806 - Neuromuscular Medicine Physica, 11002 - Neurology, 11006 - Neurology & Psychiatry, 11008 - Neurology Chemical, 11014 - Neurology/Psychiatry Hospice, 11015 - Sleep Medicine - Neurology, 11102 - Neuromuscular Medicine Psychia, 11103 - Epilepsy, 2NE - Neurology, 91044 - Botox injections Neurology, 91062 - Epilepsy Surgery, 91081 - Movement Disorders, 91082 - Multiple Sclerosis, 91084 - Neuromuscular Medicine, 91086 - Neurovascular Surgery, 91149 - Sleep Medicine-Neurology',
        'Neurosurgery': '10803 - Spinal Cord Injury Medicine, 2NS - Neurosurgery, 2NSS - Neurosurgery (Spine), 30301 - Surgery, Neurological, 90348 - Neurosurgery (Spine), 91118 - Spinal Cord Stimulation, 91119 - Stereotactic & Functional Neur',
        'Oncology Medical/Surgical': '10309 - Hematology, 10311 - Oncology Medical, 10315 - Hematology/Oncology, 20103 - Oncology Gynecologic, 30810 - Surgery Oncology, 91085 - Neuro-Oncology, 91126 - Urologic Oncology, 91129 - Surgery Carcinoid',
        'Oncology Radiation': '2RO - Radiation Therapy, 40303 - Radiation Oncology, 40304 - Radiological Physics, 40310 - Therapeutic Radiology, 40316 - Radiation Therapy, 40318 - Radium Therapy',
        'Ophthalmology': '2O - Ophthalmology, 2OAG - Anterior Segment (Glaucoma), 2OC - Corneal Specialist, 2PSOC - Oculoplastic Surgery, 2RS - Retinal Specialist, 30401 - Opthalmology, 30402 - Retinal Opthalmology, 30403 - Sleep Medicine-Ophthalmology/O, 90089 - Retinal Specialist, 90311 - Anterior Segment (Glaucoma), 90315 - Corneal Specialist, 90343 - Neuro-Ophthalmology, 90356 - Oculoplastic Surgery, 91070 - Glaucoma Service, 91087 - Ophthamologic Cancer, 91088 - Orbital Surgery',
        'Orthopedic Surgery': '10317 - Oncology Orthopedic, 2OR - Orthopedics, 2ORFA - Orthopedics (Foot & Ankle), 2ORON - Orthopedics (Oncology), 2ORR - Orthopedics (Joint Replacement, 2ORS - Orthopedics Surgery (Spine), 2ORSM - Orthopedics (Sports Medicine), 30501 - Surgery Orthopedic, 30502 - Surgery Hand/Orthopedic, 30503 - Surgery Knee, 90361 - Orthopedics (Foot & Ankle), 90362 - Orthopedics (Joint Replacement, 90365 - Orthopedics Surgery (Spine), 91092 - Orthopedic Elbow Replacement, 91093 - Orthopedic Trauma, 91094 - Orthopedic Shoulder',
        'Physiatry Rehabilitative Med': '10801 - Physical Medicine & Rehabilita, 10802 - Rehabilitation Medicine, 10805 - Physical Medicine Hospice and, 2PM - Physical Medicine, 2RM - Rehab Medicine',
        'Plastic Surgery': '2PS - Plastic Surgery, 2PSCF - Craniofacial Surgery, 2PSOC - Oculoplastic Surgery, 30602 - Surgery Oro-Facial Plastic, 90308 - Facial Plastic and Reconstruct, 90316 - Craniofacial Surgery, 90356 - Oculoplastic Surgery, 91054 - Craniofacial Plastics, 91111 - Reconstructive Breast Surgery, 91112 - Reconstructive Breast Surgery, 91113 - Reconstructive Breast Surgery',
        'Podiatry': '2PO - Podiatry, 91213 - Foot and Ankle Surgery, 91214 - Foot Surgery, DP - Podiatrist',
        'Psychiatry': '11001 - Psychiatry, 11005 - Psychiatry Geriatric, 11006 - Neurology & Psychiatry, 11007 - Addictionology, 11011 - Addiction Psychiatry, 11014 - Neurology/Psychiatry Hospice, 11101 - Psychomatic Medicine, 2PY - Psychiatry, 2PYGR - Geriatric Psychiatry, 91244 - Psychiatry Autism Spectrum, 91250 - Psychiatry Home Based Services, 91252 - Psychiatry Trauma/Crisis',
        'Pulmonology': '10313 - Pulmonary Disease, 10318 - Medical Diseases of Chest, 2PD - Pulmonary Disease, 91139 - Sleep Medicine - Pulmonology',
        'Rheumatology': '10314 - Rheumatology, 2RH - Rheumatology, 91041 - Arthritis Reconstruction',
        'Urology': '2U - Urology, 2UGY - Uro-gynecology, 2UMI - Urology (Male Infertility), 30808 - Surgery Urological, 31001 - Urology, 90379 - Urology (Male Infertility), 90398 - Uro-Gynecology, 91126 - Urologic Oncology, 91127 - UROLOGICTR',
        'Vascular Surgery': '2IY - Peripheral Vascular Disease, 2VS - Vascular Surgery, 40317 - Vascular & Interventional Radi, 40319 - Angiography and Interventional, 90071 - Peripheral Vascular Disease',
        'Cardiothoracic Surgery': '2CS - Cardiothoracic/Cardiovascular, 2TS - Thoracic Surgery, 30805 - Surgery Thoracic Cardiovascul, 30812 - Surgery Congenital Cardiac/Th, 30901 - Surgery Thoracic, 91206 - Cardiac Surgery Program, 91215 - Heart Transplant Program',
        'Acute Inpatient Hospitals': '2HSLT - Hospitalist, 2SH - Specialty Hospital, 91002 - Hospitalist, HO - Acute Short Term Hospital, HSLT - Hospitalist, LHO - Long Term Acute Care Hospital',
        'Cardiac Surgery Program': '2CS - Cardiothoracic/Cardiovascular, 91046 - Cardiac Valve Replacement, 91205 - Cardiac Monitoring Service, 91206 - Cardiac Surgery Program',
        'Cardiac Catheterization': '10302 - Cardiac Electrophysiology, 10332 - Interventional Cardiology, 2CEP - Cardiac Electrophysiology, 91205 - Cardiac Monitoring Service',
        'Critical Care ICU': '10304 - Critical Care Medicine, 11016 - Neurocritical Care, 20102 - Critical Care Medicine/Obstetr, 2CCM - Critical Care Medicine, 30102 - Critical Care Medicine/Anesthe, 30302 - Critical Care Medicine Neurolo, 30803 - Surgery Critical care, 91125 - Trauma Surgical Critical Care, 91165 - Intensive Care Coordination',
        'Surgical Services ASC': '2FS - Free Standing Surgical Unit, 91235 - Outpatient Surgery, AC - Ambulatory Surgicenter, FEC - Freestanding Emergency Center',
        'Skilled Nursing Facility': '2SNF - Skilled Nursing Facility, 91287 - Assisted Living Center, 91294 - Skilled Nursing Facilities, 91301 - Nursing Facility Transition Di, 91302 - Recuperative Care, ALC - Assisted Living Center, LSS - Long-Term Services and Support, SK - Skilled Nursing Facility',
        'Diagnostic Radiology': '40306 - Diagnostic Roentgenology, 40311 - Neuroradiology, 40313 - Nuclear Imaging and Therapy, 40315 - Diagnostic Ultrasound, 40320 - Body Imaging, 91224 - Medical Imaging, RFA - Radiology Center',
        'Mammography': '91223 - Mammography',
        'Physical Therapy': '2HR - Hand Rehabilitation, 2PT - Physical Therapy, 90331 - Hand Rehabilitation, 91141 - Physical Therapy',
        'Occupational Therapy': '2TO - Occupational Therapy, 91142 - Occupational Therapy',
        'Speech Therapy': '2TT - Speech Therapy, 91143 - Speech Therapy, 91257 - Speech/Hearing, 91258 - Speech/Hearing Therapy, 91259 - Speech/Language/Hearing Therap, SH - Speech Pathologist, ST - Speech Therapist',
        'Inpatient Psychiatric': '91003 - Psychotic Disorders, RTF - Residential Treatment Facility',
        'Outpatient Infusion/Chemo': '2IC - Infusion Center, 91180 - Antibiotic Infusion, 91218 - Home Infusion Therapy for HIV, 91234 - Outpatient Infusion/Chemothera, HI - Home Infusion, IC - Infusion Center',
        'Outpatient Behavioral Health': '10204 - Addiction Medicine, 11007 - Addictionology, 11011 - Addiction Psychiatry, 2AC - Addictions Counselor, 2MH - Mental Health-Substance Abuse, 90001 - Addictions Counselor, 90428 - Mental Health, 91005 - Dialectic Behavioral Therapy, 91006 - Cognitive Behavioral Therapy, 91011 - Substance Abuse Professional, 91012 - Crisis Intervention, 91032 - Applied Behavioral Analysis, 91134 - Behavioral Health Rehabilitati, 91174 - Mobile Crisis Intervention (MC, 91278 - Applied Behavioral Analysis (A, ABA - Applied Behavioral Analysis, BHR - Behavioral Health Rehabilitati, CAC - Certified Addictions Counselor, CMC - Community Mental Health Center, MH - Mental Health - Substance Abus, SA - Substance Abuse Facility',
    }
    cms_order = [
        "Primary Care","Allergy and Immunology","Cardiology","Chiropractor",
        "Clinical Psychology","Clinical Social Work","Dermatology","Endocrinology",
        "ENT/Otolaryngology","Gastroenterology","General Surgery","Gynecology OB/GYN",
        "Infectious Diseases","Nephrology","Neurology","Neurosurgery",
        "Oncology Medical/Surgical","Oncology Radiation","Ophthalmology",
        "Orthopedic Surgery","Physiatry Rehabilitative Med","Plastic Surgery",
        "Podiatry","Psychiatry","Pulmonology","Rheumatology","Urology",
        "Vascular Surgery","Cardiothoracic Surgery","Acute Inpatient Hospitals",
        "Cardiac Surgery Program","Cardiac Catheterization","Critical Care ICU",
        "Surgical Services ASC","Skilled Nursing Facility","Diagnostic Radiology",
        "Mammography","Physical Therapy","Occupational Therapy","Speech Therapy",
        "Inpatient Psychiatric","Outpatient Infusion/Chemo","Outpatient Behavioral Health",
    ]
    lookup = SPECIALTY_MAPPING

    # column headers
    ws.merge_cells(f"B{row}:C{row}")
    cell(ws, f"B{row}", "CMS Specialty",
         bold=True, color=WHITE, bg=DARK_GREY, size=9, h_align="center", bdr=True)
    ws.merge_cells(f"D{row}:G{row}")
    cell(ws, f"D{row}", "Specialty Code - Description (comma separated)",
         bold=True, color=WHITE, bg=DARK_GREY, size=9, h_align="center", bdr=True)
    ws.row_dimensions[row].height = 16
    row += 1

    for i, cms in enumerate(cms_order):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        codes = lookup.get(cms, "No mapping found")

        ws.merge_cells(f"B{row}:C{row}")
        cell(ws, f"B{row}", cms, bold=True, size=9, bg=bg, bdr=True, wrap=False)

        ws.merge_cells(f"D{row}:G{row}")
        cell(ws, f"D{row}", codes, size=9, bg=bg, bdr=True, wrap=True)

        # auto height based on content length
        estimated_lines = max(2, len(codes) // 120 + 1)
        ws.row_dimensions[row].height = estimated_lines * 14
        row += 1

    row = blank(ws, row)

    # ── ASSUMPTIONS ────────────────────────────────────────
    row = section_header(ws, row, 2, 7, "  KEY ASSUMPTIONS & DATA DECISIONS")
    row = blank(ws, row)
    for label, value in [
        ("Required Provider Count",
         "Sourced directly from CMS 2026 HSD Reference File. "
         "Uses 95th percentile MA plan enrollment — not total Medicare eligibles."),
        ("Compliance Threshold",
         "90% for Large Metro and Metro counties. 85% for Micro, Rural, CEAC. "
         "Per 42 CFR 422.116(d)(4)."),
        ("Facility Minimum Count",
         "13 facility specialty types require minimum 1 per county (flat). "
         "Per 42 CFR 422.116(e)(2)(iii)."),
        ("Acute Inpatient Beds",
         "Required = CEIL(12.2 × beneficiaries_required_to_cover / 1,000). "
         "Measured in contracted BEDS not hospital count. Source: hosp_list_cmi."),
        ("Population Data",
         "ACS 2018 5-year estimates at zip code level. "
         "2020 zip-level data not available in BigQuery public data at time of analysis."),
        ("Distance Limitation",
         "Straight-line distance used. CMS uses drive time. "
         "Rural counties most affected — actual drive distances will be longer."),
        ("Telehealth Credit",
         "NOT applied. 42 CFR 422.116(d)(5) allows 10% credit for 14 specialties. "
         "No telehealth flag available in provider data."),
        ("Plan Type Independence",
         "MA-HMO and MA-PPO evaluated separately. "
         "A provider in MA-HMO does not count toward MA-PPO compliance."),
    ]:
        row = kv(ws, row, f"  {label}", value, label_bg=LIGHT_YELLOW, h=30)

    return ws


# ── TAB 2: COMPLIANCE REPORT ──────────────────────────────────

def build_tab2(wb, df):
    ws = wb.create_sheet("2. Compliance Report")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # col widths
    col_widths = {
        "A": 20, "B": 14, "C": 28, "D": 10,
        "E": 16, "F": 14, "G": 22, "H": 12,
        "I": 16, "J": 14, "K": 16,
        "L": 18, "M": 20, "N": 18,
        "O": 18, "P": 16,
        "Q": 14, "R": 14, "S": 14, "T": 16,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── ROW 1: TITLE ─────────────────────────────────────────
    ws.merge_cells("A1:T1")
    cell(ws, "A1", "Medicare Supply Demand — Compliance Report",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # ── ROW 2: COLOR BAND LABELS ──────────────────────────────
    for rng, text, bg in [
        ("A2:D2",  "  IDENTIFIERS",                                      DARK_GREY),
        ("E2:K2",  "  CMS RULES  (42 CFR 422.116 + HSD Reference File)", MID_BLUE),
        ("L2:N2",  "  POPULATION DATA  (Census ACS 2018)",               "C55A11"),
        ("O2:P2",  "  AETNA NETWORK ACCESS",                             "C55A11"),
        ("Q2:T2",  "  COMPLIANCE RESULTS",                               DARK_BLUE),
    ]:
        ws.merge_cells(rng)
        cell(ws, rng.split(":")[0], text,
             bold=True, color=WHITE, bg=bg, size=9, h_align="left")
    ws.row_dimensions[2].height = 16

    # ── ROW 3: CALLOUTS ───────────────────────────────────────
    callouts = {
        "A3": "", "B3": "", "C3": "", "D3": "",
        "E3": "Source: CMS MA State/County Penetration file",
        "F3": "CMS published annually per county type. Same for all specialties.",
        "G3": "95th_pct_ratio × total_medicare_beneficiaries",
        "H3": "From 422.116 Table 2. Varies by specialty and county type.",
        "I3": "CEIL(min_ratio × beneficiaries_required / 1,000). From HSD file.",
        "J3": "90% Large Metro/Metro | 85% Micro/Rural/CEAC",
        "K3": "From 422.116 Table 1. Applied at member county type.",
        "L3": "ACS 2018 zip population rolled to county. All ages, all insurance.",
        "M3": "SUM(zip_population WHERE has_access = TRUE per zip)",
        "N3": "population_with_access / total_county_population",
        "O3": "COUNT(DISTINCT provider_id) within max_distance_miles of at least 1 member zip",
        "P3": "SUM(Beds) from hosp_list_cmi. Acute Inpatient only. 0 for all others.",
        "Q3": "required_provider_count − actual_count. Negative = surplus.",
        "R3": "pct_covered >= compliance_threshold",
        "S3": "actual_count >= required_provider_count",
        "T3": "BOTH access AND count standards met",
    }
    for ref, txt in callouts.items():
        cell(ws, ref, txt, size=8, color="666666", bg="F9F9F9",
             italic=True, wrap=True)
    ws.row_dimensions[3].height = 28

    # ── ROW 4: COLUMN HEADERS ────────────────────────────────
    headers = [
        ("A4", "County",                            DARK_GREY),
        ("B4", "County Type",                       DARK_GREY),
        ("C4", "CMS Specialty",                     DARK_GREY),
        ("D4", "Plan Type",                         DARK_GREY),
        ("E4", "Total Medicare\nBeneficiaries",      MID_BLUE),
        ("F4", "95th Pct\nBase Ratio",               MID_BLUE),
        ("G4", "Beneficiaries\nRequired to Cover",   MID_BLUE),
        ("H4", "Min\nRatio",                         MID_BLUE),
        ("I4", "CMS Required\nCount",                MID_BLUE),
        ("J4", "Access\nThreshold",                  MID_BLUE),
        ("K4", "Max Distance\n(Miles)",              MID_BLUE),
        ("L4", "County Population\n(ACS 2018)",      "C55A11"),
        ("M4", "Population\nWith Access",            "C55A11"),
        ("N4", "% Population\nWith Access",          "C55A11"),
        ("O4", "Contracted\nProviders / Beds",       "C55A11"),
        ("P4", "Contracted Beds\n(Inpatient Only)",  "C55A11"),
        ("Q4", "Gap\n(Required - Actual)",           DARK_BLUE),
        ("R4", "Access\nStandard Met",               DARK_BLUE),
        ("S4", "Count\nStandard Met",                DARK_BLUE),
        ("T4", "Compliance\nStatus",                 DARK_BLUE),
    ]
    ws.row_dimensions[4].height = 35
    for ref, label, bg in headers:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=9, h_align="center", bdr=True)

    # ── DATA ROWS ────────────────────────────────────────────
    LIGHT_GREEN  = "E2EFDA"
    LIGHT_RED    = "FFE0E0"
    LIGHT_BLUE_D = "D6E4F0"
    LIGHT_ORANGE = "FCE4D6"

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 5
        is_compliant = str(row.get("compliance_status", "")).strip() == "COMPLIANT"
        row_bg = LIGHT_GREEN if is_compliant else LIGHT_RED

        def v(col):
            val = row.get(col, 0)
            if val is None or (isinstance(val, float) and str(val) == 'nan'):
                return 0
            return val

        access_c = "Yes" if bool(v("access_compliant")) else "No"
        count_c  = "Yes" if bool(v("count_compliant"))  else "No"
        beds = v("total_contracted_beds") or 0

        data = [
            ("A", v("county_name"),                        DARK_GREY,  row_bg),
            ("B", v("county_type"),                        DARK_GREY,  row_bg),
            ("C", v("cms_specialty"),                      DARK_GREY,  row_bg),
            ("D", v("plan_type"),                          DARK_GREY,  row_bg),
            ("E", v("county_total_beneficiaries"),         MID_BLUE,   LIGHT_BLUE_D),
            ("F", v("ratio_95th_percentile"),              MID_BLUE,   LIGHT_BLUE_D),
            ("G", v("beneficiaries_required_to_cover"),    MID_BLUE,   LIGHT_BLUE_D),
            ("H", v("min_ratio_per_1000"),                 MID_BLUE,   LIGHT_BLUE_D),
            ("I", v("required_provider_count"),            MID_BLUE,   LIGHT_BLUE_D),
            ("J", v("compliance_threshold"),               MID_BLUE,   LIGHT_BLUE_D),
            ("K", v("max_distance_miles"),                 MID_BLUE,   LIGHT_BLUE_D),
            ("L", v("total_county_population"),            "C55A11",   LIGHT_ORANGE),
            ("M", v("population_with_access"),             "C55A11",   LIGHT_ORANGE),
            ("N", v("pct_covered"),                        "C55A11",   LIGHT_ORANGE),
            ("O", v("actual_count"),                       "C55A11",   LIGHT_ORANGE),
            ("P", beds,                                    "C55A11",   LIGHT_ORANGE),
            ("Q", v("provider_gap"),                       DARK_BLUE,  row_bg),
            ("R", access_c, DARK_BLUE,
             LIGHT_GREEN if access_c == "Yes" else LIGHT_RED),
            ("S", count_c,  DARK_BLUE,
             LIGHT_GREEN if count_c  == "Yes" else LIGHT_RED),
            ("T", v("compliance_status"), DARK_BLUE,
             LIGHT_GREEN if is_compliant else LIGHT_RED),
        ]

        for col, val, txt_color, bg_color in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", color=txt_color, size=9,
                          bold=(col == "T"))
            c.fill = fill(bg_color)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=False)
            c.border = thin_border()
            if col == "N":
                c.number_format = "0.0%"
            elif col in ("F", "H"):
                c.number_format = "0.0000"
            elif col == "J":
                c.number_format = "0%"

        ws.row_dimensions[r].height = 15

    # note
    note_r = len(df) + 5 + 1
    ws.merge_cells(f"A{note_r}:T{note_r}")
    cell(ws, f"A{note_r}",
         "NOTE: Contracted Beds (col P) populated only for Acute Inpatient Hospitals — 0 for all other specialties. "
         "Gap is negative when actual count exceeds required (surplus). "
         "% Population With Access uses Census ACS zip population — not Medicare beneficiaries. "
         "Compliance Status = COMPLIANT only when BOTH Access Standard AND Count Standard are met.",
         size=8, color="666666", bg="F9F9F9", italic=True, wrap=True)
    ws.row_dimensions[note_r].height = 30

    return ws


from openpyxl.formatting.rule import ColorScaleRule

# ── TAB 3 defined here before MAIN ───────────────────────────

PROJECT        = "anbc-hcb-dev"           # table project
CLIENT_PROJECT = "anbc-dev-prv-nc-ds"     # billing/auth project
DATASET        = "provider_ds_netconf_data_hcb_dev"
PREFIX         = "A870800_medicare_supply_demand"

COMPLIANCE_QUERY = f"""
SELECT
  county_name,
  county_type,
  cms_specialty,
  plan_type,
  COALESCE(county_total_beneficiaries, 0)       AS county_total_beneficiaries,
  COALESCE(ratio_95th_percentile, 0)            AS ratio_95th_percentile,
  COALESCE(beneficiaries_required_to_cover, 0)  AS beneficiaries_required_to_cover,
  COALESCE(min_ratio_per_1000, 0)               AS min_ratio_per_1000,
  COALESCE(required_provider_count, 0)          AS required_provider_count,
  COALESCE(compliance_threshold, 0)             AS compliance_threshold,
  COALESCE(max_distance_miles, 0)               AS max_distance_miles,
  COALESCE(total_county_population, 0)          AS total_county_population,
  COALESCE(population_with_access, 0)           AS population_with_access,
  COALESCE(pct_covered, 0)                      AS pct_covered,
  COALESCE(actual_count, 0)                     AS actual_count,
  COALESCE(total_contracted_beds, 0)            AS total_contracted_beds,
  COALESCE(provider_gap, 0)                     AS provider_gap,
  access_compliant,
  count_compliant,
  compliance_status
FROM `{PROJECT}.{DATASET}.{PREFIX}_fact_gap_analysis_v2`
ORDER BY county_name, cms_specialty, plan_type
"""

SUMMARY_SPECIALTY_QUERY = f"""
SELECT
  cms_specialty,
  plan_type,
  COUNTIF(compliance_status = 'COMPLIANT')     AS compliant_counties,
  COUNTIF(compliance_status = 'NON-COMPLIANT') AS non_compliant_counties,
  COUNT(*)                                      AS total_counties,
  ROUND(
    COUNTIF(compliance_status = 'COMPLIANT') / COUNT(*), 4
  )                                             AS pct_compliant,
  COUNTIF(access_compliant = FALSE)             AS access_failures,
  COUNTIF(count_compliant = FALSE)              AS count_failures
FROM `{PROJECT}.{DATASET}.{PREFIX}_fact_gap_analysis_v2`
GROUP BY cms_specialty, plan_type
ORDER BY pct_compliant ASC, cms_specialty, plan_type
"""


# ── TAB 3: SUMMARY BY PLAN × SPECIALTY ───────────────────────

def build_tab3(wb, df_summary):
    ws = wb.create_sheet("3. Summary by Specialty")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {
        "A": 30,  # cms specialty
        "B": 12,  # plan type
        "C": 16,  # compliant counties
        "D": 20,  # non-compliant counties
        "E": 14,  # total counties
        "F": 14,  # pct compliant
        "G": 16,  # access failures
        "H": 16,  # count failures
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── ROW 1: TITLE ─────────────────────────────────────────
    ws.merge_cells("A1:H1")
    cell(ws, "A1", "Medicare Supply Demand — Specialty Compliance Summary (V2)",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # ── ROW 2: SUBTITLE ──────────────────────────────────────
    ws.merge_cells("A2:H2")
    cell(ws, "A2",
         "Grain: CMS Specialty × Plan Type  |  Each row = county-level pass/fail counts  |  "
         "Sorted by % Compliant ascending (worst performing specialties first)",
         size=9, color="666666", bg="F9F9F9", italic=True, h_align="left")
    ws.row_dimensions[2].height = 18

    # ── ROW 3: CALLOUTS ──────────────────────────────────────
    callouts = {
        "A3": "",
        "B3": "",
        "C3": "Counties where BOTH access % AND count standard are met",
        "D3": "Counties where EITHER access % OR count standard fails",
        "E3": "Total Florida counties evaluated",
        "F3": "compliant_counties / total_counties",
        "G3": "Counties where pct_covered < compliance_threshold",
        "H3": "Counties where actual_count < required_provider_count",
    }
    for ref, txt in callouts.items():
        cell(ws, ref, txt, size=8, color="666666",
             bg="F9F9F9", italic=True, wrap=True)
    ws.row_dimensions[3].height = 28

    # ── ROW 4: HEADERS ───────────────────────────────────────
    headers = [
        ("A4", "CMS Specialty",         DARK_GREY),
        ("B4", "Plan Type",             DARK_GREY),
        ("C4", "Compliant\nCounties",   "375623"),
        ("D4", "Non-Compliant\nCounties","C00000"),
        ("E4", "Total\nCounties",        DARK_BLUE),
        ("F4", "% Compliant",            DARK_BLUE),
        ("G4", "Access\nFailures",       MID_BLUE),
        ("H4", "Count\nFailures",        MID_BLUE),
    ]
    ws.row_dimensions[4].height = 35
    for ref, label, bg in headers:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=10, h_align="center", bdr=True)

    # ── DATA ROWS ────────────────────────────────────────────
    prev_specialty = None
    alt = True

    for i, (_, row) in enumerate(df_summary.iterrows()):
        r = i + 5

        # alternate shade per specialty group
        if row['cms_specialty'] != prev_specialty:
            alt = not alt
            prev_specialty = row['cms_specialty']
        row_bg = GREY if alt else WHITE

        pct = float(row.get('pct_compliant', 0) or 0)

        data = [
            ("A", row.get('cms_specialty', ''),           DARK_GREY,  row_bg),
            ("B", row.get('plan_type', ''),                DARK_GREY,  row_bg),
            ("C", int(row.get('compliant_counties', 0) or 0),    "375623", "E2EFDA"),
            ("D", int(row.get('non_compliant_counties', 0) or 0),"C00000", "FFE0E0"),
            ("E", int(row.get('total_counties', 0) or 0),         DARK_BLUE, row_bg),
            ("F", pct,                                     DARK_BLUE,  row_bg),
            ("G", int(row.get('access_failures', 0) or 0),        MID_BLUE,  LIGHT_BLUE),
            ("H", int(row.get('count_failures', 0) or 0),         MID_BLUE,  LIGHT_BLUE),
        ]

        for col, val, txt_color, bg_color in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", color=txt_color, size=10,
                          bold=(col == "F"))
            c.fill = fill(bg_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            if col == "F":
                c.number_format = "0.0%"

        ws.row_dimensions[r].height = 16

    # ── GRADIENT COLOR SCALE ON % COMPLIANT (col F) ──────────
    last_row = len(df_summary) + 4
    ws.conditional_formatting.add(
        f"F5:F{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="C00000",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB84",
            end_type="num",   end_value=1,     end_color="375623"
        )
    )

    return ws


# ── TAB 4: SUMMARY BY PLAN × COUNTY ──────────────────────────

def build_tab4(wb, df_county):
    ws = wb.create_sheet("4. Summary by County")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    set_col_widths = {
        "A": 22, "B": 14, "C": 12,
        "D": 18, "E": 22, "F": 14,
        "G": 14, "H": 16, "I": 16,
    }
    for col, w in set_col_widths.items():
        ws.column_dimensions[col].width = w

    # title
    ws.merge_cells("A1:I1")
    cell(ws, "A1", "Medicare Supply Demand — County Compliance Summary",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # subtitle
    ws.merge_cells("A2:I2")
    cell(ws, "A2",
         "Grain: County × Plan Type  |  Shows how many of the 43 CMS specialties are compliant per county  |  "
         "Sorted by % Compliant ascending (worst counties first)  |  All 67 Florida counties shown",
         size=9, color="666666", bg="F9F9F9", italic=True, h_align="left")
    ws.row_dimensions[2].height = 18

    # callouts
    callouts = {
        "A3": "", "B3": "", "C3": "",
        "D3": "Specialties where BOTH access % AND count standard met",
        "E3": "Specialties where EITHER access % OR count standard fails",
        "F3": "Total CMS specialties evaluated (43)",
        "G3": "compliant_specialties / total_specialties",
        "H3": "Specialties failing pct_covered >= threshold",
        "I3": "Specialties failing actual_count >= required_count",
    }
    for ref, txt in callouts.items():
        cell(ws, ref, txt, size=8, color="666666",
             bg="F9F9F9", italic=True, wrap=True)
    ws.row_dimensions[3].height = 28

    # headers
    headers = [
        ("A4", "County",                    DARK_GREY),
        ("B4", "County Type",               DARK_GREY),
        ("C4", "Plan Type",                 DARK_GREY),
        ("D4", "Compliant\nSpecialties",    "375623"),
        ("E4", "Non-Compliant\nSpecialties","C00000"),
        ("F4", "Total\nSpecialties",         DARK_BLUE),
        ("G4", "% Compliant",               DARK_BLUE),
        ("H4", "Access\nFailures",           MID_BLUE),
        ("I4", "Count\nFailures",            MID_BLUE),
    ]
    ws.row_dimensions[4].height = 35
    for ref, label, bg in headers:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=10, h_align="center", bdr=True)

    # data rows
    prev_county = None
    alt = True

    for i, (_, row) in enumerate(df_county.iterrows()):
        r = i + 5
        if row['county_name'] != prev_county:
            alt = not alt
            prev_county = row['county_name']
        row_bg = GREY if alt else WHITE
        pct = float(row.get('pct_compliant', 0) or 0)

        data = [
            ("A", row.get('county_name', ''),              DARK_GREY,  row_bg),
            ("B", row.get('county_type', ''),              DARK_GREY,  row_bg),
            ("C", row.get('plan_type', ''),                DARK_GREY,  row_bg),
            ("D", int(row.get('compliant_specialties', 0) or 0),  "375623", "E2EFDA"),
            ("E", int(row.get('non_compliant_specialties', 0) or 0), "C00000", "FFE0E0"),
            ("F", int(row.get('total_specialties', 0) or 0),      DARK_BLUE, row_bg),
            ("G", pct,                                     DARK_BLUE,  row_bg),
            ("H", int(row.get('access_failures', 0) or 0),        MID_BLUE,  LIGHT_BLUE),
            ("I", int(row.get('count_failures', 0) or 0),         MID_BLUE,  LIGHT_BLUE),
        ]
        for col, val, txt_color, bg_color in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", color=txt_color, size=10, bold=(col == "G"))
            c.fill = fill(bg_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            if col == "G":
                c.number_format = "0.0%"
        ws.row_dimensions[r].height = 16

    # gradient on % compliant
    last_row = len(df_county) + 4
    ws.conditional_formatting.add(
        f"G5:G{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="C00000",
            mid_type="num",   mid_value=0.5,   mid_color="FFEB84",
            end_type="num",   end_value=1,     end_color="375623"
        )
    )
    return ws


# ── TAB 5: DATA DICTIONARY ────────────────────────────────────

def build_tab5(wb):
    ws = wb.create_sheet("5. Data Dictionary")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 20

    # title
    ws.merge_cells("B1:F1")
    cell(ws, "B1", "Medicare Supply Demand — Data Dictionary",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # headers
    for ref, label, bg in [
        ("B2", "Column Name",    MID_BLUE),
        ("C2", "Source Table",   MID_BLUE),
        ("D2", "Description",    MID_BLUE),
        ("E2", "Formula / Logic",MID_BLUE),
        ("F2", "Data Type",      MID_BLUE),
    ]:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=10, h_align="center", bdr=True)
    ws.row_dimensions[2].height = 20

    entries = [
        # col name, source, description, formula, type
        ("County",                   "ref_county_classification",  "Florida county name",                                                          "From Census geo_us_boundaries.counties",                   "STRING"),
        ("County Type",              "ref_county_classification",  "CMS county classification per 422.116(c)",                                     "Large Metro / Metro / Micro / Rural / CEAC",               "STRING"),
        ("CMS Specialty",            "ref_specialty_crosswalk_expanded", "One of 43 CMS-defined provider or facility specialty types",              "Per 42 CFR 422.116(b)",                                    "STRING"),
        ("Plan Type",                "stg_providers_multi_specialty","MA plan product type",                                                        "MA-HMO or MA-PPO",                                         "STRING"),
        ("Total Medicare Beneficiaries","ref_hsd_required_counts", "Total Medicare eligible people in the county",                                 "From CMS MA State/County Penetration data",                "INT64"),
        ("Beneficiaries Required to Cover","ref_hsd_required_counts","CMS benchmark enrollment for a large MA plan",                               "95th_percentile_ratio × total_medicare_beneficiaries",     "INT64"),
        ("95th Pct Base Ratio",      "ref_hsd_required_counts",    "Proportion of Medicare benes enrolled in 95th percentile MA plan",             "CMS calculated annually per county type. Published in HSD file.", "FLOAT64"),
        ("CMS Required Count",       "ref_hsd_required_counts",    "Minimum number of providers or beds CMS requires",                             "CEIL(min_ratio × beneficiaries_required_to_cover / 1,000). Flat 1 for most facility types.", "INT64"),
        ("Access Threshold",         "ref_county_classification",  "Minimum % of members that must have access per 422.116(d)(4)",                 "0.90 for Large Metro/Metro. 0.85 for Micro/Rural/CEAC",    "FLOAT64"),
        ("Max Distance (Miles)",     "ref_time_distance",          "CMS maximum allowed distance per specialty per county type",                    "From 422.116 Table 1. Applied at MEMBER county type.",     "FLOAT64"),
        ("County Population (ACS 2018)","stg_beneficiaries",       "Total residential population in the county",                                   "ACS 2018 5-year zip estimates rolled up to county",        "INT64"),
        ("Population With Access",   "fact_zip_access",            "Population in member zips that have at least 1 provider within threshold",     "SUM(zip_population WHERE has_access = TRUE)",              "INT64"),
        ("% Members With Access",    "fact_gap_analysis",          "Share of county population with at least 1 provider within CMS threshold",     "population_with_access / total_county_population",         "FLOAT64"),
        ("Contracted Providers / Beds","fact_gap_analysis",        "Distinct contracted providers within threshold for this county and specialty",  "COUNT(DISTINCT provider_id) where distance <= max_distance_miles and within at least 1 member zip", "INT64"),
        ("Contracted Beds (Inpatient Only)","hosp_list_cmi",       "Sum of contracted inpatient beds for Acute Inpatient Hospitals only",          "SUM(Beds) from hosp_list_cmi. NULL beds excluded. 0 for all other specialties.", "INT64"),
        ("Gap (Required - Actual)",  "fact_gap_analysis",          "Difference between CMS required count and actual contracted count",             "required_provider_count − actual_count. Negative = surplus.", "INT64"),
        ("Access Standard Met",      "fact_gap_analysis",          "Whether county passes CMS Test 1 — access percentage standard",                "pct_covered >= compliance_threshold",                      "BOOLEAN"),
        ("Count Standard Met",       "fact_gap_analysis",          "Whether county passes CMS Test 2 — minimum provider count standard",           "actual_count >= required_provider_count",                  "BOOLEAN"),
        ("Compliance Status",        "fact_gap_analysis",          "Overall compliance result. BOTH tests must pass per 422.116",                  "COMPLIANT if access_compliant AND count_compliant. Otherwise NON-COMPLIANT.", "STRING"),
    ]

    for i, (col_nm, source, desc, formula, dtype) in enumerate(entries):
        r = i + 3
        bg = GREY if i % 2 == 0 else WHITE
        for ref, val, w in [
            (f"B{r}", col_nm,  True),
            (f"C{r}", source,  False),
            (f"D{r}", desc,    False),
            (f"E{r}", formula, False),
            (f"F{r}", dtype,   False),
        ]:
            cell(ws, ref, val, bold=w, size=9, bg=bg, bdr=True, wrap=True)
        ws.row_dimensions[r].height = 30

    return ws


# ── TAB 6: CMS RULES ─────────────────────────────────────────

def build_tab6(wb):
    ws = wb.create_sheet("6. CMS Rules")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    for col, w in {"A": 32, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    cell(ws, "A1", "42 CFR 422.116 — Network Adequacy Time & Distance Standards",
         bold=True, color=WHITE, bg=DARK_BLUE, size=13, h_align="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    cell(ws, "A2",
         "Source: https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-B/part-422/subpart-C/section-422.116  |  "
         "CMS 2026 HSD Reference File: https://www.cms.gov/medicare/health-drug-plans/medicare-advantage-application",
         size=8, color="666666", bg="F9F9F9", italic=True, h_align="left")
    ws.row_dimensions[2].height = 18

    # headers
    for i, h in enumerate(["CMS Specialty", "Large Metro", "Metro", "Micro", "Rural", "CEAC"]):
        c = ws.cell(row=3, column=i+1)
        cell(ws, c.coordinate, h, bold=True, color=WHITE,
             bg=MID_BLUE, size=10, h_align="center", bdr=True)
    ws.row_dimensions[3].height = 20

    provider_td = [
        ("Primary Care",               "10 min / 5 mi",   "15 min / 10 mi",  "30 min / 20 mi",  "40 min / 30 mi",  "70 min / 60 mi"),
        ("Allergy and Immunology",     "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Cardiology",                 "20 min / 10 mi",  "30 min / 20 mi",  "50 min / 35 mi",  "75 min / 60 mi",  "95 min / 85 mi"),
        ("Chiropractor",               "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Clinical Psychology",        "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "145 min / 130 mi"),
        ("Clinical Social Work",       "20 min / 10 mi",  "30 min / 20 mi",  "50 min / 35 mi",  "75 min / 60 mi",  "125 min / 110 mi"),
        ("Dermatology",                "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Endocrinology",              "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("ENT/Otolaryngology",         "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Gastroenterology",           "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("General Surgery",            "20 min / 10 mi",  "30 min / 20 mi",  "50 min / 35 mi",  "75 min / 60 mi",  "95 min / 85 mi"),
        ("Gynecology OB/GYN",          "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Infectious Diseases",        "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Nephrology",                 "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Neurology",                  "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Neurosurgery",               "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Oncology Medical/Surgical",  "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Oncology Radiation",         "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Ophthalmology",              "20 min / 10 mi",  "30 min / 20 mi",  "50 min / 35 mi",  "75 min / 60 mi",  "95 min / 85 mi"),
        ("Orthopedic Surgery",         "20 min / 10 mi",  "30 min / 20 mi",  "50 min / 35 mi",  "75 min / 60 mi",  "95 min / 85 mi"),
        ("Physiatry Rehab Med",        "30 min / 15 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "90 min / 75 mi",  "125 min / 110 mi"),
        ("Plastic Surgery",            "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Podiatry",                   "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Psychiatry",                 "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Pulmonology",                "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Rheumatology",               "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Urology",                    "20 min / 10 mi",  "45 min / 30 mi",  "60 min / 45 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Vascular Surgery",           "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
        ("Cardiothoracic Surgery",     "30 min / 15 mi",  "60 min / 40 mi",  "100 min / 75 mi", "110 min / 90 mi", "145 min / 130 mi"),
    ]

    facility_td = [
        ("Acute Inpatient Hospitals",  "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Cardiac Surgery Program",    "30 min / 15 mi",  "60 min / 40 mi",  "160 min / 120 mi","145 min / 120 mi","155 min / 140 mi"),
        ("Cardiac Catheterization",    "30 min / 15 mi",  "60 min / 40 mi",  "160 min / 120 mi","145 min / 120 mi","155 min / 140 mi"),
        ("Critical Care ICU",          "20 min / 10 mi",  "45 min / 30 mi",  "160 min / 120 mi","145 min / 120 mi","155 min / 140 mi"),
        ("Surgical Services ASC",      "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Skilled Nursing Facility",   "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "95 min / 85 mi"),
        ("Diagnostic Radiology",       "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Mammography",                "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Physical Therapy",           "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Occupational Therapy",       "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Speech Therapy",             "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Inpatient Psychiatric",      "30 min / 15 mi",  "70 min / 45 mi",  "100 min / 75 mi", "90 min / 75 mi",  "155 min / 140 mi"),
        ("Outpatient Infusion/Chemo",  "20 min / 10 mi",  "45 min / 30 mi",  "80 min / 60 mi",  "75 min / 60 mi",  "110 min / 100 mi"),
        ("Outpatient Behavioral Health","20 min / 10 mi", "40 min / 25 mi",  "55 min / 40 mi",  "60 min / 50 mi",  "110 min / 100 mi"),
    ]

    r = 4
    for i, row_data in enumerate(provider_td):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data):
            cell(ws, ws.cell(row=r, column=j+1).coordinate,
                 val, size=9, bg=bg, h_align="center", bdr=True)
        ws.row_dimensions[r].height = 16
        r += 1

    # facility separator
    ws.merge_cells(f"A{r}:F{r}")
    cell(ws, f"A{r}",
         "  FACILITY SPECIALTY TYPES  |  Minimum = 1 per county (flat) per 422.116(e)(2)(iii)  |  "
         "EXCEPTION: Acute Inpatient Hospitals = CEIL(12.2 × beneficiaries_required / 1,000) BEDS",
         bold=True, color=WHITE, bg="375623", size=9, h_align="left")
    ws.row_dimensions[r].height = 20
    r += 1

    for i, row_data in enumerate(facility_td):
        bg = "E2EFDA" if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data):
            cell(ws, ws.cell(row=r, column=j+1).coordinate,
                 val, size=9, bg=bg, h_align="center", bdr=True)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    cell(ws, f"A{r}",
         "COMPLIANCE THRESHOLDS per 422.116(d)(4):  "
         "Large Metro + Metro → 90% of beneficiaries must have access  |  "
         "Micro + Rural + CEAC → 85%  |  "
         "BOTH access % AND provider count tests must pass for COMPLIANT status",
         bold=True, color=DARK_BLUE, bg=LIGHT_BLUE, size=10, h_align="left", wrap=True)
    ws.row_dimensions[r].height = 30

    return ws


# ── MAIN ─────────────────────────────────────────────────────
import pandas as pd
from google.cloud import bigquery

SUMMARY_COUNTY_QUERY = f"""
SELECT
  county_name,
  county_type,
  plan_type,
  COUNTIF(compliance_status = 'COMPLIANT')     AS compliant_specialties,
  COUNTIF(compliance_status = 'NON-COMPLIANT') AS non_compliant_specialties,
  COUNT(*)                                      AS total_specialties,
  ROUND(
    COUNTIF(compliance_status = 'COMPLIANT') / COUNT(*), 4
  )                                             AS pct_compliant,
  COUNTIF(access_compliant = FALSE)             AS access_failures,
  COUNTIF(count_compliant = FALSE)              AS count_failures
FROM `{PROJECT}.{DATASET}.{PREFIX}_fact_gap_analysis_v2`
GROUP BY county_name, county_type, plan_type
ORDER BY pct_compliant ASC, county_name, plan_type
"""

INVENTORY_QUERY = f"""
SELECT
  cms_specialty,
  plan_type,
  county_name,
  ma_contracted_providers,
  aetna_participating_providers,
  cms_medicare_providers
FROM `{PROJECT}.{DATASET}.{PREFIX}_week3_data_inventory`
ORDER BY county_name, cms_specialty, plan_type
"""

PAR_FLAG_QUERY = f"""
SELECT
  county_name,
  cms_specialty,
  plan_type,
  participation_status,
  COUNT(DISTINCT provider_id)            AS provider_count,
  SUM(claim_count)                       AS total_claims,
  ROUND(SUM(total_allowed_amt), 2)       AS total_allowed_amt,
  SUM(tot_benes)                         AS total_cms_benes,
  ROUND(AVG(NULLIF(tot_mdcr_pymt_amt, 0)), 2) AS avg_cms_payment
FROM `{PROJECT}.{DATASET}.{PREFIX}_provider_par_flag`
GROUP BY county_name, cms_specialty, plan_type, participation_status
ORDER BY county_name, cms_specialty, plan_type, participation_status
"""


# ── TAB 7: WEEK 3 DELIVERABLE 1 — DATA INVENTORY ─────────────

def build_tab7(wb, df_inventory):
    ws = wb.create_sheet("7. W3 Data Inventory")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {
        "A": 30, "B": 12, "C": 22,
        "D": 22, "E": 24, "F": 24,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # title
    ws.merge_cells("A1:F1")
    cell(ws, "A1", "Medicare Supply Demand — Week 3 Deliverable 1: Provider Data Inventory",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # subtitle
    ws.merge_cells("A2:F2")
    cell(ws, "A2",
         "Grain: CMS Specialty × Plan Type × County  |  "
         "Counts are correct at county level only — do NOT sum across counties",
         size=9, color="666666", bg="F9F9F9", italic=True, h_align="left")
    ws.row_dimensions[2].height = 18

    # callouts
    callouts = {
        "A3": "",
        "B3": "",
        "C3": "",
        "D3": "All providers in Aetna MA contracted network",
        "E3": "Providers with at least 1 claim (allowed_amt > 0) in 2024-2025",
        "F3": "Providers participating in Original Medicare (rndrng_prvdr_mdcr_prtcptg_ind = Y)",
    }
    for ref, txt in callouts.items():
        cell(ws, ref, txt, size=8, color="666666", bg="F9F9F9", italic=True, wrap=True)
    ws.row_dimensions[3].height = 28

    # headers
    headers = [
        ("A4", "CMS Specialty",                  DARK_GREY),
        ("B4", "Plan Type",                       DARK_GREY),
        ("C4", "County",                          DARK_GREY),
        ("D4", "MA Contracted\nProviders",         MID_BLUE),
        ("E4", "Aetna Participating\nProviders",   "375623"),
        ("F4", "CMS Medicare\nProviders",          "C55A11"),
    ]
    ws.row_dimensions[4].height = 35
    for ref, label, bg in headers:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=10, h_align="center", bdr=True)

    prev_specialty = None
    alt = True

    for i, (_, row) in enumerate(df_inventory.iterrows()):
        r = i + 5
        if row['cms_specialty'] != prev_specialty:
            alt = not alt
            prev_specialty = row['cms_specialty']
        row_bg = GREY if alt else WHITE

        data = [
            ("A", row.get('cms_specialty', ''),                    DARK_GREY, row_bg),
            ("B", row.get('plan_type', ''),                        DARK_GREY, row_bg),
            ("C", row.get('county_name', ''),                      DARK_GREY, row_bg),
            ("D", int(row.get('ma_contracted_providers', 0) or 0),  MID_BLUE,  LIGHT_BLUE),
            ("E", int(row.get('aetna_participating_providers', 0) or 0), "375623", "E2EFDA"),
            ("F", int(row.get('cms_medicare_providers', 0) or 0),   "C55A11",  "FCE4D6"),
        ]
        for col, val, txt_color, bg_color in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", color=txt_color, size=9)
            c.fill = fill(bg_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        ws.row_dimensions[r].height = 15

    return ws


# ── TAB 8: WEEK 3 DELIVERABLE 3 — PARTICIPATION FLAGS ─────────

STATUS_COLORS = {
    'ACTIVE BOTH':                                  ("375623", "E2EFDA"),
    'AETNA ACTIVE - NOT IN ORIGINAL MEDICARE':      ("7F6000", "FFF2CC"),
    'AETNA ACTIVE - NO NPI MATCH':                  ("7F6000", "FFFDE7"),
    'CONTRACTED NOT ACTIVE - IN ORIGINAL MEDICARE': ("C55A11", "FCE4D6"),
    'CONTRACTED NOT ACTIVE - NO CMS RECORD':        ("C00000", "FFE0E0"),
    'CONTRACTED NOT ACTIVE - NOT IN ORIGINAL MEDICARE': ("C00000", "FFE0E0"),
}

def build_tab8(wb, df_par):
    ws = wb.create_sheet("8. W3 Par Flags")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    col_widths = {
        "A": 22, "B": 30, "C": 12,
        "D": 44, "E": 14,
        "F": 14, "G": 18, "H": 16, "I": 18,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # title
    ws.merge_cells("A1:I1")
    cell(ws, "A1", "Medicare Supply Demand — Week 3 Deliverable 3: Provider Participation Flags",
         bold=True, color=WHITE, bg=DARK_BLUE, size=14, h_align="center")
    ws.row_dimensions[1].height = 35

    # subtitle
    ws.merge_cells("A2:I2")
    cell(ws, "A2",
         "Grain: County × CMS Specialty × Plan Type × Participation Status  |  "
         "Participation = Aetna claims activity (2024-2025) + CMS Original Medicare flag",
         size=9, color="666666", bg="F9F9F9", italic=True, h_align="left")
    ws.row_dimensions[2].height = 18

    # callouts
    callouts = {
        "A3": "", "B3": "", "C3": "", "D3": "",
        "E3": "COUNT(DISTINCT provider_id) per status",
        "F3": "SUM of claim_count for active providers",
        "G3": "SUM of Aetna allowed amounts 2024-2025",
        "H3": "SUM of CMS Medicare beneficiaries served (2023 FFS)",
        "I3": "AVG CMS payment per provider (active only)",
    }
    for ref, txt in callouts.items():
        cell(ws, ref, txt, size=8, color="666666", bg="F9F9F9", italic=True, wrap=True)
    ws.row_dimensions[3].height = 28

    # headers
    headers = [
        ("A4", "County",               DARK_GREY),
        ("B4", "CMS Specialty",        DARK_GREY),
        ("C4", "Plan Type",            DARK_GREY),
        ("D4", "Participation Status", DARK_GREY),
        ("E4", "Provider\nCount",      MID_BLUE),
        ("F4", "Total\nClaims",        MID_BLUE),
        ("G4", "Total Aetna\nAllowed", MID_BLUE),
        ("H4", "CMS Benes\nServed",    "C55A11"),
        ("I4", "Avg CMS\nPayment",     "C55A11"),
    ]
    ws.row_dimensions[4].height = 35
    for ref, label, bg in headers:
        cell(ws, ref, label, bold=True, color=WHITE,
             bg=bg, size=10, h_align="center", bdr=True)

    prev_key = None
    alt = True

    for i, (_, row) in enumerate(df_par.iterrows()):
        r = i + 5
        key = (row.get('county_name', ''), row.get('cms_specialty', ''))
        if key != prev_key:
            alt = not alt
            prev_key = key
        row_bg = GREY if alt else WHITE

        status = str(row.get('participation_status', ''))
        txt_c, status_bg = STATUS_COLORS.get(status, (DARK_GREY, row_bg))

        data = [
            ("A", row.get('county_name', ''),       DARK_GREY, row_bg),
            ("B", row.get('cms_specialty', ''),      DARK_GREY, row_bg),
            ("C", row.get('plan_type', ''),          DARK_GREY, row_bg),
            ("D", status,                            txt_c,     status_bg),
            ("E", int(row.get('provider_count', 0) or 0),       MID_BLUE, LIGHT_BLUE),
            ("F", int(row.get('total_claims', 0) or 0),         MID_BLUE, LIGHT_BLUE),
            ("G", float(row.get('total_allowed_amt', 0) or 0),  MID_BLUE, LIGHT_BLUE),
            ("H", int(row.get('total_cms_benes', 0) or 0),      "C55A11", "FCE4D6"),
            ("I", float(row.get('avg_cms_payment', 0) or 0),    "C55A11", "FCE4D6"),
        ]
        for col, val, txt_color, bg_color in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(name="Arial", color=txt_color, size=9,
                          bold=(col == "D"))
            c.fill = fill(bg_color)
            c.alignment = Alignment(horizontal="center" if col not in ("A","B","D") else "left",
                                    vertical="center")
            c.border = thin_border()
            if col in ("G", "I"):
                c.number_format = "#,##0.00"
        ws.row_dimensions[r].height = 15

    return ws

if __name__ == "__main__":
    client = bigquery.Client(project=CLIENT_PROJECT)

    print("Querying compliance data...")
    df = client.query(COMPLIANCE_QUERY).to_dataframe()
    print(f"  {len(df):,} rows")

    print("Querying specialty summary...")
    df_summary = client.query(SUMMARY_SPECIALTY_QUERY).to_dataframe()
    print(f"  {len(df_summary):,} rows")

    print("Querying county summary...")
    df_county = client.query(SUMMARY_COUNTY_QUERY).to_dataframe()
    print(f"  {len(df_county):,} rows")

    wb = Workbook()
    wb.remove(wb.active)

    print("Building Tab 1...")
    build_tab1(wb)

    print("Building Tab 2...")
    build_tab2(wb, df)

    print("Building Tab 3...")
    build_tab3(wb, df_summary)

    print("Building Tab 4...")
    build_tab4(wb, df_county)

    print("Building Tab 5...")
    build_tab5(wb)

    print("Building Tab 6...")
    build_tab6(wb)

    output = "medicare_supply_demand.xlsx"
    wb.save(output)
    print(f"Saved: {output}")
